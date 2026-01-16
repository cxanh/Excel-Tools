"""
删除页眉页脚插件 Worker 脚本
删除Excel文件的页眉页脚
"""
import openpyxl
import io

def process(file_content: bytes, options: dict) -> dict:
    """
    处理Excel文件，删除页眉页脚
    
    Args:
        file_content: Excel文件的二进制内容
        options: 处理选项
            - apply_to_all: 是否应用到所有工作表
    
    Returns:
        包含处理结果的字典
    """
    logs = []
    
    try:
        # 加载工作簿
        logs.append("正在加载Excel文件...")
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        apply_to_all = options.get('apply_to_all', True)
        sheets_processed = 0
        sheets_with_header_footer = 0
        
        # 处理工作表
        sheets_to_process = wb.worksheets if apply_to_all else [wb.active]
        
        for sheet in sheets_to_process:
            # 检查是否有页眉页脚
            has_header_footer = False
            if (sheet.oddHeader.left.text or sheet.oddHeader.center.text or sheet.oddHeader.right.text or
                sheet.oddFooter.left.text or sheet.oddFooter.center.text or sheet.oddFooter.right.text):
                has_header_footer = True
                sheets_with_header_footer += 1
            
            if has_header_footer:
                logs.append(f"正在处理工作表: {sheet.title}")
                
                # 清除页眉
                sheet.oddHeader.left.text = ""
                sheet.oddHeader.center.text = ""
                sheet.oddHeader.right.text = ""
                sheet.evenHeader.left.text = ""
                sheet.evenHeader.center.text = ""
                sheet.evenHeader.right.text = ""
                
                # 清除页脚
                sheet.oddFooter.left.text = ""
                sheet.oddFooter.center.text = ""
                sheet.oddFooter.right.text = ""
                sheet.evenFooter.left.text = ""
                sheet.evenFooter.center.text = ""
                sheet.evenFooter.right.text = ""
                
                sheets_processed += 1
                logs.append(f"✓ 工作表 {sheet.title} 页眉页脚已清除")
            else:
                logs.append(f"○ 工作表 {sheet.title} 无页眉页脚，跳过")
        
        if sheets_with_header_footer == 0:
            logs.append("⚠ 文件中没有页眉页脚，无需处理")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logs.append(f"✓ 处理完成！共清除 {sheets_processed} 个工作表的页眉页脚")
        
        return {
            'success': True,
            'buffer': output.getvalue(),
            'logs': logs,
            'statistics': {
                'totalSheets': len(wb.sheetnames),
                'sheetsWithHeaderFooter': sheets_with_header_footer,
                'sheetsProcessed': sheets_processed
            }
        }
        
    except Exception as e:
        logs.append(f"✗ 错误: {str(e)}")
        return {
            'success': False,
            'buffer': None,
            'logs': logs,
            'error': str(e)
        }

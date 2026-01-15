import io
import openpyxl
from PIL import Image as PILImage
import base64
from collections import defaultdict

def process(data):
    """从Excel文档中提取所有图片"""
    file_content = data['file']
    file_name = data['fileName']
    settings = data.get('settings', {})
    logs = []
    
    try:
        logs.append(f"正在加载文件: {file_name}")
        # 加载Excel文件
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        logs.append(f"文件加载成功，包含 {len(wb.sheetnames)} 个工作表")
        
        # 提取所有图片
        extracted_images = []
        image_counter = 1
        
        for sheet_name in wb.sheetnames:
            logs.append(f"正在处理工作表: {sheet_name}")
            ws = wb[sheet_name]
            
            # 提取工作表中的图片
            try:
                # 检查是否有图片
                if hasattr(ws, 'images'):
                    sheet_images = extract_images_from_sheet(ws, sheet_name, image_counter)
                    extracted_images.extend(sheet_images)
                    image_counter += len(sheet_images)
                    logs.append(f"  成功提取 {len(sheet_images)} 张图片")
                else:
                    logs.append(f"  未发现图片")
            except Exception as e:
                logs.append(f"  处理图片时出错: {str(e)}")
        
        if not extracted_images:
            logs.append("未提取到任何图片")
            return {
                'success': False,
                'error': '未提取到任何图片',
                'logs': logs
            }
        
        logs.append(f"总计提取 {len(extracted_images)} 张图片")
        
        return {
            'success': True,
            'logs': logs,
            'results': extracted_images,
            'details': {
                'imageCount': len(extracted_images)
            }
        }
        
    except Exception as e:
        logs.append(f"提取图片失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'logs': logs
        }


def extract_images_from_sheet(ws, sheet_name, start_index):
    """从单个工作表中提取图片"""
    extracted_images = []
    
    # 遍历工作表中的所有图片
    for img_index, img in enumerate(ws.images):
        try:
            # 图片的元数据
            img_info = {
                'index': start_index + img_index,
                'sheet_name': sheet_name,
                'original_filename': getattr(img, 'filename', f'image_{start_index + img_index}'),
                'width': img.width,
                'height': img.height,
                'anchor': str(img.anchor)
            }
            
            # 获取图片数据
            if hasattr(img, 'image'):
                # 对于较新版本的openpyxl
                image_data = img.image
            elif hasattr(img, '_data'):
                # 对于旧版本的openpyxl
                image_data = img._data
            else:
                continue
            
            # 处理不同类型的图片数据
            if isinstance(image_data, bytes):
                # 直接是字节数据
                img_bytes = image_data
            elif hasattr(image_data, 'data'):
                # 有data属性
                img_bytes = image_data.data
            elif hasattr(image_data, 'tobytes'):
                # 有tobytes方法
                img_bytes = image_data.tobytes()
            else:
                continue
            
            # 确定图片格式
            img_format = 'png'  # 默认格式
            try:
                # 尝试使用PIL检测图片格式
                pil_img = PILImage.open(io.BytesIO(img_bytes))
                img_format = pil_img.format.lower() or 'png'
                
                # 确保图片格式为常见格式
                if img_format not in ['jpg', 'jpeg', 'png', 'gif', 'bmp']:
                    img_format = 'png'
                    # 转换为PNG格式
                    png_buffer = io.BytesIO()
                    pil_img.save(png_buffer, format='PNG')
                    img_bytes = png_buffer.getvalue()
            except Exception as e:
                # 如果检测失败，使用默认PNG格式
                img_format = 'png'
            
            # 生成文件名
            base_filename = ws.title.lower().replace(' ', '_')
            file_name = f"{base_filename}_{img_info['index']}.{img_format}"
            
            img_info['file_name'] = file_name
            img_info['format'] = img_format
            img_info['buffer'] = img_bytes
            
            extracted_images.append(img_info)
            
        except Exception as e:
            continue
    
    return extracted_images

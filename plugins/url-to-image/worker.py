# 图片地址转图片插件 - Python处理脚本
import time
import io
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter, column_index_from_string
from PIL import Image as PILImage
import requests

def process(data):
    """
    将Excel中的图片URL转换为实际的嵌入图片
    
    参数:
        data: 包含file(bytes)、params的字典
        params包含:
            - urlColumn: URL所在列 (如'A')
            - insertPosition: 插入位置 ('same', 'next', 'custom')
            - targetColumn: 目标列 (当insertPosition='custom'时)
            - imageWidth: 图片宽度
            - imageHeight: 图片高度
            - clearUrl: 是否清空URL单元格
            - skipErrors: 是否跳过错误继续处理
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        url_column = params.get('urlColumn', 'A')
        insert_position = params.get('insertPosition', 'next')
        target_column = params.get('targetColumn', '')
        image_width = params.get('imageWidth', 100)
        image_height = params.get('imageHeight', 100)
        clear_url = params.get('clearUrl', False)
        skip_errors = params.get('skipErrors', True)
        
        logs.append(f"配置: URL列={url_column}, 插入位置={insert_position}, 图片尺寸={image_width}x{image_height}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        # 计算目标列
        url_col_idx = column_index_from_string(url_column)
        if insert_position == 'same':
            target_col_idx = url_col_idx
        elif insert_position == 'next':
            target_col_idx = url_col_idx + 1
        else:  # custom
            target_col_idx = column_index_from_string(target_column) if target_column else url_col_idx + 1
        
        target_col_letter = get_column_letter(target_col_idx)
        logs.append(f"目标列: {target_col_letter}")
        
        total_images_converted = 0
        failed_urls = 0
        sheets_processed = 0
        
        # URL验证正则
        url_pattern = re.compile(r'^https?://.+\.(jpg|jpeg|png|gif|bmp|webp)(\?.*)?$', re.IGNORECASE)
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            sheet_converted = 0
            sheet_failed = 0
            
            # 遍历URL列的所有行
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=url_col_idx)
                url = cell.value
                
                # 跳过空单元格
                if not url or not isinstance(url, str):
                    continue
                
                url = url.strip()
                
                # 验证URL格式
                if not url_pattern.match(url):
                    logs.append(f"  行{row}: 跳过无效URL - {url[:50]}")
                    if not skip_errors:
                        raise ValueError(f"无效的图片URL: {url}")
                    sheet_failed += 1
                    continue
                
                try:
                    # 下载图片
                    logs.append(f"  行{row}: 下载图片 - {url[:50]}...")
                    response = requests.get(url, timeout=10, headers={
                        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
                    })
                    response.raise_for_status()
                    
                    # 加载图片
                    img_data = io.BytesIO(response.content)
                    pil_img = PILImage.open(img_data)
                    
                    # 转换为PNG格式（确保兼容性）
                    img_stream = io.BytesIO()
                    pil_img.save(img_stream, format='PNG')
                    img_stream.seek(0)
                    
                    # 创建openpyxl图片对象
                    excel_img = Image(img_stream)
                    excel_img.width = image_width
                    excel_img.height = image_height
                    
                    # 设置图片位置（锚定到目标单元格）
                    excel_img.anchor = f"{target_col_letter}{row}"
                    
                    # 添加图片到工作表
                    ws.add_image(excel_img)
                    
                    # 清空URL单元格（如果需要）
                    if clear_url:
                        cell.value = None
                    
                    sheet_converted += 1
                    logs.append(f"  行{row}: 转换成功")
                    
                except Exception as e:
                    error_msg = str(e)
                    logs.append(f"  行{row}: 转换失败 - {error_msg}")
                    if not skip_errors:
                        raise
                    sheet_failed += 1
            
            if sheet_converted > 0 or sheet_failed > 0:
                logs.append(f"  工作表统计: 成功{sheet_converted}个, 失败{sheet_failed}个")
                sheets_processed += 1
            
            total_images_converted += sheet_converted
            failed_urls += sheet_failed
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！共转换 {total_images_converted} 个图片")
        if failed_urls > 0:
            logs.append(f"失败 {failed_urls} 个URL")
        logs.append(f"处理了 {sheets_processed} 个工作表")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'imagesConverted': total_images_converted,
                    'failedUrls': failed_urls,
                    'sheetsProcessed': sheets_processed,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

# 删除空白行插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def process(data):
    """
    处理Excel文件，删除所有空白行
    
    参数:
        data: 包含file(bytes)和params(dict)的字典
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        total_deleted = 0
        total_rows_before = 0
        total_rows_after = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 获取最大行数和列数
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 0:
                logs.append(f"  工作表为空，跳过")
                continue
            
            total_rows_before += max_row
            
            # 从后向前遍历，避免删除行后索引变化
            empty_rows = []
            for row_idx in range(max_row, 0, -1):
                # 检查该行是否为空
                is_empty = True
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None and str(cell.value).strip() != '':
                        is_empty = False
                        break
                
                if is_empty:
                    empty_rows.append(row_idx)
            
            # 删除空白行
            deleted_count = 0
            for row_idx in empty_rows:
                ws.delete_rows(row_idx, 1)
                deleted_count += 1
            
            total_deleted += deleted_count
            total_rows_after += (max_row - deleted_count)
            
            logs.append(f"  原始行数: {max_row}, 删除: {deleted_count}, 剩余: {max_row - deleted_count}")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！总共删除 {total_deleted} 行空白行")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'totalRows': total_rows_before,
                    'deletedRows': total_deleted,
                    'remainingRows': total_rows_after,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

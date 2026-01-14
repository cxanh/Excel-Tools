import openpyxl
import io
import re
import json
from collections import defaultdict

def process(data):
    # 解析输入数据
    file_content = data['file']
    file_name = data['fileName']
    settings = data['settings']
    
    logs = []
    deleted_count = 0
    total_kept_count = 0
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        logs.append(f"成功加载文件: {file_name}")
        logs.append(f"包含工作表: {workbook.sheetnames}")
        
        # 获取需要处理的工作表
        sheets_to_process = []
        if settings.get('sheetOption', 'all') == 'all':
            sheets_to_process = workbook.sheetnames
            logs.append("将处理所有工作表")
        else:
            # 解析指定工作表
            specific_sheets = settings.get('specificSheets', '')
            if specific_sheets:
                # 分割工作表名称
                sheet_patterns = [s.strip() for s in specific_sheets.split(',') if s.strip()]
                
                # 匹配工作表名称
                for sheet_name in workbook.sheetnames:
                    for pattern in sheet_patterns:
                        # 转换为正则表达式模式
                        regex_pattern = pattern.replace('*', '.*').replace('?', '.')
                        if re.match(regex_pattern, sheet_name, re.IGNORECASE):
                            sheets_to_process.append(sheet_name)
                            break
                
                if sheets_to_process:
                    logs.append(f"将处理工作表: {sheets_to_process}")
                else:
                    logs.append(f"没有匹配的工作表: {specific_sheets}")
                    return {
                        'success': False,
                        'error': '没有匹配的工作表',
                        'logs': logs,
                        'deletedCount': deleted_count
                    }
            else:
                logs.append("没有指定工作表")
                return {
                    'success': False,
                    'error': '没有指定工作表',
                    'logs': logs,
                    'deletedCount': deleted_count
                }
        
        # 获取重复行检测条件
        column_option = settings.get('columnOption', 'all')
        specific_columns = settings.get('specificColumns', '')
        
        # 获取重复行处理选项
        keep_option = settings.get('keepOption', 'first')
        
        # 获取高级选项
        advanced_options = settings.get('advancedOptions', ['ignoreBlank', 'ignoreEmptyRows'])
        case_sensitive = 'caseSensitive' in advanced_options
        ignore_blank = 'ignoreBlank' in advanced_options
        ignore_empty_rows = 'ignoreEmptyRows' in advanced_options
        
        logs.append(f"重复行检测: {column_option}{f' ({specific_columns})' if column_option == 'specific' else ''}")
        logs.append(f"重复行处理: {'保留第一个' if keep_option == 'first' else '保留最后一个'}")
        logs.append(f"高级选项: {'区分大小写' if case_sensitive else '不区分大小写'}, {'忽略空白单元格' if ignore_blank else '不忽略空白单元格'}, {'忽略空行' if ignore_empty_rows else '不忽略空行'}")
        
        # 处理每个工作表
        for sheet_name in sheets_to_process:
            sheet = workbook[sheet_name]
            sheet_deleted = 0
            sheet_total = 0
            
            logs.append(f"开始处理工作表: {sheet_name}")
            
            # 获取工作表的行列信息
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row < 1:
                logs.append(f"工作表 {sheet_name} 没有数据行")
                continue
            
            # 解析指定的列
            target_columns = []
            if column_option == 'specific' and specific_columns:
                # 支持列字母(A,B,C)或列名称
                column_patterns = [s.strip() for s in specific_columns.split(',') if s.strip()]
                
                # 获取表头信息
                headers = []
                if max_row > 0:
                    headers = [sheet.cell(row=1, column=col).value for col in range(1, max_col + 1)]
                    headers = [str(h) if h is not None else f"column_{col+1}" for col, h in enumerate(headers)]
                
                # 解析列信息
                for pattern in column_patterns:
                    # 检查是否是列字母
                    if re.match(r'^[A-Za-z]+$', pattern):
                        # 转换为列索引
                        col_idx = 0
                        for c in pattern.upper():
                            col_idx = col_idx * 26 + (ord(c) - ord('A') + 1)
                        target_columns.append(col_idx)
                    else:
                        # 按列名称匹配
                        for idx, header in enumerate(headers):
                            if header and (header == pattern or (not case_sensitive and header.lower() == pattern.lower())):
                                target_columns.append(idx + 1)  # 转换为1-based索引
                                break
                
                # 去重并排序
                target_columns = sorted(list(set(target_columns)))
                
                if not target_columns:
                    logs.append(f"没有找到匹配的列: {specific_columns}")
                    target_columns = list(range(1, max_col + 1))  # 默认使用所有列
            else:
                # 使用所有列
                target_columns = list(range(1, max_col + 1))
            
            logs.append(f"使用列: {target_columns}")
            
            # 收集所有数据行
            rows_data = []
            for row_idx in range(1, max_row + 1):
                row_data = []
                is_empty = True
                
                for col_idx in target_columns:
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell_value = cell.value
                    
                    # 处理单元格值
                    if cell_value is not None:
                        is_empty = False
                        if isinstance(cell_value, str):
                            if not case_sensitive:
                                cell_value = cell_value.lower()
                    elif ignore_blank:
                        # 忽略空白单元格
                        cell_value = ""
                    
                    row_data.append(cell_value)
                
                # 检查是否忽略空行
                if ignore_empty_rows and is_empty:
                    continue
                
                rows_data.append((row_idx, tuple(row_data)))
                sheet_total += 1
            
            # 检测重复行
            seen_rows = defaultdict(list)
            for row_idx, row_tuple in rows_data:
                seen_rows[row_tuple].append(row_idx)
            
            # 标记需要删除的行
            rows_to_delete = set()
            for row_tuple, row_indices in seen_rows.items():
                if len(row_indices) > 1:
                    # 有重复行
                    if keep_option == 'first':
                        # 保留第一个，删除其他
                        rows_to_delete.update(row_indices[1:])
                        sheet_deleted += len(row_indices) - 1
                    else:
                        # 保留最后一个，删除其他
                        rows_to_delete.update(row_indices[:-1])
                        sheet_deleted += len(row_indices) - 1
            
            # 按降序删除行（避免索引问题）
            for row_idx in sorted(rows_to_delete, reverse=True):
                sheet.delete_rows(row_idx)
            
            # 更新计数
            deleted_count += sheet_deleted
            sheet_kept = sheet_total - sheet_deleted
            total_kept_count += sheet_kept
            
            logs.append(f"工作表 {sheet_name} 处理完成: 删除了 {sheet_deleted} 个重复行, 保留了 {sheet_kept} 行数据")
        
        # 保存处理后的文件
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        logs.append(f"文件 {file_name} 处理完成")
        logs.append(f"总计删除: {deleted_count} 个重复行")
        logs.append(f"总计保留: {total_kept_count} 行数据")
        
        return {
            'success': True,
            'buffer': output_buffer.getvalue(),
            'deletedCount': deleted_count,
            'keptCount': total_kept_count,
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"处理失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'deletedCount': deleted_count,
            'logs': logs
        }

# 测试函数（用于开发和调试）
def test():
    # 创建一个简单的测试数据
    test_data = {
        'file': open('test.xlsx', 'rb').read(),
        'fileName': 'test.xlsx',
        'settings': {
            'sheetOption': 'all',
            'specificSheets': '',
            'columnOption': 'all',
            'specificColumns': '',
            'keepOption': 'first',
            'advancedOptions': ['ignoreBlank', 'ignoreEmptyRows']
        }
    }
    
    result = process(test_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    if result['success']:
        with open('处理后_test.xlsx', 'wb') as f:
            f.write(result['buffer'])
        print("处理后的文件已保存为: 处理后_test.xlsx")

if __name__ == '__main__':
    test()

import io
import openpyxl
import csv
import json
import html


def process(data):
    """将Excel文件转换为其他格式"""
    file_content = data['file']
    file_name = data['fileName']
    settings = data.get('settings', {})
    target_format = settings.get('targetFormat', 'csv')
    logs = []
    
    try:
        logs.append(f"正在加载文件: {file_name}")
        # 加载Excel文件
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=True)
        logs.append(f"文件加载成功，包含 {len(wb.sheetnames)} 个工作表")
        
        # 选择要转换的工作表
        sheet_name = settings.get('sheetName', wb.active.title)
        if sheet_name not in wb.sheetnames:
            sheet_name = wb.active.title
        ws = wb[sheet_name]
        logs.append(f"正在转换工作表: {sheet_name}")
        
        # 转换为目标格式
        result_buffer = io.BytesIO()
        
        if target_format.lower() == 'csv':
            # 转换为CSV
            logs.append("正在转换为CSV格式")
            # 创建CSV写入器
            writer = csv.writer(io.TextIOWrapper(result_buffer, encoding='utf-8', newline=''))
            
            # 写入表头
            header = [cell.value for cell in ws[1]]
            writer.writerow(header)
            
            # 写入数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                writer.writerow(row)
            
            logs.append("CSV转换完成")
            result_buffer.seek(0)
            return {
                'success': True,
                'buffer': result_buffer.read(),
                'logs': logs,
                'details': {
                    'targetFormat': 'csv',
                    'sheetName': sheet_name,
                    'rowCount': ws.max_row,
                    'columnCount': ws.max_column
                }
            }
        
        elif target_format.lower() == 'json':
            # 转换为JSON
            logs.append("正在转换为JSON格式")
            
            # 读取数据
            data = []
            header = [cell.value for cell in ws[1]]
            
            for row in ws.iter_rows(min_row=2, values_only=True):
                row_data = {}
                for i, value in enumerate(row):
                    if i < len(header):
                        row_data[header[i]] = value
                data.append(row_data)
            
            # 写入JSON
            json.dump(data, io.TextIOWrapper(result_buffer, encoding='utf-8'), ensure_ascii=False, indent=2)
            
            logs.append("JSON转换完成")
            result_buffer.seek(0)
            return {
                'success': True,
                'buffer': result_buffer.read(),
                'logs': logs,
                'details': {
                    'targetFormat': 'json',
                    'sheetName': sheet_name,
                    'rowCount': len(data),
                    'columnCount': len(header)
                }
            }
        
        elif target_format.lower() == 'html':
            # 转换为HTML
            logs.append("正在转换为HTML格式")
            
            # 创建HTML内容
            html_content = ['<html>', '<head><title>Excel转换结果</title>']
            html_content.append('<style>')
            html_content.append('table { border-collapse: collapse; width: 100%; }')
            html_content.append('th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }')
            html_content.append('th { background-color: #f2f2f2; }')
            html_content.append('tr:nth-child(even) { background-color: #f9f9f9; }')
            html_content.append('</style>')
            html_content.append('</head>')
            html_content.append('<body>')
            html_content.append(f'<h1>{sheet_name}</h1>')
            html_content.append('<table>')
            
            # 写入表头
            html_content.append('<tr>')
            for cell in ws[1]:
                html_content.append(f'<th>{html.escape(str(cell.value) if cell.value else "")}</th>')
            html_content.append('</tr>')
            
            # 写入数据
            for row in ws.iter_rows(min_row=2, values_only=True):
                html_content.append('<tr>')
                for value in row:
                    html_content.append(f'<td>{html.escape(str(value) if value else "")}</td>')
                html_content.append('</tr>')
            
            html_content.append('</table>')
            html_content.append('</body>')
            html_content.append('</html>')
            
            # 写入HTML到缓冲区
            result_buffer.write('\n'.join(html_content).encode('utf-8'))
            
            logs.append("HTML转换完成")
            result_buffer.seek(0)
            return {
                'success': True,
                'buffer': result_buffer.read(),
                'logs': logs,
                'details': {
                    'targetFormat': 'html',
                    'sheetName': sheet_name,
                    'rowCount': ws.max_row,
                    'columnCount': ws.max_column
                }
            }
        
        else:
            # 不支持的格式
            logs.append(f"不支持的目标格式: {target_format}")
            return {
                'success': False,
                'error': f"不支持的目标格式: {target_format}",
                'logs': logs
            }
        
    except Exception as e:
        logs.append(f"转换失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'logs': logs
        }
    finally:
        result_buffer.close()

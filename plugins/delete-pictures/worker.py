import openpyxl
import io
import re
import json
from openpyxl.drawing.image import Image as OpenpyxlImage

def process(data):
    # 解析输入数据
    file_content = data['file']
    file_name = data['fileName']
    settings = data['settings']
    
    logs = []
    deleted_count = 0
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(io.BytesIO(file_content))
        logs.append(f"成功加载文件: {file_name}")
        logs.append(f"包含工作表: {workbook.sheetnames}")
        
        # 获取需要处理的工作表
        sheets_to_process = []
        if settings.get('sheetOption', 'all') == 'all':
            sheets_to_process = workbook.sheetnames
            logs.append("将处理所有工作表")
        else:
            # 解析指定工作表
            specific_sheets = settings.get('specificSheets', '')
            if specific_sheets:
                # 分割工作表名称
                sheet_patterns = [s.strip() for s in specific_sheets.split(',') if s.strip()]
                
                # 匹配工作表名称
                for sheet_name in workbook.sheetnames:
                    for pattern in sheet_patterns:
                        # 转换为正则表达式模式
                        regex_pattern = pattern.replace('*', '.*').replace('?', '.')
                        if re.match(regex_pattern, sheet_name, re.IGNORECASE):
                            sheets_to_process.append(sheet_name)
                            break
                
                if sheets_to_process:
                    logs.append(f"将处理工作表: {sheets_to_process}")
                else:
                    logs.append(f"没有匹配的工作表: {specific_sheets}")
                    return {
                        'success': False,
                        'error': '没有匹配的工作表',
                        'logs': logs,
                        'deletedCount': deleted_count
                    }
            else:
                logs.append("没有指定工作表")
                return {
                    'success': False,
                    'error': '没有指定工作表',
                    'logs': logs,
                    'deletedCount': deleted_count
                }
        
        # 获取删除范围
        delete_range = settings.get('deleteRange', ['shapes'])
        delete_shapes = 'shapes' in delete_range
        delete_comments = 'comments' in delete_range
        
        logs.append(f"删除范围: {'形状' if delete_shapes else ''}{', 批注' if delete_comments else ''}")
        
        # 获取保留条件
        keep_conditions = settings.get('keepConditions', [])
        keep_small = 'small' in keep_conditions
        keep_large = 'large' in keep_conditions
        
        # 获取尺寸阈值
        size_threshold = settings.get('sizeThreshold', {'width': 100, 'height': 100})
        threshold_width = size_threshold.get('width', 100)
        threshold_height = size_threshold.get('height', 100)
        
        if keep_conditions:
            logs.append(f"保留条件: {'小图片' if keep_small else ''}{', 大图片' if keep_large else ''}")
            logs.append(f"尺寸阈值: {threshold_width}x{threshold_height} 像素")
        
        # 处理每个工作表
        for sheet_name in sheets_to_process:
            sheet = workbook[sheet_name]
            sheet_deleted = 0
            
            logs.append(f"开始处理工作表: {sheet_name}")
            
            # 处理形状（图片、图表等）
            if delete_shapes:
                # 获取所有形状
                shapes = list(sheet._images) if hasattr(sheet, '_images') else []
                
                # 删除需要移除的形状
                for shape in shapes:
                    delete_shape = True
                    
                    # 检查是否需要根据尺寸保留
                    if keep_conditions and hasattr(shape, 'width') and hasattr(shape, 'height'):
                        # 转换为像素尺寸
                        width_pixels = shape.width * 3.7795275591  # 转换为像素
                        height_pixels = shape.height * 3.7795275591  # 转换为像素
                        
                        # 检查尺寸条件
                        if keep_small and width_pixels < threshold_width and height_pixels < threshold_height:
                            delete_shape = False  # 保留小图片
                        elif keep_large and width_pixels > threshold_width and height_pixels > threshold_height:
                            delete_shape = False  # 保留大图片
                    
                    if delete_shape:
                        try:
                            if hasattr(sheet, '_images'):
                                sheet._images.remove(shape)
                            sheet_deleted += 1
                        except Exception as e:
                            logs.append(f"删除形状失败: {str(e)}")
            
            # 处理批注图片
            if delete_comments:
                for row in sheet.iter_rows():
                    for cell in row:
                        if cell.comment:
                            # 检查批注是否包含图片
                            if hasattr(cell.comment, '_richText'):
                                for run in cell.comment._richText:
                                    if hasattr(run, '_r'):
                                        r_tag = run._r
                                        if hasattr(r_tag, 'pic'):
                                            # 删除批注
                                            cell.comment = None
                                            sheet_deleted += 1
                                            break
            
            if sheet_deleted > 0:
                logs.append(f"工作表 {sheet_name} 删除了 {sheet_deleted} 个图片")
                deleted_count += sheet_deleted
            else:
                logs.append(f"工作表 {sheet_name} 没有找到需要删除的图片")
        
        # 保存处理后的文件
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        logs.append(f"文件 {file_name} 处理完成，共删除 {deleted_count} 个图片")
        
        return {
            'success': True,
            'buffer': output_buffer.getvalue(),
            'deletedCount': deleted_count,
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"处理失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'deletedCount': deleted_count,
            'logs': logs
        }

# 测试函数（用于开发和调试）
def test():
    # 创建一个简单的测试数据
    test_data = {
        'file': open('test.xlsx', 'rb').read(),
        'fileName': 'test.xlsx',
        'settings': {
            'sheetOption': 'all',
            'specificSheets': '',
            'deleteRange': ['shapes', 'comments'],
            'keepConditions': [],
            'sizeThreshold': {
                'width': 100,
                'height': 100
            }
        }
    }
    
    result = process(test_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    if result['success']:
        with open('处理后_test.xlsx', 'wb') as f:
            f.write(result['buffer'])
        print("处理后的文件已保存为: 处理后_test.xlsx")

if __name__ == '__main__':
    test()

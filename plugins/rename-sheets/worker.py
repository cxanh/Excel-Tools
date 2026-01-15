#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
重命名 Excel 工作表插件
"""

import io
import json
import re
from openpyxl import load_workbook


def process(input_data):
    """处理工作表重命名请求
    
    Args:
        input_data: 包含输入文件和设置的数据
    
    Returns:
        dict: 处理结果，包含处理后的文件buffer和日志
    """
    result = {
        "success": False,
        "logs": [],
        "buffer": None
    }
    
    try:
        # 解析输入数据
        file = input_data["file"]
        file_name = input_data["fileName"]
        settings = input_data.get("settings", {})
        
        rename_rules = settings.get("renameRules", [])
        
        # 加载Excel文件
        result["logs"].append(f"正在加载文件: {file_name}")
        wb = load_workbook(io.BytesIO(file))
        
        # 遍历工作表重命名
        original_sheet_names = wb.sheetnames.copy()
        result["logs"].append(f"文件包含 {len(original_sheet_names)} 个工作表")
        
        renamed_count = 0
        
        for sheet_name in original_sheet_names:
            ws = wb[sheet_name]
            new_name = None
            
            # 查找匹配的重命名规则
            for rule in rename_rules:
                if rule.get("enabled", False):
                    match_type = rule.get("matchType", "exact")
                    old_name_pattern = rule.get("oldName", "")
                    new_name_format = rule.get("newName", "")
                    
                    if match_type == "exact" and sheet_name == old_name_pattern:
                        new_name = new_name_format
                        break
                    elif match_type == "contains" and old_name_pattern in sheet_name:
                        new_name = new_name_format
                        break
                    elif match_type == "regex":
                        try:
                            if re.match(old_name_pattern, sheet_name):
                                new_name = re.sub(old_name_pattern, new_name_format, sheet_name)
                                break
                        except re.error:
                            result["logs"].append(f"  正则表达式错误: {old_name_pattern}")
                    elif match_type == "prefix" and sheet_name.startswith(old_name_pattern):
                        new_name = new_name_format + sheet_name[len(old_name_pattern):]
                        break
                    elif match_type == "suffix" and sheet_name.endswith(old_name_pattern):
                        new_name = sheet_name[:-len(old_name_pattern)] + new_name_format
                        break
            
            # 如果找到匹配的规则，执行重命名
            if new_name and new_name != sheet_name:
                # 确保新名称不重复
                counter = 1
                original_new_name = new_name
                while new_name in wb.sheetnames:
                    new_name = f"{original_new_name}_{counter}"
                    counter += 1
                
                result["logs"].append(f"  将工作表 '{sheet_name}' 重命名为 '{new_name}'")
                ws.title = new_name
                renamed_count += 1
            else:
                result["logs"].append(f"  工作表 '{sheet_name}' 未找到匹配的重命名规则或无需重命名")
        
        # 保存处理后的文件
        result["logs"].append(f"共重命名了 {renamed_count} 个工作表")
        result["logs"].append("保存处理后的文件")
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        result["buffer"] = output_stream.getvalue()
        result["success"] = True
        result["logs"].append("文件处理完成")
        
    except Exception as e:
        result["success"] = False
        result["logs"].append(f"处理错误: {str(e)}")
        result["error"] = str(e)
    
    return result

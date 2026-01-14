import io
import openpyxl
from PIL import Image
from openpyxl.drawing.image import Image as ExcelImage
import base64
import sys
import json

# 默认设置
DEFAULT_SETTINGS = {
    "replaceMode": "all",
    "imageSize": "original",
    "customWidth": 0,
    "customHeight": 0,
    "imagePosition": "original",
    "sheetOption": "all",
    "selectedSheets": "",
    "ignoreHiddenSheets": True,
    "enableDetailedLogs": True
}

# 处理图片替换
def process(data):
    # 解析输入数据
    excel_file_bytes = data['file']
    replace_image_bytes = data['settings']['replaceImageBytes']
    settings = data.get('settings', None)
    
    logs = []
    
    # 使用默认设置或传入的设置
    current_settings = DEFAULT_SETTINGS.copy()
    if settings:
        current_settings.update(settings)
    
    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(io.BytesIO(excel_file_bytes))
        logs.append(f"成功加载Excel文件")
        logs.append(f"包含 {len(wb.sheetnames)} 个工作表")
        
        if current_settings['enableDetailedLogs']:
            logs.append(f"工作表列表: {', '.join(wb.sheetnames)}")
        
        # 加载替换图片
        replace_image = Image.open(io.BytesIO(replace_image_bytes))
        logs.append(f"成功加载替换图片，尺寸: {replace_image.size}")
        
        # 确定要处理的工作表
        sheets_to_process = []
        
        if current_settings['sheetOption'] == 'all':
            sheets_to_process = wb.sheetnames
            logs.append("将处理所有工作表")
        elif current_settings['sheetOption'] == 'active':
            sheets_to_process = [wb.active.title]
            logs.append(f"将处理活动工作表: {wb.active.title}")
        elif current_settings['sheetOption'] == 'selected':
            selected_sheets = [name.strip() for name in current_settings['selectedSheets'].split(',')]
            sheets_to_process = [sheet for sheet in selected_sheets if sheet in wb.sheetnames]
            if not sheets_to_process:
                logs.append(f"警告: 未找到指定的工作表，将处理所有工作表")
                sheets_to_process = wb.sheetnames
            else:
                logs.append(f"将处理工作表: {sheets_to_process}")
        
        total_replaced = 0
        
        # 遍历需要处理的工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            
            # 检查是否需要忽略隐藏工作表
            if current_settings['ignoreHiddenSheets'] and not ws.sheet_state == "visible":
                logs.append(f"忽略隐藏工作表: {sheet_name}")
                continue
                
            logs.append(f"处理工作表: {sheet_name}")
            
            # 获取所有图片
            images = list(ws._images)
            img_count = len(images)
            
            if img_count > 0:
                logs.append(f"  发现 {img_count} 张图片，开始替换...")
                
                if current_settings['enableDetailedLogs']:
                    logs.append(f"  替换模式: {current_settings['replaceMode']}")
                    logs.append(f"  图片尺寸: {current_settings['imageSize']}")
                    if current_settings['imageSize'] == "custom":
                        logs.append(f"  自定义尺寸: {current_settings['customWidth']}x{current_settings['customHeight']}")
                    logs.append(f"  图片位置: {current_settings['imagePosition']}")
                
                # 根据替换模式确定要替换的图片数量
                if current_settings['replaceMode'] == "first" and img_count > 0:
                    images_to_process = images[:1]
                    logs.append(f"  只替换第一张图片")
                else:
                    images_to_process = images
                    logs.append(f"  替换所有图片")
                
                replaced_in_sheet = 0
                
                # 处理每张图片
                for img in images_to_process:
                    try:
                        # 保存原图片位置和尺寸
                        original_anchor = img.anchor
                        
                        # 根据设置调整替换图片尺寸
                        processed_image = replace_image.copy()
                        
                        if current_settings['imageSize'] == "original":
                            # 保持原图片尺寸
                            if hasattr(img, 'width') and hasattr(img, 'height'):
                                processed_image = processed_image.resize((img.width, img.height))
                        elif current_settings['imageSize'] == "custom":
                            # 自定义尺寸
                            if current_settings['customWidth'] > 0 and current_settings['customHeight'] > 0:
                                processed_image = processed_image.resize((current_settings['customWidth'], current_settings['customHeight']))
                        # elif current_settings['imageSize'] == "fit":
                        #     # 自适应填充（需要更复杂的逻辑）
                        #     # 这里暂时保持原尺寸
                        #     pass
                        
                        # 将PIL图片转换为ExcelImage
                        output = io.BytesIO()
                        processed_image.save(output, format='PNG')
                        output.seek(0)
                        excel_image = ExcelImage(output)
                        
                        # 设置图片位置
                        if current_settings['imagePosition'] == "original" and original_anchor:
                            # 保持原位置
                            excel_image.anchor = original_anchor
                        elif current_settings['imagePosition'] == "center":
                            # 居中放置（需要更复杂的定位逻辑）
                            # 这里暂时保持原位置
                            if original_anchor:
                                excel_image.anchor = original_anchor
                        
                        # 移除原图片
                        ws._images.remove(img)
                        
                        # 添加新图片
                        ws.add_image(excel_image)
                        
                        replaced_in_sheet += 1
                        total_replaced += 1
                        
                    except Exception as e:
                        logs.append(f"  替换图片时出错: {str(e)}")
                
                logs.append(f"  工作表 {sheet_name} 处理完成，替换了 {replaced_in_sheet} 张图片")
            else:
                logs.append(f"  未发现图片")
        
        logs.append(f"所有工作表处理完成，共替换 {total_replaced} 张图片")
        
        # 将处理结果写入内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'success': True,
            'buffer': output.read(),
            'logs': logs,
            'replacedCount': total_replaced
        }
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logs.append(f"错误: {str(e)}")
        logs.append(f"详细错误: {error_trace}")
        return {
            'success': False,
            'error': str(e),
            'logs': logs
        }

# 为了保持向后兼容性，保留原函数接口
replace_pictures_in_excel = process

# 测试函数（用于开发和调试）
def test():
    # 创建一个简单的测试数据
    test_data = {
        'file': open('test.xlsx', 'rb').read(),
        'settings': {
            'replaceImageBytes': open('replace.jpg', 'rb').read(),
            'replaceMode': 'all',
            'imageSize': 'original',
            'sheetOption': 'all',
            'enableDetailedLogs': True
        }
    }
    
    result = process(test_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    if result['success']:
        with open('处理后_test.xlsx', 'wb') as f:
            f.write(result['buffer'])
        print("处理后的文件已保存为: 处理后_test.xlsx")

# 主入口函数
if __name__ == "__main__":
    # 在浏览器环境中，Pyodide会通过其他方式调用该函数
    test()
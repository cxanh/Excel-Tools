import io
import openpyxl
import re
import sys
import json

# 默认设置
DEFAULT_SETTINGS = {
    'caseSensitive': False,
    'matchEntireCell': False,
    'ignoreHiddenCells': True,
    'sheetOption': 'all',
    'selectedSheets': '',
    'rangeOption': 'all',
    'customRange': '',
    'enableDetailedLogs': True
}

# 处理内容替换
def process(data):
    # 解析输入数据
    file_content = data['file']
    file_name = data['fileName']
    settings = data.get('settings', {}) or {}
    replacement_rules = settings.get('replacementRules', []) if settings else []
    
    logs = []
    
    # 使用默认设置或传入的设置
    current_settings = DEFAULT_SETTINGS.copy()
    if settings:
        current_settings.update(settings)
    
    try:
        # 加载Excel文件
        wb = openpyxl.load_workbook(io.BytesIO(file_content), read_only=False)
        logs.append(f"成功加载Excel文件: {file_name}")
        logs.append(f"包含 {len(wb.sheetnames)} 个工作表")
        
        if current_settings['enableDetailedLogs']:
            logs.append(f"工作表列表: {', '.join(wb.sheetnames)}")
        
        # 验证替换规则
        valid_rules = []
        for i, rule in enumerate(replacement_rules):
            if not rule.get('findText'):
                logs.append(f"警告: 规则 {i+1} 缺少查找文本，将被忽略")
                continue
            valid_rules.append(rule)
        
        if not valid_rules:
            logs.append("错误: 没有有效的替换规则")
            return {
                'success': False,
                'error': '没有有效的替换规则',
                'logs': logs
            }
        
        logs.append(f"共 {len(valid_rules)} 个有效替换规则")
        
        # 确定要处理的工作表
        sheets_to_process = []
        
        if current_settings['sheetOption'] == 'all':
            sheets_to_process = wb.sheetnames
            logs.append("将处理所有工作表")
        elif current_settings['sheetOption'] == 'active':
            sheets_to_process = [wb.active.title]
            logs.append(f"将处理活动工作表: {wb.active.title}")
        elif current_settings['sheetOption'] == 'selected':
            selected_sheets = [name.strip() for name in current_settings['selectedSheets'].split(',') if name.strip()]
            if not selected_sheets:
                logs.append(f"警告: 未指定工作表名称，将处理所有工作表")
                sheets_to_process = wb.sheetnames
            else:
                sheets_to_process = [name for name in selected_sheets if name in wb.sheetnames]
                if not sheets_to_process:
                    logs.append(f"警告: 未找到指定的工作表，将处理所有工作表")
                    sheets_to_process = wb.sheetnames
                elif len(sheets_to_process) < len(selected_sheets):
                    missing_sheets = [sheet for sheet in selected_sheets if sheet not in wb.sheetnames]
                    logs.append(f"警告: 未找到以下工作表: {', '.join(missing_sheets)}，将只处理找到的工作表")
        
        total_replacements = 0
        total_cells_processed = 0
        
        # 遍历需要处理的工作表
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            sheet_replacements = 0
            cell_replacements = 0
            cells_processed = 0
            
            logs.append(f"处理工作表: {sheet_name}")
            
            # 获取处理范围
            try:
                if current_settings['rangeOption'] == 'custom' and current_settings['customRange']:
                    # 解析自定义范围，例如 "A1:D100", "B:C", "1:10", "A1"
                    custom_range = current_settings['customRange'].strip()
                    if not custom_range:
                        raise ValueError("自定义范围为空")
                        
                    # 尝试使用openpyxl的range_boundaries函数
                    try:
                        from openpyxl.utils import range_boundaries
                        min_col, min_row, max_col, max_row = range_boundaries(custom_range)
                        start_col, start_row = min_col, min_row
                        end_col, end_row = max_col, max_row
                        logs.append(f"  自定义范围: {custom_range} (A1坐标: {start_row},{start_col} 到 {end_row},{end_col})")
                    except Exception as e:
                        # 回退到自定义解析
                        import re
                        range_pattern = r'^([A-Z]+)(\d+)?:([A-Z]+)(\d+)?$'
                        column_pattern = r'^([A-Z]+):([A-Z]+)$'
                        row_pattern = r'^(\d+):(\d+)$'
                        single_cell_pattern = r'^([A-Z]+)(\d+)$'
                        
                        if re.match(single_cell_pattern, custom_range.upper()):
                            # 单个单元格
                            match = re.match(single_cell_pattern, custom_range.upper())
                            start_col = openpyxl.utils.column_index_from_string(match.group(1))
                            start_row = int(match.group(2))
                            end_col, end_row = start_col, start_row
                            logs.append(f"  单个单元格: {custom_range}")
                        elif re.match(range_pattern, custom_range.upper()):
                            # 常规范围
                            match = re.match(range_pattern, custom_range.upper())
                            start_col = openpyxl.utils.column_index_from_string(match.group(1))
                            start_row = int(match.group(2)) if match.group(2) else 1
                            end_col = openpyxl.utils.column_index_from_string(match.group(3))
                            end_row = int(match.group(4)) if match.group(4) else ws.max_row
                            logs.append(f"  自定义范围: {custom_range} (A1坐标: {start_row},{start_col} 到 {end_row},{end_col})")
                        elif re.match(column_pattern, custom_range.upper()):
                            # 整列范围
                            match = re.match(column_pattern, custom_range.upper())
                            start_col = openpyxl.utils.column_index_from_string(match.group(1))
                            end_col = openpyxl.utils.column_index_from_string(match.group(2))
                            start_row, end_row = 1, ws.max_row
                            logs.append(f"  整列范围: {custom_range} (A1坐标: {start_row},{start_col} 到 {end_row},{end_col})")
                        elif re.match(row_pattern, custom_range):
                            # 整行范围
                            match = re.match(row_pattern, custom_range)
                            start_row = int(match.group(1))
                            end_row = int(match.group(2))
                            start_col, end_col = 1, ws.max_column
                            logs.append(f"  整行范围: {custom_range} (A1坐标: {start_row},{start_col} 到 {end_row},{end_col})")
                        else:
                            raise ValueError(f"无法解析范围格式: {custom_range}")
                elif current_settings['rangeOption'] == 'data':
                    # 仅处理数据区域
                    start_row, start_col = 1, 1
                    end_row, end_col = ws.max_row, ws.max_column
                    logs.append(f"  数据区域范围: A{start_row}:{openpyxl.utils.get_column_letter(end_col)}{end_row}")
                else:
                    # 处理整个工作表
                    start_row, start_col = 1, 1
                    end_row, end_col = ws.max_row, ws.max_column
                    logs.append(f"  整个工作表范围: A{start_row}:{openpyxl.utils.get_column_letter(end_col)}{end_row}")
                    
                # 验证范围边界
                start_row = max(1, start_row)
                start_col = max(1, start_col)
                end_row = min(ws.max_row, end_row)
                end_col = min(ws.max_column, end_col)
                
                if start_row > end_row or start_col > end_col:
                    logs.append(f"  警告: 无效的范围边界，将处理整个工作表")
                    start_row, start_col = 1, 1
                    end_row, end_col = ws.max_row, ws.max_column
                    
            except Exception as e:
                logs.append(f"  解析范围时出错: {str(e)}，将处理整个工作表")
                start_row, start_col = 1, 1
                end_row, end_col = ws.max_row, ws.max_column
            
            if current_settings['enableDetailedLogs']:
                logs.append(f"  开始行: {start_row}, 结束行: {end_row}")
                logs.append(f"  开始列: {openpyxl.utils.get_column_letter(start_col)}, 结束列: {openpyxl.utils.get_column_letter(end_col)}")
                logs.append(f"  预计处理单元格数量: {(end_row - start_row + 1) * (end_col - start_col + 1)}")
            
            # 遍历单元格
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    cells_processed += 1
                    
                    # 检查是否需要忽略隐藏单元格
                    if current_settings['ignoreHiddenCells']:
                        # 检查行是否隐藏
                        if row in ws.row_dimensions and ws.row_dimensions[row].hidden:
                            continue
                        # 检查列是否隐藏
                        col_letter = openpyxl.utils.get_column_letter(col)
                        if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden:
                            continue
                    
                    # 只处理有值的单元格
                    if cell.value is not None:
                        original_value = str(cell.value)
                        modified_value = original_value
                        cell_changed = False
                        original_type = type(cell.value)
                        
                        # 应用所有有效替换规则
                        for rule in valid_rules:
                            find_text = rule['findText']
                            replace_text = rule['replaceText']
                            match_mode = rule['matchMode']
                            
                            if not find_text:
                                continue
                            
                            # 确定正则标志
                            flags = re.IGNORECASE if not current_settings['caseSensitive'] else 0
                            
                            if match_mode == 'regex':
                                # 使用正则表达式替换
                                try:
                                    if current_settings['matchEntireCell']:
                                        if re.fullmatch(find_text, modified_value, flags=flags):
                                            original = modified_value
                                            modified_value = re.sub(find_text, replace_text, modified_value, flags=flags)
                                            if original != modified_value:
                                                cell_changed = True
                                                if current_settings['enableDetailedLogs']:
                                                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                    logs.append(f"    单元格 {cell_ref}: 正则替换 '{original}' → '{modified_value}'")
                                    else:
                                        if re.search(find_text, modified_value, flags=flags):
                                            original = modified_value
                                            modified_value = re.sub(find_text, replace_text, modified_value, flags=flags)
                                            if original != modified_value:
                                                cell_changed = True
                                                if current_settings['enableDetailedLogs']:
                                                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                    logs.append(f"    单元格 {cell_ref}: 正则替换 '{original}' → '{modified_value}'")
                                except re.error as e:
                                    logs.append(f"  正则表达式错误: {str(e)}")
                                    continue
                            else:
                                # 使用普通文本替换
                                if not current_settings['caseSensitive']:
                                    # 不区分大小写的普通替换
                                    if current_settings['matchEntireCell']:
                                        if modified_value.lower() == find_text.lower():
                                            modified_value = replace_text
                                            cell_changed = True
                                            if current_settings['enableDetailedLogs']:
                                                cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                logs.append(f"    单元格 {cell_ref}: 整单元格替换 '{original_value}' → '{replace_text}'")
                                    else:
                                        # 复杂的不区分大小写替换，保持原始大小写结构
                                        if find_text.lower() in modified_value.lower():
                                            # 使用正则表达式实现智能替换，保持原始大小写
                                            pattern = re.compile(re.escape(find_text), re.IGNORECASE)
                                            original = modified_value
                                            modified_value = pattern.sub(replace_text, modified_value)
                                            if original != modified_value:
                                                cell_changed = True
                                                if current_settings['enableDetailedLogs']:
                                                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                    logs.append(f"    单元格 {cell_ref}: 不区分大小写替换 '{original}' → '{modified_value}'")
                                else:
                                    # 区分大小写的普通替换
                                    if current_settings['matchEntireCell']:
                                        if modified_value == find_text:
                                            modified_value = replace_text
                                            cell_changed = True
                                            if current_settings['enableDetailedLogs']:
                                                cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                logs.append(f"    单元格 {cell_ref}: 整单元格替换 '{original_value}' → '{replace_text}'")
                                    else:
                                        if find_text in modified_value:
                                            original = modified_value
                                            modified_value = modified_value.replace(find_text, replace_text)
                                            if original != modified_value:
                                                cell_changed = True
                                                if current_settings['enableDetailedLogs']:
                                                    cell_ref = f"{openpyxl.utils.get_column_letter(col)}{row}"
                                                    logs.append(f"    单元格 {cell_ref}: 替换 '{original}' → '{modified_value}'")
                        
                        # 如果单元格内容发生了变化，更新单元格
                        if cell_changed:
                            try:
                                # 尝试保持原始数据类型
                                if original_type in [int, float]:
                                    if modified_value.replace('.', '', 1).isdigit():
                                        if '.' in modified_value:
                                            cell.value = float(modified_value)
                                        else:
                                            cell.value = int(modified_value)
                                    else:
                                        # 无法保持数字类型，使用字符串
                                        cell.value = modified_value
                                else:
                                    # 非数字类型，直接使用字符串
                                    cell.value = modified_value
                                
                                cell_replacements += 1
                                sheet_replacements += 1
                                
                            except ValueError as e:
                                # 如果类型转换失败，使用字符串
                                cell.value = modified_value
                                cell_replacements += 1
                                sheet_replacements += 1
                                
                                if current_settings['enableDetailedLogs']:
                                    logs.append(f"    类型转换失败，保持为字符串: {str(e)}")
            
            logs.append(f"  工作表 {sheet_name} 处理完成")
            logs.append(f"    - 处理单元格数量: {cells_processed}")
            logs.append(f"    - 成功替换: {sheet_replacements} 处内容")
            total_replacements += sheet_replacements
            total_cells_processed += cells_processed
        
        # 处理完成总结
        logs.append(f"\n=== 处理完成总结 ===")
        logs.append(f"总工作表数: {len(sheets_to_process)}")
        logs.append(f"总处理单元格: {total_cells_processed}")
        logs.append(f"总替换次数: {total_replacements}")
        logs.append(f"使用的有效规则数: {len(valid_rules)}")
        logs.append(f"==================")
        
        # 将处理结果写入内存
        try:
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            return {
                'success': True,
                'buffer': output.read(),
                'logs': logs,
                'details': {
                    'statistics': {
                        'totalSheets': len(sheets_to_process),
                        'totalCells': total_cells_processed,
                        'totalReplacements': total_replacements,
                        'validRules': len(valid_rules)
                    }
                }
            }
        except Exception as e:
            logs.append(f"保存文件时出错: {str(e)}")
            import traceback
            error_trace = traceback.format_exc()
            logs.append(f"详细错误: {error_trace}")
            return {
                'success': False,
                'error': f"保存文件失败: {str(e)}",
                'logs': logs
            }
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        logs.append(f"错误: {str(e)}")
        logs.append(f"详细错误: {error_trace}")
        return {
            'success': False,
            'error': str(e),
            'logs': logs
        }

# 为了保持向后兼容性，保留原函数接口
replace_content_in_excel = process

# 测试函数（用于开发和调试）
def test():
    # 创建一个简单的测试数据
    test_data = {
        'file': open('test.xlsx', 'rb').read(),
        'fileName': 'test.xlsx',
        'settings': {
            'replacementRules': [
                {
                    'findText': '测试',
                    'replaceText': '替换',
                    'matchMode': 'normal'
                }
            ],
            'caseSensitive': False,
            'matchEntireCell': False,
            'ignoreHiddenCells': True,
            'sheetOption': 'all',
            'enableDetailedLogs': True
        }
    }
    
    result = process(test_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    if result['success']:
        with open('处理后_test.xlsx', 'wb') as f:
            f.write(result['buffer'])
        print("处理后的文件已保存为: 处理后_test.xlsx")

# 主入口函数
if __name__ == "__main__":
    # 在浏览器环境中，Pyodide会通过其他方式调用该函数
    test()

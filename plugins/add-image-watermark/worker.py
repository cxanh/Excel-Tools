"""
图片添加水印插件 - Python处理脚本
为Excel中的图片添加水印
"""

import io
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image, ImageDraw, ImageFont

def add_image_watermark(file_buffer, params):
    """
    为Excel中的图片添加水印
    
    Args:
        file_buffer: Excel文件的字节数据
        params: 参数字典，包含：
            - watermark_type: 水印类型 ('text' 或 'image')
            - watermark_text: 水印文字（文本水印）
            - watermark_image: 水印图片base64（图片水印）
            - font_size: 字体大小（文本水印）
            - color: 颜色（文本水印）
            - opacity: 透明度 (0-100)
            - position: 位置 ('center', 'top-left', 'top-right', 'bottom-left', 'bottom-right')
    
    Returns:
        处理后的Excel文件字节数据
    """
    logs = []
    
    try:
        # 加载工作簿
        wb = load_workbook(io.BytesIO(file_buffer))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        # 获取参数
        watermark_type = params.get('watermark_type', 'text')
        opacity = int(params.get('opacity', 50))
        position = params.get('position', 'center')
        
        total_images = 0
        processed_images = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 检查工作表是否有图片
            if not hasattr(ws, '_images') or not ws._images:
                continue
            
            sheet_images = len(ws._images)
            total_images += sheet_images
            logs.append(f"工作表 '{sheet_name}' 包含 {sheet_images} 张图片")
            
            # 处理每张图片
            new_images = []
            for img in ws._images:
                try:
                    # 获取原始图片数据
                    if hasattr(img, '_data'):
                        img_data = img._data()
                    elif hasattr(img, 'ref'):
                        # 从图片引用获取数据
                        img_data = img.ref
                    else:
                        logs.append(f"  ⚠ 无法获取图片数据，跳过")
                        new_images.append(img)
                        continue
                    
                    # 打开图片
                    original_img = Image.open(io.BytesIO(img_data))
                    
                    # 转换为RGBA模式以支持透明度
                    if original_img.mode != 'RGBA':
                        original_img = original_img.convert('RGBA')
                    
                    # 创建水印层
                    watermark_layer = Image.new('RGBA', original_img.size, (0, 0, 0, 0))
                    draw = ImageDraw.Draw(watermark_layer)
                    
                    if watermark_type == 'text':
                        # 文本水印
                        watermark_text = params.get('watermark_text', 'WATERMARK')
                        font_size = int(params.get('font_size', 48))
                        color = params.get('color', '#000000')
                        
                        # 解析颜色
                        if color.startswith('#'):
                            r = int(color[1:3], 16)
                            g = int(color[3:5], 16)
                            b = int(color[5:7], 16)
                        else:
                            r, g, b = 0, 0, 0
                        
                        # 计算透明度
                        alpha = int(255 * opacity / 100)
                        text_color = (r, g, b, alpha)
                        
                        # 使用默认字体
                        try:
                            font = ImageFont.truetype("arial.ttf", font_size)
                        except:
                            font = ImageFont.load_default()
                        
                        # 获取文本大小
                        bbox = draw.textbbox((0, 0), watermark_text, font=font)
                        text_width = bbox[2] - bbox[0]
                        text_height = bbox[3] - bbox[1]
                        
                        # 计算位置
                        img_width, img_height = original_img.size
                        if position == 'center':
                            x = (img_width - text_width) // 2
                            y = (img_height - text_height) // 2
                        elif position == 'top-left':
                            x, y = 10, 10
                        elif position == 'top-right':
                            x = img_width - text_width - 10
                            y = 10
                        elif position == 'bottom-left':
                            x = 10
                            y = img_height - text_height - 10
                        elif position == 'bottom-right':
                            x = img_width - text_width - 10
                            y = img_height - text_height - 10
                        else:
                            x = (img_width - text_width) // 2
                            y = (img_height - text_height) // 2
                        
                        # 绘制文本
                        draw.text((x, y), watermark_text, font=font, fill=text_color)
                    
                    elif watermark_type == 'image':
                        # 图片水印
                        watermark_image_data = params.get('watermark_image', '')
                        if watermark_image_data:
                            # 解码base64图片
                            import base64
                            watermark_bytes = base64.b64decode(watermark_image_data.split(',')[1] if ',' in watermark_image_data else watermark_image_data)
                            watermark_img = Image.open(io.BytesIO(watermark_bytes))
                            
                            # 转换为RGBA
                            if watermark_img.mode != 'RGBA':
                                watermark_img = watermark_img.convert('RGBA')
                            
                            # 调整透明度
                            watermark_img = watermark_img.copy()
                            alpha = watermark_img.split()[3]
                            alpha = alpha.point(lambda p: int(p * opacity / 100))
                            watermark_img.putalpha(alpha)
                            
                            # 计算位置
                            img_width, img_height = original_img.size
                            wm_width, wm_height = watermark_img.size
                            
                            if position == 'center':
                                x = (img_width - wm_width) // 2
                                y = (img_height - wm_height) // 2
                            elif position == 'top-left':
                                x, y = 10, 10
                            elif position == 'top-right':
                                x = img_width - wm_width - 10
                                y = 10
                            elif position == 'bottom-left':
                                x = 10
                                y = img_height - wm_height - 10
                            elif position == 'bottom-right':
                                x = img_width - wm_width - 10
                                y = img_height - wm_height - 10
                            else:
                                x = (img_width - wm_width) // 2
                                y = (img_height - wm_height) // 2
                            
                            # 粘贴水印
                            watermark_layer.paste(watermark_img, (x, y), watermark_img)
                    
                    # 合并图层
                    result_img = Image.alpha_composite(original_img, watermark_layer)
                    
                    # 转换回RGB（如果原图是RGB）
                    if result_img.mode == 'RGBA':
                        # 创建白色背景
                        rgb_img = Image.new('RGB', result_img.size, (255, 255, 255))
                        rgb_img.paste(result_img, mask=result_img.split()[3])
                        result_img = rgb_img
                    
                    # 保存处理后的图片
                    img_buffer = io.BytesIO()
                    result_img.save(img_buffer, format='PNG')
                    img_buffer.seek(0)
                    
                    # 创建新的图片对象
                    new_img = OpenpyxlImage(img_buffer)
                    new_img.anchor = img.anchor
                    new_images.append(new_img)
                    
                    processed_images += 1
                    
                except Exception as e:
                    logs.append(f"  ⚠ 处理图片时出错: {str(e)}")
                    new_images.append(img)
            
            # 替换工作表中的图片
            ws._images = new_images
        
        if total_images == 0:
            logs.append("⚠ 文件中没有找到图片")
            return {
                'buffer': file_buffer,
                'logs': logs
            }
        
        logs.append(f"\n✓ 处理完成！")
        logs.append(f"  - 总图片数: {total_images}")
        logs.append(f"  - 成功处理: {processed_images}")
        if processed_images < total_images:
            logs.append(f"  - 跳过: {total_images - processed_images}")
        
        # 保存工作簿
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'buffer': output.getvalue(),
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        import traceback
        logs.append(f"详细错误: {traceback.format_exc()}")
        return {
            'buffer': file_buffer,
            'logs': logs
        }

# 主函数
def process(file_buffer, params_json):
    """
    主处理函数
    
    Args:
        file_buffer: 文件字节数据
        params_json: JSON格式的参数字符串
    
    Returns:
        包含buffer和logs的字典
    """
    params = json.loads(params_json) if isinstance(params_json, str) else params_json
    return add_image_watermark(file_buffer, params)

# 删除重复行插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook

def process(data):
    """
    处理Excel文件，删除重复行
    
    参数:
        data: 包含file(bytes)、params(dict)的字典
        params包含:
            - compareMode: 'all' 或 'columns'
            - compareColumns: 列号列表（当compareMode='columns'时）
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据和参数
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        params = data.get('params', {})
        compare_mode = params.get('compareMode', 'all')
        compare_columns = params.get('compareColumns', [])
        
        logs.append(f"比较模式: {'全行比较' if compare_mode == 'all' else f'指定列比较 (列: {compare_columns})'}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        total_deleted = 0
        total_rows_before = 0
        total_rows_after = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 获取最大行数和列数
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 0:
                logs.append(f"  工作表为空，跳过")
                continue
            
            total_rows_before += max_row
            
            # 收集所有行的数据用于比较
            seen_rows = set()
            rows_to_delete = []
            
            for row_idx in range(1, max_row + 1):
                # 根据比较模式获取行数据
                if compare_mode == 'all':
                    # 全行比较：获取所有单元格的值
                    row_data = tuple(
                        str(ws.cell(row=row_idx, column=col_idx).value) if ws.cell(row=row_idx, column=col_idx).value is not None else ''
                        for col_idx in range(1, max_col + 1)
                    )
                else:
                    # 指定列比较：只获取指定列的值
                    row_data = tuple(
                        str(ws.cell(row=row_idx, column=col_idx).value) if ws.cell(row=row_idx, column=col_idx).value is not None else ''
                        for col_idx in compare_columns
                        if col_idx <= max_col
                    )
                
                # 检查是否重复
                if row_data in seen_rows:
                    rows_to_delete.append(row_idx)
                else:
                    seen_rows.add(row_data)
            
            # 从后向前删除重复行（避免索引变化）
            deleted_count = 0
            for row_idx in reversed(rows_to_delete):
                ws.delete_rows(row_idx, 1)
                deleted_count += 1
            
            total_deleted += deleted_count
            total_rows_after += (max_row - deleted_count)
            
            logs.append(f"  原始行数: {max_row}, 删除: {deleted_count}, 剩余: {max_row - deleted_count}")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！总共删除 {total_deleted} 行重复行")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'totalRows': total_rows_before,
                    'duplicateRows': total_deleted,
                    'remainingRows': total_rows_after,
                    'compareMode': '全行比较' if compare_mode == 'all' else f'指定列比较 (列: {compare_columns})',
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

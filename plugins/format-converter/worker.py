# Excel格式转换插件 - Python处理脚本
import time
import io
import json
import zipfile
from openpyxl import load_workbook
import pandas as pd

def process(data):
    """
    将Excel文件转换为其他格式
    
    参数:
        data: 包含file(bytes)、params的字典
        params包含:
            - targetFormat: 目标格式 ('csv', 'html', 'json')
            - csvEncoding: CSV编码
            - csvDelimiter: CSV分隔符
            - htmlIncludeStyles: HTML是否包含样式
            - jsonFormat: JSON格式 ('records', 'columns')
            - jsonPretty: JSON是否格式化
            - sheetHandling: 工作表处理 ('all', 'first')
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        file_name = data.get('fileName', 'workbook.xlsx')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        target_format = params.get('targetFormat', 'csv')
        csv_encoding = params.get('csvEncoding', 'utf-8')
        csv_delimiter = params.get('csvDelimiter', ',')
        html_include_styles = params.get('htmlIncludeStyles', True)
        json_format = params.get('jsonFormat', 'records')
        json_pretty = params.get('jsonPretty', True)
        sheet_handling = params.get('sheetHandling', 'all')
        
        logs.append(f"目标格式: {target_format.upper()}")
        logs.append(f"工作表处理: {'所有工作表' if sheet_handling == 'all' else '仅第一个工作表'}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        # 确定要处理的工作表
        if sheet_handling == 'first':
            sheet_names = [wb.sheetnames[0]] if wb.sheetnames else []
        else:
            sheet_names = wb.sheetnames
        
        logs.append(f"将处理 {len(sheet_names)} 个工作表")
        
        # 创建ZIP文件（用于多文件输出）
        zip_buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)
        
        files_converted = 0
        sheets_processed = 0
        
        # 处理每个工作表
        for sheet_name in sheet_names:
            try:
                logs.append(f"处理工作表: {sheet_name}")
                
                # 读取工作表数据
                ws = wb[sheet_name]
                
                # 转换为pandas DataFrame
                data_rows = []
                for row in ws.iter_rows(values_only=True):
                    data_rows.append(row)
                
                if not data_rows:
                    logs.append(f"  工作表为空，跳过")
                    continue
                
                # 使用第一行作为列名
                df = pd.DataFrame(data_rows[1:], columns=data_rows[0])
                
                logs.append(f"  数据: {len(df)} 行 x {len(df.columns)} 列")
                
                # 根据目标格式转换
                if target_format == 'csv':
                    # 转换为CSV
                    output_buffer = io.StringIO()
                    df.to_csv(output_buffer, index=False, encoding=csv_encoding, sep=csv_delimiter)
                    output_bytes = output_buffer.getvalue().encode(csv_encoding)
                    
                    # 确定文件名
                    if len(sheet_names) > 1:
                        output_name = f"{sheet_name}.csv"
                    else:
                        base_name = file_name.rsplit('.', 1)[0]
                        output_name = f"{base_name}.csv"
                    
                    zip_file.writestr(output_name, output_bytes)
                    logs.append(f"  转换为CSV: {output_name}")
                    
                elif target_format == 'html':
                    # 转换为HTML
                    html_content = df.to_html(index=False, border=1)
                    
                    # 添加基本样式
                    if html_include_styles:
                        html_content = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{sheet_name}</title>
    <style>
        table {{ border-collapse: collapse; width: 100%; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #4CAF50; color: white; }}
        tr:nth-child(even) {{ background-color: #f2f2f2; }}
    </style>
</head>
<body>
    <h1>{sheet_name}</h1>
    {html_content}
</body>
</html>
"""
                    
                    # 确定文件名
                    if len(sheet_names) > 1:
                        output_name = f"{sheet_name}.html"
                    else:
                        base_name = file_name.rsplit('.', 1)[0]
                        output_name = f"{base_name}.html"
                    
                    zip_file.writestr(output_name, html_content.encode('utf-8'))
                    logs.append(f"  转换为HTML: {output_name}")
                    
                elif target_format == 'json':
                    # 转换为JSON
                    if json_format == 'records':
                        json_data = df.to_dict(orient='records')
                    else:  # columns
                        json_data = df.to_dict(orient='list')
                    
                    if json_pretty:
                        json_content = json.dumps(json_data, indent=2, ensure_ascii=False)
                    else:
                        json_content = json.dumps(json_data, ensure_ascii=False)
                    
                    # 确定文件名
                    if len(sheet_names) > 1:
                        output_name = f"{sheet_name}.json"
                    else:
                        base_name = file_name.rsplit('.', 1)[0]
                        output_name = f"{base_name}.json"
                    
                    zip_file.writestr(output_name, json_content.encode('utf-8'))
                    logs.append(f"  转换为JSON: {output_name}")
                
                files_converted += 1
                sheets_processed += 1
                
            except Exception as e:
                logs.append(f"  处理工作表失败: {str(e)}")
                continue
        
        # 关闭ZIP文件
        zip_file.close()
        zip_buffer.seek(0)
        result_bytes = zip_buffer.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        if files_converted == 0:
            raise ValueError("没有成功转换任何文件")
        
        logs.append(f"转换完成！共转换 {files_converted} 个文件")
        logs.append(f"处理了 {sheets_processed} 个工作表")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'filesConverted': files_converted,
                    'sheetsProcessed': sheets_processed,
                    'targetFormat': target_format,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

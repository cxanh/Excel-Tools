# 根据模板生成Excel插件 - Python处理脚本
import time
import io
import re
import zipfile
from openpyxl import load_workbook
import pandas as pd

def process(data):
    """
    基于模板和数据源批量生成Excel文件
    
    参数:
        data: 包含file(模板)、params的字典
        params包含:
            - dataFile: 数据源文件字节
            - dataFileName: 数据源文件名
            - namingPattern: 命名模式 ('sequential', 'field')
            - namingField: 命名字段
            - preserveFormulas: 是否保留公式
            - preserveFormatting: 是否保留格式
    
    返回:
        包含success、buffer(zip文件)、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理...")
        
        # 获取模板文件
        template_bytes = data.get('file')
        if not template_bytes:
            raise ValueError("未提供模板文件")
        
        # 获取参数
        params = data.get('params', {})
        data_bytes = params.get('dataFile')
        if not data_bytes:
            raise ValueError("未提供数据源文件")
        
        naming_pattern = params.get('namingPattern', 'sequential')
        naming_field = params.get('namingField', '')
        preserve_formulas = params.get('preserveFormulas', True)
        preserve_formatting = params.get('preserveFormatting', True)
        
        logs.append(f"配置: 命名模式={naming_pattern}, 保留公式={preserve_formulas}")
        
        # 加载模板
        logs.append("正在加载模板文件...")
        template_wb = load_workbook(io.BytesIO(template_bytes))
        
        # 加载数据源
        logs.append("正在加载数据源...")
        try:
            # 尝试作为Excel读取
            df = pd.read_excel(io.BytesIO(data_bytes))
        except:
            # 尝试作为CSV读取
            df = pd.read_csv(io.BytesIO(data_bytes))
        
        logs.append(f"数据源包含 {len(df)} 条记录，{len(df.columns)} 个字段")
        logs.append(f"字段列表: {', '.join(df.columns.tolist())}")
        
        # 识别模板中的占位符
        logs.append("正在识别模板占位符...")
        placeholders = set()
        for sheet_name in template_wb.sheetnames:
            ws = template_wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # 查找 {{变量名}} 格式的占位符
                        matches = re.findall(r'\{\{(\w+)\}\}', cell.value)
                        placeholders.update(matches)
        
        logs.append(f"找到 {len(placeholders)} 个占位符: {', '.join(placeholders)}")
        
        # 验证占位符是否在数据源中
        missing_fields = placeholders - set(df.columns)
        if missing_fields:
            logs.append(f"警告: 以下占位符在数据源中不存在: {', '.join(missing_fields)}")
        
        # 创建ZIP文件
        zip_buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)
        
        files_generated = 0
        variables_replaced = 0
        file_list = []
        
        # 为每条数据记录生成一个文件
        for index, row in df.iterrows():
            try:
                # 创建模板副本
                output_wb = load_workbook(io.BytesIO(template_bytes))
                
                # 确定输出文件名
                if naming_pattern == 'field' and naming_field and naming_field in df.columns:
                    file_name = f"{row[naming_field]}.xlsx"
                else:
                    file_name = f"output_{index + 1}.xlsx"
                
                logs.append(f"生成文件 {index + 1}/{len(df)}: {file_name}")
                
                record_variables = 0
                
                # 遍历所有工作表，替换占位符
                for sheet_name in output_wb.sheetnames:
                    ws = output_wb[sheet_name]
                    
                    for row_cells in ws.iter_rows():
                        for cell in row_cells:
                            if cell.value and isinstance(cell.value, str):
                                original_value = cell.value
                                new_value = original_value
                                
                                # 替换所有占位符
                                for placeholder in placeholders:
                                    pattern = f'{{{{{placeholder}}}}}'
                                    if pattern in new_value:
                                        # 获取数据值
                                        if placeholder in df.columns:
                                            data_value = row[placeholder]
                                            # 处理NaN值
                                            if pd.isna(data_value):
                                                data_value = ''
                                            new_value = new_value.replace(pattern, str(data_value))
                                            record_variables += 1
                                
                                # 如果值发生了变化，更新单元格
                                if new_value != original_value:
                                    # 保存原始格式
                                    if preserve_formatting:
                                        original_number_format = cell.number_format
                                        original_font = cell.font.copy() if cell.font else None
                                        original_fill = cell.fill.copy() if cell.fill else None
                                        original_border = cell.border.copy() if cell.border else None
                                        original_alignment = cell.alignment.copy() if cell.alignment else None
                                    
                                    # 更新值
                                    cell.value = new_value
                                    
                                    # 恢复格式
                                    if preserve_formatting:
                                        if original_number_format:
                                            cell.number_format = original_number_format
                                        if original_font:
                                            cell.font = original_font
                                        if original_fill:
                                            cell.fill = original_fill
                                        if original_border:
                                            cell.border = original_border
                                        if original_alignment:
                                            cell.alignment = original_alignment
                
                # 保存到内存
                output_buffer = io.BytesIO()
                output_wb.save(output_buffer)
                output_buffer.seek(0)
                
                # 添加到ZIP
                zip_file.writestr(file_name, output_buffer.read())
                
                files_generated += 1
                variables_replaced += record_variables
                file_list.append(file_name)
                
                logs.append(f"  替换了 {record_variables} 个变量")
                
            except Exception as e:
                logs.append(f"  生成文件失败: {str(e)}")
                continue
        
        # 关闭ZIP文件
        zip_file.close()
        zip_buffer.seek(0)
        result_bytes = zip_buffer.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！共生成 {files_generated} 个文件")
        logs.append(f"替换了 {variables_replaced} 个变量")
        logs.append(f"处理了 {len(df)} 条数据记录")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'filesGenerated': files_generated,
                    'variablesReplaced': variables_replaced,
                    'recordsProcessed': len(df),
                    'processingTime': processing_time,
                    'fileList': file_list
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

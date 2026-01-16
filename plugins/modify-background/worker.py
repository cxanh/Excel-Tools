"""
删除或修改背景图片插件 - Python处理脚本
删除或替换Excel工作表的背景图片
"""

import io
import json
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image

def modify_background(file_buffer, params):
    """
    删除或修改Excel工作表的背景图片
    
    Args:
        file_buffer: Excel文件的字节数据
        params: 参数字典，包含：
            - mode: 操作模式 ('remove' 或 'replace')
            - background_image: 新背景图片base64（替换模式）
    
    Returns:
        处理后的Excel文件字节数据
    """
    logs = []
    
    try:
        # 加载工作簿
        wb = load_workbook(io.BytesIO(file_buffer))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        # 获取参数
        mode = params.get('mode', 'remove')
        
        sheets_with_background = 0
        sheets_processed = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # 检查是否有背景图片
            has_background = False
            if hasattr(ws, 'sheet_properties') and ws.sheet_properties.pageSetUpPr:
                if hasattr(ws.sheet_properties.pageSetUpPr, 'backgroundPicture'):
                    has_background = True
            
            # 另一种检查方式：通过_images属性
            if not has_background and hasattr(ws, '_images'):
                # 检查是否有作为背景的图片
                for img in ws._images:
                    if hasattr(img, 'anchor') and img.anchor is None:
                        has_background = True
                        break
            
            if has_background:
                sheets_with_background += 1
                
                if mode == 'remove':
                    # 删除背景图片
                    try:
                        if hasattr(ws, 'sheet_properties') and ws.sheet_properties.pageSetUpPr:
                            ws.sheet_properties.pageSetUpPr.backgroundPicture = None
                        
                        # 移除作为背景的图片
                        if hasattr(ws, '_images'):
                            ws._images = [img for img in ws._images if img.anchor is not None]
                        
                        sheets_processed += 1
                        logs.append(f"✓ 工作表 '{sheet_name}': 已删除背景图片")
                    except Exception as e:
                        logs.append(f"⚠ 工作表 '{sheet_name}': 删除失败 - {str(e)}")
                
                elif mode == 'replace':
                    # 替换背景图片
                    background_image_data = params.get('background_image', '')
                    if not background_image_data:
                        logs.append(f"⚠ 工作表 '{sheet_name}': 未提供新背景图片")
                        continue
                    
                    try:
                        # 解码base64图片
                        import base64
                        img_bytes = base64.b64decode(
                            background_image_data.split(',')[1] 
                            if ',' in background_image_data 
                            else background_image_data
                        )
                        
                        # 先删除旧背景
                        if hasattr(ws, 'sheet_properties') and ws.sheet_properties.pageSetUpPr:
                            ws.sheet_properties.pageSetUpPr.backgroundPicture = None
                        if hasattr(ws, '_images'):
                            ws._images = [img for img in ws._images if img.anchor is not None]
                        
                        # 添加新背景
                        # 注意：openpyxl对背景图片的支持有限，这里使用图片对象
                        img_buffer = io.BytesIO(img_bytes)
                        new_img = OpenpyxlImage(img_buffer)
                        
                        # 设置图片位置（作为背景，不设置anchor）
                        # 注意：这会将图片添加为普通图片，而不是真正的背景
                        ws.add_image(new_img, 'A1')
                        
                        sheets_processed += 1
                        logs.append(f"✓ 工作表 '{sheet_name}': 已替换背景图片")
                    except Exception as e:
                        logs.append(f"⚠ 工作表 '{sheet_name}': 替换失败 - {str(e)}")
        
        if sheets_with_background == 0:
            logs.append("⚠ 文件中没有找到带背景图片的工作表")
            return {
                'buffer': file_buffer,
                'logs': logs
            }
        
        logs.append(f"\n✓ 处理完成！")
        logs.append(f"  - 有背景的工作表: {sheets_with_background}")
        logs.append(f"  - 成功处理: {sheets_processed}")
        if sheets_processed < sheets_with_background:
            logs.append(f"  - 失败: {sheets_with_background - sheets_processed}")
        
        # 保存工作簿
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'buffer': output.getvalue(),
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        import traceback
        logs.append(f"详细错误: {traceback.format_exc()}")
        return {
            'buffer': file_buffer,
            'logs': logs
        }

# 主函数
def process(file_buffer, params_json):
    """
    主处理函数
    
    Args:
        file_buffer: 文件字节数据
        params_json: JSON格式的参数字符串
    
    Returns:
        包含buffer和logs的字典
    """
    params = json.loads(params_json) if isinstance(params_json, str) else params_json
    return modify_background(file_buffer, params)

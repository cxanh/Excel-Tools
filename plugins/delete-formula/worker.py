import openpyxl
import io
import re
import json

def process(data):
    # 解析输入数据
    file_content = data['file']
    file_name = data['fileName']
    settings = data['settings']
    
    logs = []
    deleted_count = 0
    
    try:
        # 加载Excel文件
        workbook = openpyxl.load_workbook(io.BytesIO(file_content), data_only=False)
        logs.append(f"成功加载文件: {file_name}")
        logs.append(f"包含工作表: {workbook.sheetnames}")
        
        # 获取需要处理的工作表
        sheets_to_process = []
        if settings.get('sheetOption', 'all') == 'all':
            sheets_to_process = workbook.sheetnames
            logs.append("将处理所有工作表")
        else:
            # 解析指定工作表
            specific_sheets = settings.get('specificSheets', '')
            if specific_sheets:
                # 分割工作表名称
                sheet_patterns = [s.strip() for s in specific_sheets.split(',') if s.strip()]
                
                # 匹配工作表名称
                for sheet_name in workbook.sheetnames:
                    for pattern in sheet_patterns:
                        # 转换为正则表达式模式
                        regex_pattern = pattern.replace('*', '.*').replace('?', '.')
                        if re.match(regex_pattern, sheet_name, re.IGNORECASE):
                            sheets_to_process.append(sheet_name)
                            break
                
                if sheets_to_process:
                    logs.append(f"将处理工作表: {sheets_to_process}")
                else:
                    logs.append(f"没有匹配的工作表: {specific_sheets}")
                    return {
                        'success': False,
                        'error': '没有匹配的工作表',
                        'logs': logs,
                        'deletedCount': deleted_count
                    }
            else:
                logs.append("没有指定工作表")
                return {
                    'success': False,
                    'error': '没有指定工作表',
                    'logs': logs,
                    'deletedCount': deleted_count
                }
        
        # 获取处理选项
        process_options = settings.get('processOptions', ['formula', 'arrayFormula', 'blankCells'])
        delete_formula = 'formula' in process_options
        delete_array_formula = 'arrayFormula' in process_options
        ignore_blank = 'blankCells' in process_options
        
        logs.append(f"处理选项: {'公式' if delete_formula else ''}{'，数组公式' if delete_array_formula else ''}{'，忽略空白' if ignore_blank else ''}")
        
        # 获取范围选项
        range_option = settings.get('rangeOption', 'used')
        specific_range = settings.get('specificRange', '')
        
        logs.append(f"数据区域: {range_option}{f' ({specific_range})' if range_option == 'specific' else ''}")
        
        # 处理每个工作表
        for sheet_name in sheets_to_process:
            sheet = workbook[sheet_name]
            sheet_deleted = 0
            
            logs.append(f"开始处理工作表: {sheet_name}")
            
            # 获取需要处理的单元格范围
            cells_to_process = []
            
            if range_option == 'all':
                # 处理整个工作表
                cells_to_process = sheet.iter_rows()
            elif range_option == 'used':
                # 只处理已使用区域
                if sheet.dimensions:
                    cells_to_process = sheet.iter_rows(min_row=1, max_row=sheet.max_row, 
                                                      min_col=1, max_col=sheet.max_column)
                else:
                    logs.append(f"工作表 {sheet_name} 没有已使用区域")
                    continue
            else:  # specific
                # 处理指定区域
                if specific_range:
                    # 支持多个区域用逗号分隔
                    ranges = [r.strip() for r in specific_range.split(',') if r.strip()]
                    for rng in ranges:
                        try:
                            # 解析区域范围
                            cell_range = sheet[rng]
                            for row in cell_range:
                                for cell in row:
                                    cells_to_process.append([cell])  # 转换为与iter_rows兼容的格式
                        except Exception as e:
                            logs.append(f"解析区域 {rng} 失败: {str(e)}")
                else:
                    logs.append(f"没有指定区域")
                    continue
            
            # 处理单元格
            for row in cells_to_process:
                for cell in row:
                    # 跳过空白单元格
                    if ignore_blank and cell.value is None:
                        continue
                    
                    # 检查是否有公式
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        # 检查是否是数组公式
                        is_array_formula = hasattr(cell, '_is_array_formula') and cell._is_array_formula
                        
                        # 检查是否需要删除这个公式
                        if (delete_formula and not is_array_formula) or (delete_array_formula and is_array_formula):
                            # 获取公式的计算值
                            # 需要重新加载工作簿以获取计算值
                            temp_buffer = io.BytesIO()
                            workbook.save(temp_buffer)
                            temp_buffer.seek(0)
                            
                            # 以data_only=True重新加载以获取计算值
                            temp_workbook = openpyxl.load_workbook(temp_buffer, data_only=True)
                            temp_sheet = temp_workbook[sheet_name]
                            
                            # 获取计算值
                            calculated_value = temp_sheet.cell(row=cell.row, column=cell.column).value
                            
                            # 替换公式为计算值
                            cell.value = calculated_value
                            
                            # 如果是数组公式，清除数组公式标记
                            if is_array_formula and hasattr(cell, '_is_array_formula'):
                                cell._is_array_formula = False
                                cell._array_formula = None
                            
                            sheet_deleted += 1
                            deleted_count += 1
                        
                        # 清理临时工作簿
                        temp_workbook.close()
            
            if sheet_deleted > 0:
                logs.append(f"工作表 {sheet_name} 删除了 {sheet_deleted} 个公式")
            else:
                logs.append(f"工作表 {sheet_name} 没有找到需要删除的公式")
        
        # 保存处理后的文件
        output_buffer = io.BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        logs.append(f"文件 {file_name} 处理完成，共删除 {deleted_count} 个公式")
        
        return {
            'success': True,
            'buffer': output_buffer.getvalue(),
            'deletedCount': deleted_count,
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"处理失败: {str(e)}")
        return {
            'success': False,
            'error': str(e),
            'deletedCount': deleted_count,
            'logs': logs
        }

# 测试函数（用于开发和调试）
def test():
    # 创建一个简单的测试数据
    test_data = {
        'file': open('test.xlsx', 'rb').read(),
        'fileName': 'test.xlsx',
        'settings': {
            'sheetOption': 'all',
            'specificSheets': '',
            'processOptions': ['formula', 'arrayFormula', 'blankCells'],
            'rangeOption': 'used',
            'specificRange': ''
        }
    }
    
    result = process(test_data)
    print(json.dumps(result, ensure_ascii=False, indent=2))
    
    if result['success']:
        with open('处理后_test.xlsx', 'wb') as f:
            f.write(result['buffer'])
        print("处理后的文件已保存为: 处理后_test.xlsx")

if __name__ == '__main__':
    test()

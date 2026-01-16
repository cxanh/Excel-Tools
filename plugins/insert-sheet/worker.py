"""
插入Sheet插件 - Python处理脚本
在Excel的指定位置插入新工作表
"""

import io
import json
from openpyxl import load_workbook

def insert_sheet(file_buffer, params):
    """
    在Excel的指定位置插入新工作表
    
    Args:
        file_buffer: Excel文件的字节数据
        params: 参数字典，包含：
            - sheet_name: 新工作表名称
            - position: 插入位置（索引，-1表示末尾）
            - count: 插入数量
    
    Returns:
        处理后的Excel文件字节数据
    """
    logs = []
    
    try:
        wb = load_workbook(io.BytesIO(file_buffer))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        sheet_name = params.get('sheet_name', 'NewSheet')
        position = params.get('position', -1)
        count = params.get('count', 1)
        
        inserted_count = 0
        
        for i in range(count):
            # 生成唯一的工作表名称
            final_name = sheet_name if count == 1 else f"{sheet_name}{i+1}"
            
            # 检查名称冲突
            counter = 1
            original_name = final_name
            while final_name in wb.sheetnames:
                final_name = f"{original_name}_{counter}"
                counter += 1
            
            # 插入工作表
            if position == -1:
                ws = wb.create_sheet(final_name)
            else:
                ws = wb.create_sheet(final_name, position + i)
            
            inserted_count += 1
            logs.append(f"✓ 已插入工作表: {final_name}")
        
        logs.append(f"\n✓ 处理完成！共插入 {inserted_count} 个工作表")
        logs.append(f"当前工作表总数: {len(wb.sheetnames)}")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'buffer': output.getvalue(),
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        import traceback
        logs.append(f"详细错误: {traceback.format_exc()}")
        return {
            'buffer': file_buffer,
            'logs': logs
        }

def process(file_buffer, params_json):
    params = json.loads(params_json) if isinstance(params_json, str) else params_json
    return insert_sheet(file_buffer, params)

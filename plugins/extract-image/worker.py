# 提取图片插件 - Python处理脚本
import time
import io
import json
import zipfile
from openpyxl import load_workbook
from PIL import Image as PILImage

def process(data):
    """
    从Excel文件中提取所有图片
    
    参数:
        data: 包含file(bytes)、fileName、params的字典
        params包含:
            - imageFormat: 图片格式 ('original', 'png', 'jpg')
            - namingPattern: 命名模式 ('sequential', 'sheet', 'position')
            - createSubfolders: 是否创建子文件夹
            - includeMetadata: 是否包含元数据
    
    返回:
        包含success、buffer(zip文件)、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        file_name = data.get('fileName', 'workbook.xlsx')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        image_format = params.get('imageFormat', 'original')
        naming_pattern = params.get('namingPattern', 'sequential')
        create_subfolders = params.get('createSubfolders', True)
        include_metadata = params.get('includeMetadata', False)
        
        logs.append(f"配置: 格式={image_format}, 命名={naming_pattern}, 子文件夹={create_subfolders}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        # 创建ZIP文件
        zip_buffer = io.BytesIO()
        zip_file = zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED)
        
        total_images_extracted = 0
        sheets_processed = 0
        total_size = 0
        image_list = []
        metadata_list = []
        
        # 全局图片计数器
        global_counter = 1
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 检查是否有图片
            if not hasattr(ws, '_images') or not ws._images:
                logs.append(f"  没有找到图片")
                continue
            
            images_count = len(ws._images)
            logs.append(f"  找到 {images_count} 个图片")
            sheets_processed += 1
            
            # 工作表图片计数器
            sheet_counter = 1
            
            # 提取每个图片
            for img in ws._images:
                try:
                    # 获取图片数据
                    img_data = img._data()
                    pil_img = PILImage.open(io.BytesIO(img_data))
                    
                    # 确定文件扩展名
                    if image_format == 'original':
                        # 保持原格式
                        img_format = pil_img.format or 'PNG'
                        ext = img_format.lower()
                    elif image_format == 'png':
                        img_format = 'PNG'
                        ext = 'png'
                    else:  # jpg
                        img_format = 'JPEG'
                        ext = 'jpg'
                        # JPG不支持透明度，需要转换
                        if pil_img.mode in ('RGBA', 'LA', 'P'):
                            background = PILImage.new('RGB', pil_img.size, (255, 255, 255))
                            if pil_img.mode == 'P':
                                pil_img = pil_img.convert('RGBA')
                            background.paste(pil_img, mask=pil_img.split()[-1] if pil_img.mode == 'RGBA' else None)
                            pil_img = background
                    
                    # 确定文件名
                    if naming_pattern == 'sequential':
                        img_name = f"image_{global_counter}.{ext}"
                    elif naming_pattern == 'sheet':
                        img_name = f"{sheet_name}_{sheet_counter}.{ext}"
                    else:  # position
                        # 尝试获取锚点位置
                        anchor = img.anchor
                        if hasattr(anchor, '_from') and hasattr(anchor._from, 'col'):
                            col = anchor._from.col
                            row = anchor._from.row
                            from openpyxl.utils import get_column_letter
                            cell_ref = f"{get_column_letter(col + 1)}{row + 1}"
                            img_name = f"{cell_ref}.{ext}"
                        else:
                            img_name = f"image_{global_counter}.{ext}"
                    
                    # 确定ZIP中的路径
                    if create_subfolders:
                        # 使用Excel文件名（去除扩展名）作为文件夹名
                        base_name = file_name.rsplit('.', 1)[0]
                        zip_path = f"{base_name}/{sheet_name}/{img_name}"
                    else:
                        zip_path = img_name
                    
                    # 保存图片到内存
                    img_buffer = io.BytesIO()
                    pil_img.save(img_buffer, format=img_format)
                    img_bytes = img_buffer.getvalue()
                    img_size = len(img_bytes)
                    
                    # 添加到ZIP
                    zip_file.writestr(zip_path, img_bytes)
                    
                    # 更新统计
                    total_images_extracted += 1
                    total_size += img_size
                    image_list.append(zip_path)
                    
                    # 收集元数据
                    if include_metadata:
                        metadata_list.append({
                            'fileName': img_name,
                            'path': zip_path,
                            'sheet': sheet_name,
                            'format': img_format,
                            'size': img_size,
                            'width': pil_img.width,
                            'height': pil_img.height
                        })
                    
                    logs.append(f"  提取: {zip_path} ({img_size} 字节)")
                    
                    global_counter += 1
                    sheet_counter += 1
                    
                except Exception as e:
                    logs.append(f"  提取图片失败: {str(e)}")
                    continue
        
        # 添加元数据文件
        if include_metadata and metadata_list:
            metadata_json = json.dumps(metadata_list, indent=2, ensure_ascii=False)
            zip_file.writestr('metadata.json', metadata_json)
            logs.append("已生成元数据文件: metadata.json")
        
        # 关闭ZIP文件
        zip_file.close()
        zip_buffer.seek(0)
        result_bytes = zip_buffer.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        if total_images_extracted == 0:
            logs.append("警告: 没有提取到任何图片")
        else:
            logs.append(f"处理完成！共提取 {total_images_extracted} 个图片")
            logs.append(f"处理了 {sheets_processed} 个工作表")
            logs.append(f"总大小: {total_size} 字节")
            logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'imagesExtracted': total_images_extracted,
                    'sheetsProcessed': sheets_processed,
                    'totalSize': total_size,
                    'processingTime': processing_time,
                    'imageList': image_list
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

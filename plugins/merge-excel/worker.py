# Excel合并插件 - Python处理脚本
import time
import io
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

def process(data):
    """
    合并多个Excel文件
    
    参数:
        data: 包含files(list)、params(dict)的字典
        files: 文件列表，每个包含name和data
        params包含:
            - mergeMode: 'sheets' 或 'single'
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始合并文件...")
        
        # 获取文件列表和参数
        files = data.get('files', [])
        if not files or len(files) < 2:
            raise ValueError("至少需要2个文件进行合并")
        
        params = data.get('params', {})
        merge_mode = params.get('mergeMode', 'sheets')
        
        logs.append(f"合并模式: {'保留工作表' if merge_mode == 'sheets' else '合并到单个工作表'}")
        logs.append(f"待合并文件数: {len(files)}")
        
        # 创建新工作簿
        merged_wb = Workbook()
        merged_wb.remove(merged_wb.active)  # 删除默认工作表
        
        total_sheets = 0
        total_rows = 0
        
        if merge_mode == 'sheets':
            # 保留工作表模式：复制所有工作表到新文件
            sheet_names = set()
            
            for file_idx, file_info in enumerate(files):
                logs.append(f"处理文件 {file_idx + 1}/{len(files)}: {file_info['name']}")
                
                # 加载源文件
                source_wb = load_workbook(io.BytesIO(file_info['data']))
                
                # 复制每个工作表
                for sheet_name in source_wb.sheetnames:
                    source_ws = source_wb[sheet_name]
                    
                    # 处理重名工作表
                    new_sheet_name = sheet_name
                    counter = 1
                    while new_sheet_name in sheet_names:
                        new_sheet_name = f"{sheet_name}_{counter}"
                        counter += 1
                    
                    sheet_names.add(new_sheet_name)
                    
                    # 创建新工作表
                    target_ws = merged_wb.create_sheet(title=new_sheet_name)
                    
                    # 复制数据和格式
                    for row in source_ws.iter_rows():
                        for cell in row:
                            target_cell = target_ws.cell(
                                row=cell.row,
                                column=cell.column,
                                value=cell.value
                            )
                            
                            # 复制格式
                            if cell.has_style:
                                target_cell.font = copy(cell.font)
                                target_cell.border = copy(cell.border)
                                target_cell.fill = copy(cell.fill)
                                target_cell.number_format = copy(cell.number_format)
                                target_cell.protection = copy(cell.protection)
                                target_cell.alignment = copy(cell.alignment)
                    
                    # 复制列宽
                    for col_letter in source_ws.column_dimensions:
                        if col_letter in source_ws.column_dimensions:
                            target_ws.column_dimensions[col_letter].width = \
                                source_ws.column_dimensions[col_letter].width
                    
                    # 复制行高
                    for row_num in source_ws.row_dimensions:
                        if row_num in source_ws.row_dimensions:
                            target_ws.row_dimensions[row_num].height = \
                                source_ws.row_dimensions[row_num].height
                    
                    total_sheets += 1
                    total_rows += source_ws.max_row
                    
                    logs.append(f"  复制工作表: {new_sheet_name} ({source_ws.max_row}行)")
        
        else:
            # 合并到单个工作表模式
            merged_ws = merged_wb.create_sheet(title="合并数据")
            current_row = 1
            
            for file_idx, file_info in enumerate(files):
                logs.append(f"处理文件 {file_idx + 1}/{len(files)}: {file_info['name']}")
                
                # 加载源文件
                source_wb = load_workbook(io.BytesIO(file_info['data']))
                
                # 合并所有工作表的数据
                for sheet_name in source_wb.sheetnames:
                    source_ws = source_wb[sheet_name]
                    total_sheets += 1
                    
                    logs.append(f"  合并工作表: {sheet_name} ({source_ws.max_row}行)")
                    
                    # 复制数据
                    for row in source_ws.iter_rows():
                        for cell in row:
                            if cell.value is not None:
                                target_cell = merged_ws.cell(
                                    row=current_row,
                                    column=cell.column,
                                    value=cell.value
                                )
                                
                                # 复制格式
                                if cell.has_style:
                                    target_cell.font = copy(cell.font)
                                    target_cell.border = copy(cell.border)
                                    target_cell.fill = copy(cell.fill)
                                    target_cell.number_format = copy(cell.number_format)
                                    target_cell.alignment = copy(cell.alignment)
                        
                        current_row += 1
                    
                    total_rows += source_ws.max_row
        
        # 保存到内存
        logs.append("正在保存合并后的文件...")
        output = io.BytesIO()
        merged_wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"合并完成！共处理 {len(files)} 个文件，{total_sheets} 个工作表，{total_rows} 行数据")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'filesProcessed': len(files),
                    'totalSheets': total_sheets,
                    'totalRows': total_rows,
                    'mergeMode': '保留工作表' if merge_mode == 'sheets' else '合并到单个工作表',
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

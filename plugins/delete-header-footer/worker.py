#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
删除 Excel 页眉页脚插件
"""

import io
import json
from openpyxl import load_workbook
from openpyxl.worksheet.header_footer import HeaderFooter


def process(input_data):
    """处理删除页眉页脚请求
    
    Args:
        input_data: 包含输入文件和设置的数据
    
    Returns:
        dict: 处理结果，包含处理后的文件buffer和日志
    """
    result = {
        "success": False,
        "logs": [],
        "buffer": None
    }
    
    try:
        # 解析输入数据
        file = input_data["file"]
        file_name = input_data["fileName"]
        settings = input_data.get("settings", {})
        
        # 加载Excel文件
        result["logs"].append(f"正在加载文件: {file_name}")
        wb = load_workbook(io.BytesIO(file))
        
        # 获取所有需要处理的工作表
        sheet_names = settings.get("sheetNames", ["*"])
        if "*" in sheet_names:
            sheets_to_process = wb.sheetnames
        else:
            sheets_to_process = [sheet for sheet in sheet_names if sheet in wb.sheetnames]
        
        result["logs"].append(f"将处理 {len(sheets_to_process)} 个工作表")
        
        # 遍历工作表处理
        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            result["logs"].append(f"处理工作表: {sheet_name}")
            
            # 创建空的HeaderFooter对象
            header_footer = HeaderFooter()
            
            # 清空页眉页脚设置
            header_footer.left_header.text = ""
            header_footer.center_header.text = ""
            header_footer.right_header.text = ""
            header_footer.left_footer.text = ""
            header_footer.center_footer.text = ""
            header_footer.right_footer.text = ""
            
            # 应用空的页眉页脚设置
            ws.header_footer = header_footer
            
            result["logs"].append(f"  已删除页眉页脚")
        
        # 保存处理后的文件
        result["logs"].append("保存处理后的文件")
        output_stream = io.BytesIO()
        wb.save(output_stream)
        output_stream.seek(0)
        
        result["buffer"] = output_stream.getvalue()
        result["success"] = True
        result["logs"].append("文件处理完成")
        
    except Exception as e:
        result["success"] = False
        result["logs"].append(f"处理错误: {str(e)}")
        result["error"] = str(e)
    
    return result

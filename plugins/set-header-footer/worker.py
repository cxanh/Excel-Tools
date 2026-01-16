"""
设置页眉页脚插件 Worker 脚本
批量设置Excel文件的页眉页脚
"""
import openpyxl
from openpyxl.worksheet.header_footer import HeaderFooter
import io
import json

def process(file_content: bytes, options: dict) -> dict:
    """
    处理Excel文件，设置页眉页脚
    
    Args:
        file_content: Excel文件的二进制内容
        options: 处理选项
            - header_left: 页眉左侧内容
            - header_center: 页眉中间内容
            - header_right: 页眉右侧内容
            - footer_left: 页脚左侧内容
            - footer_center: 页脚中间内容
            - footer_right: 页脚右侧内容
            - different_odd_even: 是否区分奇偶页
            - apply_to_all: 是否应用到所有工作表
    
    Returns:
        包含处理结果的字典
    """
    logs = []
    
    try:
        # 加载工作簿
        logs.append("正在加载Excel文件...")
        wb = openpyxl.load_workbook(io.BytesIO(file_content))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        # 获取配置
        header_left = options.get('header_left', '')
        header_center = options.get('header_center', '')
        header_right = options.get('header_right', '')
        footer_left = options.get('footer_left', '')
        footer_center = options.get('footer_center', '')
        footer_right = options.get('footer_right', '')
        different_odd_even = options.get('different_odd_even', False)
        apply_to_all = options.get('apply_to_all', True)
        
        sheets_processed = 0
        
        # 处理工作表
        sheets_to_process = wb.worksheets if apply_to_all else [wb.active]
        
        for sheet in sheets_to_process:
            logs.append(f"正在处理工作表: {sheet.title}")
            
            # 设置页眉
            if header_left:
                sheet.oddHeader.left.text = header_left
                if different_odd_even:
                    sheet.evenHeader.left.text = header_left
            
            if header_center:
                sheet.oddHeader.center.text = header_center
                if different_odd_even:
                    sheet.evenHeader.center.text = header_center
            
            if header_right:
                sheet.oddHeader.right.text = header_right
                if different_odd_even:
                    sheet.evenHeader.right.text = header_right
            
            # 设置页脚
            if footer_left:
                sheet.oddFooter.left.text = footer_left
                if different_odd_even:
                    sheet.evenFooter.left.text = footer_left
            
            if footer_center:
                sheet.oddFooter.center.text = footer_center
                if different_odd_even:
                    sheet.evenFooter.center.text = footer_center
            
            if footer_right:
                sheet.oddFooter.right.text = footer_right
                if different_odd_even:
                    sheet.evenFooter.right.text = footer_right
            
            sheets_processed += 1
            logs.append(f"✓ 工作表 {sheet.title} 页眉页脚设置完成")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        logs.append(f"✓ 处理完成！共设置 {sheets_processed} 个工作表的页眉页脚")
        
        return {
            'success': True,
            'buffer': output.getvalue(),
            'logs': logs,
            'statistics': {
                'totalSheets': len(wb.sheetnames),
                'sheetsProcessed': sheets_processed,
                'hasOddEven': different_odd_even
            }
        }
        
    except Exception as e:
        logs.append(f"✗ 错误: {str(e)}")
        return {
            'success': False,
            'buffer': None,
            'logs': logs,
            'error': str(e)
        }

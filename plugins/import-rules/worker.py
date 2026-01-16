# 导入规则修改内容插件 - Python处理脚本
import time
import io
import json
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd

def process(data):
    """
    从文件导入规则并应用到Excel
    
    参数:
        data: 包含file(Excel)、params的字典
        params包含:
            - rulesFile: 规则文件字节
            - rulesFileName: 规则文件名
            - caseSensitive: 是否区分大小写
            - stopOnError: 遇到错误时是否停止
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理...")
        
        # 获取Excel文件
        excel_bytes = data.get('file')
        if not excel_bytes:
            raise ValueError("未提供Excel文件")
        
        # 获取参数
        params = data.get('params', {})
        rules_bytes = params.get('rulesFile')
        rules_file_name = params.get('rulesFileName', 'rules.json')
        case_sensitive = params.get('caseSensitive', False)
        stop_on_error = params.get('stopOnError', False)
        
        if not rules_bytes:
            raise ValueError("未提供规则文件")
        
        logs.append(f"配置: 区分大小写={case_sensitive}, 遇错停止={stop_on_error}")
        
        # 加载规则文件
        logs.append(f"正在加载规则文件: {rules_file_name}")
        rules = []
        
        if rules_file_name.endswith('.json'):
            # JSON格式
            rules_text = rules_bytes.decode('utf-8')
            rules = json.loads(rules_text)
        elif rules_file_name.endswith('.csv'):
            # CSV格式
            df = pd.read_csv(io.BytesIO(rules_bytes))
            rules = df.to_dict('records')
        else:
            raise ValueError("不支持的规则文件格式，请使用JSON或CSV")
        
        logs.append(f"加载了 {len(rules)} 条规则")
        
        # 验证规则格式
        required_fields = ['sheet', 'column', 'find', 'replace']
        for i, rule in enumerate(rules):
            missing_fields = [f for f in required_fields if f not in rule]
            if missing_fields:
                error_msg = f"规则 {i+1} 缺少必需字段: {', '.join(missing_fields)}"
                if stop_on_error:
                    raise ValueError(error_msg)
                else:
                    logs.append(f"警告: {error_msg}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(excel_bytes))
        
        rules_applied = 0
        total_replacements = 0
        sheets_processed = set()
        
        # 应用每条规则
        for rule_idx, rule in enumerate(rules):
            try:
                sheet_name = rule.get('sheet', '')
                column_name = rule.get('column', '')
                find_text = str(rule.get('find', ''))
                replace_text = str(rule.get('replace', ''))
                use_regex = rule.get('regex', False)
                
                logs.append(f"规则 {rule_idx + 1}: 工作表='{sheet_name}', 列='{column_name}', 查找='{find_text}', 替换='{replace_text}', 正则={use_regex}")
                
                # 检查工作表是否存在
                if sheet_name not in wb.sheetnames:
                    error_msg = f"  工作表 '{sheet_name}' 不存在"
                    logs.append(f"警告: {error_msg}")
                    if stop_on_error:
                        raise ValueError(error_msg)
                    continue
                
                ws = wb[sheet_name]
                sheets_processed.add(sheet_name)
                
                # 确定列索引
                try:
                    if column_name.isdigit():
                        col_idx = int(column_name)
                    else:
                        col_idx = column_index_from_string(column_name)
                except:
                    error_msg = f"  无效的列名: '{column_name}'"
                    logs.append(f"警告: {error_msg}")
                    if stop_on_error:
                        raise ValueError(error_msg)
                    continue
                
                rule_replacements = 0
                
                # 遍历列中的所有单元格
                for row in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row, column=col_idx)
                    
                    if cell.value is None:
                        continue
                    
                    cell_value = str(cell.value)
                    
                    # 执行替换
                    if use_regex:
                        # 正则表达式替换
                        try:
                            flags = 0 if case_sensitive else re.IGNORECASE
                            new_value = re.sub(find_text, replace_text, cell_value, flags=flags)
                        except Exception as e:
                            error_msg = f"  正则表达式错误: {str(e)}"
                            logs.append(f"警告: {error_msg}")
                            if stop_on_error:
                                raise
                            continue
                    else:
                        # 普通文本替换
                        if case_sensitive:
                            if find_text in cell_value:
                                new_value = cell_value.replace(find_text, replace_text)
                            else:
                                new_value = cell_value
                        else:
                            # 不区分大小写的替换
                            pattern = re.compile(re.escape(find_text), re.IGNORECASE)
                            new_value = pattern.sub(replace_text, cell_value)
                    
                    # 如果值发生了变化，更新单元格
                    if new_value != cell_value:
                        cell.value = new_value
                        rule_replacements += 1
                
                logs.append(f"  替换了 {rule_replacements} 个单元格")
                total_replacements += rule_replacements
                rules_applied += 1
                
            except Exception as e:
                error_msg = f"  应用规则失败: {str(e)}"
                logs.append(f"错误: {error_msg}")
                if stop_on_error:
                    raise
                continue
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！应用了 {rules_applied} 条规则")
        logs.append(f"共替换了 {total_replacements} 个单元格")
        logs.append(f"处理了 {len(sheets_processed)} 个工作表")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'rulesApplied': rules_applied,
                    'replacements': total_replacements,
                    'sheetsProcessed': len(sheets_processed),
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

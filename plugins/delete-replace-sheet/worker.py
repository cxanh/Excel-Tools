"""
删除或替换Sheet插件 - Python处理脚本
删除或替换Excel的工作表
"""

import io
import json
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def delete_replace_sheet(file_buffer, params):
    """
    删除或替换Excel的工作表
    
    Args:
        file_buffer: Excel文件的字节数据
        params: 参数字典，包含：
            - mode: 操作模式 ('delete' 或 'replace')
            - sheet_names: 要操作的工作表名称列表
            - replace_file: 替换用的Excel文件buffer（替换模式）
    
    Returns:
        处理后的Excel文件字节数据
    """
    logs = []
    
    try:
        # 加载工作簿
        wb = load_workbook(io.BytesIO(file_buffer))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        logs.append(f"工作表列表: {', '.join(wb.sheetnames)}")
        
        # 获取参数
        mode = params.get('mode', 'delete')
        sheet_names = params.get('sheet_names', [])
        
        if not sheet_names:
            logs.append("⚠ 未指定要操作的工作表")
            return {
                'buffer': file_buffer,
                'logs': logs
            }
        
        processed_count = 0
        
        if mode == 'delete':
            # 删除模式
            for sheet_name in sheet_names:
                if sheet_name not in wb.sheetnames:
                    logs.append(f"⚠ 工作表 '{sheet_name}' 不存在，跳过")
                    continue
                
                # 检查是否是最后一个工作表
                if len(wb.sheetnames) == 1:
                    logs.append(f"⚠ 工作表 '{sheet_name}' 是最后一个工作表，无法删除")
                    continue
                
                try:
                    wb.remove(wb[sheet_name])
                    processed_count += 1
                    logs.append(f"✓ 已删除工作表: {sheet_name}")
                except Exception as e:
                    logs.append(f"✗ 删除工作表 '{sheet_name}' 失败: {str(e)}")
        
        elif mode == 'replace':
            # 替换模式
            replace_file_buffer = params.get('replace_file')
            if not replace_file_buffer:
                logs.append("✗ 替换模式需要提供替换文件")
                return {
                    'buffer': file_buffer,
                    'logs': logs
                }
            
            try:
                # 加载替换文件
                import base64
                if isinstance(replace_file_buffer, str):
                    replace_bytes = base64.b64decode(
                        replace_file_buffer.split(',')[1] 
                        if ',' in replace_file_buffer 
                        else replace_file_buffer
                    )
                else:
                    replace_bytes = replace_file_buffer
                
                replace_wb = load_workbook(io.BytesIO(replace_bytes))
                logs.append(f"成功加载替换文件，共 {len(replace_wb.sheetnames)} 个工作表")
                
                for sheet_name in sheet_names:
                    if sheet_name not in wb.sheetnames:
                        logs.append(f"⚠ 工作表 '{sheet_name}' 不存在，跳过")
                        continue
                    
                    # 使用替换文件的第一个工作表
                    if len(replace_wb.sheetnames) == 0:
                        logs.append(f"⚠ 替换文件中没有工作表")
                        break
                    
                    try:
                        # 删除原工作表
                        old_index = wb.sheetnames.index(sheet_name)
                        wb.remove(wb[sheet_name])
                        
                        # 复制替换工作表
                        source_sheet = replace_wb[replace_wb.sheetnames[0]]
                        target_sheet = wb.create_sheet(sheet_name, old_index)
                        
                        # 复制数据和格式
                        for row in source_sheet.iter_rows():
                            for cell in row:
                                target_cell = target_sheet[cell.coordinate]
                                target_cell.value = cell.value
                                if cell.has_style:
                                    target_cell.font = cell.font.copy()
                                    target_cell.border = cell.border.copy()
                                    target_cell.fill = cell.fill.copy()
                                    target_cell.number_format = cell.number_format
                                    target_cell.protection = cell.protection.copy()
                                    target_cell.alignment = cell.alignment.copy()
                        
                        # 复制列宽
                        for col in source_sheet.column_dimensions:
                            if col in source_sheet.column_dimensions:
                                target_sheet.column_dimensions[col].width = source_sheet.column_dimensions[col].width
                        
                        # 复制行高
                        for row in source_sheet.row_dimensions:
                            if row in source_sheet.row_dimensions:
                                target_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height
                        
                        processed_count += 1
                        logs.append(f"✓ 已替换工作表: {sheet_name}")
                    except Exception as e:
                        logs.append(f"✗ 替换工作表 '{sheet_name}' 失败: {str(e)}")
            
            except Exception as e:
                logs.append(f"✗ 加载替换文件失败: {str(e)}")
                return {
                    'buffer': file_buffer,
                    'logs': logs
                }
        
        if processed_count == 0:
            logs.append("⚠ 没有工作表被处理")
        else:
            logs.append(f"\n✓ 处理完成！共处理 {processed_count} 个工作表")
        
        # 保存工作簿
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {
            'buffer': output.getvalue(),
            'logs': logs
        }
        
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        import traceback
        logs.append(f"详细错误: {traceback.format_exc()}")
        return {
            'buffer': file_buffer,
            'logs': logs
        }

# 主函数
def process(file_buffer, params_json):
    """
    主处理函数
    
    Args:
        file_buffer: 文件字节数据
        params_json: JSON格式的参数字符串
    
    Returns:
        包含buffer和logs的字典
    """
    params = json.loads(params_json) if isinstance(params_json, str) else params_json
    return delete_replace_sheet(file_buffer, params)

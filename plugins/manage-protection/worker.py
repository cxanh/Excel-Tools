"""添加或删除保护插件"""
import io
import json
from openpyxl import load_workbook

def process(file_buffer, params_json):
    logs = []
    try:
        params = json.loads(params_json) if isinstance(params_json, str) else params_json
        wb = load_workbook(io.BytesIO(file_buffer))
        logs.append(f"成功加载工作簿，共 {len(wb.sheetnames)} 个工作表")
        
        mode = params.get('mode', 'add')
        password = params.get('password', '')
        
        processed_count = 0
        
        if mode == 'add':
            # 添加保护
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password
                processed_count += 1
                logs.append(f"✓ 已保护工作表: {sheet_name}")
        else:
            # 删除保护
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                ws.protection.sheet = False
                ws.protection.password = None
                processed_count += 1
                logs.append(f"✓ 已取消保护工作表: {sheet_name}")
        
        logs.append(f"\n✓ 处理完成！共处理 {processed_count} 个工作表")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return {'buffer': output.getvalue(), 'logs': logs}
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        return {'buffer': file_buffer, 'logs': logs}

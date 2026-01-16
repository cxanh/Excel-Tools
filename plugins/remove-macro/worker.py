# 删除Excel宏插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook

def process(data):
    """
    删除Excel文件中的所有VBA宏代码和模块
    
    参数:
        data: 包含file(bytes)、fileName、params的字典
        params包含:
            - convertToXlsx: 是否转换为.xlsx格式
            - showStatistics: 是否显示统计信息
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        file_name = data.get('fileName', 'unknown.xlsx')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        convert_to_xlsx = params.get('convertToXlsx', True)
        
        logs.append(f"配置: 转换为xlsx={convert_to_xlsx}")
        
        # 检测文件格式
        is_macro_enabled = file_name.lower().endswith('.xlsm')
        logs.append(f"文件格式: {'.xlsm (启用宏)' if is_macro_enabled else '.xlsx (无宏)'}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes), keep_vba=True)
        
        # 检查是否包含宏
        macros_removed = 0
        if hasattr(wb, 'vba_archive') and wb.vba_archive:
            logs.append("检测到VBA宏代码")
            # 移除VBA存档
            wb.vba_archive = None
            macros_removed = 1  # 标记已移除宏
            logs.append("已删除所有VBA宏代码和模块")
        else:
            logs.append("文件不包含VBA宏代码")
        
        # 统计工作表数量
        sheets_processed = len(wb.sheetnames)
        logs.append(f"处理了 {sheets_processed} 个工作表")
        
        # 确定输出格式
        output_format = '.xlsx'
        if is_macro_enabled and convert_to_xlsx:
            logs.append("将文件格式从.xlsm转换为.xlsx")
        elif not convert_to_xlsx:
            output_format = '.xlsm' if is_macro_enabled else '.xlsx'
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        
        # 注意：保存时不保留VBA，因为我们已经将vba_archive设置为None
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        if macros_removed == 0:
            logs.append("提示: 文件已是无宏版本")
        else:
            logs.append(f"处理完成！已删除宏代码")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'macrosRemoved': macros_removed,
                    'sheetsProcessed': sheets_processed,
                    'fileFormat': output_format,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        import traceback
        logs.append(f"详细错误: {traceback.format_exc()}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

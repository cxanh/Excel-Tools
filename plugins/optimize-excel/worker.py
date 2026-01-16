"""Excel优化与压缩插件"""
import io
import json
from openpyxl import load_workbook

def process(file_buffer, params_json):
    logs = []
    try:
        params = json.loads(params_json) if isinstance(params_json, str) else params_json
        wb = load_workbook(io.BytesIO(file_buffer))
        
        original_size = len(file_buffer)
        logs.append(f"原始文件大小: {original_size / 1024:.2f} KB")
        
        # 优化操作
        optimizations = 0
        
        # 1. 删除未使用的样式
        # 2. 清理空白单元格
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # 清理空白行
            rows_to_delete = []
            for row in ws.iter_rows():
                if all(cell.value is None for cell in row):
                    rows_to_delete.append(row[0].row)
            
            for row_num in reversed(rows_to_delete):
                ws.delete_rows(row_num)
                optimizations += 1
        
        logs.append(f"✓ 已清理 {optimizations} 个空白行")
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        optimized_size = len(output.getvalue())
        reduction = ((original_size - optimized_size) / original_size * 100) if original_size > 0 else 0
        
        logs.append(f"优化后文件大小: {optimized_size / 1024:.2f} KB")
        logs.append(f"减小: {reduction:.1f}%")
        
        return {'buffer': output.getvalue(), 'logs': logs}
    except Exception as e:
        logs.append(f"✗ 处理失败: {str(e)}")
        return {'buffer': file_buffer, 'logs': logs}

# 删除公式插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook

def process(data):
    """
    删除Excel文件中的所有公式，保留计算结果值
    
    参数:
        data: 包含file(bytes)、params的字典
        params包含:
            - preserveFormatting: 是否保留格式
            - showStatistics: 是否显示统计信息
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        preserve_formatting = params.get('preserveFormatting', True)
        
        logs.append(f"配置: 保留格式={preserve_formatting}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes), data_only=False)
        
        total_formulas_removed = 0
        sheets_processed = 0
        cells_processed = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            sheet_formulas = 0
            sheet_cells = 0
            
            # 遍历所有单元格
            for row in ws.iter_rows():
                for cell in row:
                    sheet_cells += 1
                    
                    # 检查是否为公式单元格
                    if cell.data_type == 'f':  # 'f' 表示公式
                        # 获取公式的计算值
                        formula_value = cell.value
                        calculated_value = None
                        
                        # 尝试获取计算值
                        # 注意：openpyxl在data_only=False模式下，cell.value是公式字符串
                        # 我们需要重新加载文件以获取计算值
                        try:
                            # 创建一个临时工作簿来获取计算值
                            wb_data = load_workbook(io.BytesIO(file_bytes), data_only=True)
                            ws_data = wb_data[sheet_name]
                            calculated_value = ws_data[cell.coordinate].value
                            wb_data.close()
                        except:
                            # 如果无法获取计算值，使用None
                            calculated_value = None
                        
                        # 保存原始格式
                        if preserve_formatting:
                            original_number_format = cell.number_format
                            original_font = cell.font.copy() if cell.font else None
                            original_fill = cell.fill.copy() if cell.fill else None
                            original_border = cell.border.copy() if cell.border else None
                            original_alignment = cell.alignment.copy() if cell.alignment else None
                        
                        # 将公式替换为计算值
                        cell.value = calculated_value
                        
                        # 恢复格式
                        if preserve_formatting:
                            if original_number_format:
                                cell.number_format = original_number_format
                            if original_font:
                                cell.font = original_font
                            if original_fill:
                                cell.fill = original_fill
                            if original_border:
                                cell.border = original_border
                            if original_alignment:
                                cell.alignment = original_alignment
                        
                        sheet_formulas += 1
                        logs.append(f"  {cell.coordinate}: 删除公式 '{formula_value}' -> 值 '{calculated_value}'")
            
            if sheet_formulas > 0:
                logs.append(f"  工作表统计: 删除{sheet_formulas}个公式，处理{sheet_cells}个单元格")
                sheets_processed += 1
            else:
                logs.append(f"  没有找到公式")
            
            total_formulas_removed += sheet_formulas
            cells_processed += sheet_cells
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        if total_formulas_removed == 0:
            logs.append("提示: 文件中没有找到公式")
        else:
            logs.append(f"处理完成！共删除 {total_formulas_removed} 个公式")
            logs.append(f"处理了 {sheets_processed} 个工作表")
            logs.append(f"检查了 {cells_processed} 个单元格")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'formulasRemoved': total_formulas_removed,
                    'sheetsProcessed': sheets_processed,
                    'cellsProcessed': cells_processed,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

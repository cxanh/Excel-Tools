# 提取指定内容插件 - Python处理脚本
import time
import io
import re
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
import pandas as pd

def process(data):
    """
    按条件提取Excel中的特定内容
    
    参数:
        data: 包含file(bytes)、params的字典
        params包含:
            - conditions: 筛选条件列表
            - includeHeader: 是否包含表头
            - preserveFormatting: 是否保留格式
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        conditions = params.get('conditions', [])
        include_header = params.get('includeHeader', True)
        preserve_formatting = params.get('preserveFormatting', True)
        
        if not conditions:
            raise ValueError("未提供筛选条件")
        
        logs.append(f"配置: {len(conditions)} 个筛选条件, 包含表头={include_header}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        # 创建新工作簿用于存储提取的内容
        from openpyxl import Workbook
        new_wb = Workbook()
        new_wb.remove(new_wb.active)  # 删除默认工作表
        
        total_rows_extracted = 0
        total_rows_original = 0
        
        # 处理每个工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 读取所有数据
            data_rows = []
            for row in ws.iter_rows(values_only=True):
                data_rows.append(row)
            
            if not data_rows:
                logs.append(f"  工作表为空，跳过")
                continue
            
            # 使用第一行作为列名
            headers = data_rows[0]
            data_rows = data_rows[1:]
            total_rows_original += len(data_rows)
            
            logs.append(f"  原始数据: {len(data_rows)} 行")
            
            # 创建DataFrame便于筛选
            df = pd.DataFrame(data_rows, columns=headers)
            
            # 应用筛选条件
            mask = None
            for idx, condition in enumerate(conditions):
                column = condition.get('column', '')
                operator = condition.get('operator', 'equals')
                value = condition.get('value', '')
                value2 = condition.get('value2', '')
                logic = condition.get('logic', 'AND')
                
                logs.append(f"  条件 {idx + 1}: 列='{column}', 操作='{operator}', 值='{value}'")
                
                # 确定列索引
                try:
                    if column.isdigit():
                        col_idx = int(column) - 1
                        col_name = headers[col_idx] if col_idx < len(headers) else None
                    elif column in df.columns:
                        col_name = column
                    else:
                        # 尝试作为列字母
                        col_idx = column_index_from_string(column) - 1
                        col_name = headers[col_idx] if col_idx < len(headers) else None
                    
                    if col_name is None:
                        logs.append(f"    警告: 列 '{column}' 不存在，跳过此条件")
                        continue
                    
                except Exception as e:
                    logs.append(f"    警告: 无效的列名 '{column}': {str(e)}")
                    continue
                
                # 构建筛选条件
                try:
                    if operator == 'equals':
                        condition_mask = df[col_name].astype(str) == str(value)
                    elif operator == 'contains':
                        condition_mask = df[col_name].astype(str).str.contains(str(value), na=False)
                    elif operator == 'startswith':
                        condition_mask = df[col_name].astype(str).str.startswith(str(value), na=False)
                    elif operator == 'endswith':
                        condition_mask = df[col_name].astype(str).str.endswith(str(value), na=False)
                    elif operator == 'regex':
                        condition_mask = df[col_name].astype(str).str.match(str(value), na=False)
                    elif operator == 'greater':
                        condition_mask = pd.to_numeric(df[col_name], errors='coerce') > float(value)
                    elif operator == 'less':
                        condition_mask = pd.to_numeric(df[col_name], errors='coerce') < float(value)
                    elif operator == 'between':
                        numeric_col = pd.to_numeric(df[col_name], errors='coerce')
                        condition_mask = (numeric_col >= float(value)) & (numeric_col <= float(value2))
                    else:
                        logs.append(f"    警告: 不支持的操作符 '{operator}'")
                        continue
                    
                    # 组合条件
                    if mask is None:
                        mask = condition_mask
                    else:
                        if logic == 'AND':
                            mask = mask & condition_mask
                        else:  # OR
                            mask = mask | condition_mask
                    
                except Exception as e:
                    logs.append(f"    警告: 应用条件失败: {str(e)}")
                    continue
            
            # 如果没有有效的筛选条件，跳过
            if mask is None:
                logs.append(f"  没有有效的筛选条件")
                continue
            
            # 应用筛选
            filtered_df = df[mask]
            rows_extracted = len(filtered_df)
            total_rows_extracted += rows_extracted
            
            logs.append(f"  提取了 {rows_extracted} 行数据")
            
            if rows_extracted == 0:
                logs.append(f"  没有匹配的数据")
                continue
            
            # 创建新工作表
            new_ws = new_wb.create_sheet(title=sheet_name)
            
            # 写入表头
            if include_header:
                for col_idx, header in enumerate(headers, 1):
                    new_ws.cell(row=1, column=col_idx, value=header)
                start_row = 2
            else:
                start_row = 1
            
            # 写入数据
            for row_idx, (_, row_data) in enumerate(filtered_df.iterrows(), start_row):
                for col_idx, value in enumerate(row_data, 1):
                    new_ws.cell(row=row_idx, column=col_idx, value=value)
            
            # 如果需要保留格式，复制原始格式
            if preserve_formatting and include_header:
                # 复制表头格式
                for col_idx in range(1, len(headers) + 1):
                    source_cell = ws.cell(row=1, column=col_idx)
                    target_cell = new_ws.cell(row=1, column=col_idx)
                    
                    if source_cell.font:
                        target_cell.font = source_cell.font.copy()
                    if source_cell.fill:
                        target_cell.fill = source_cell.fill.copy()
                    if source_cell.border:
                        target_cell.border = source_cell.border.copy()
                    if source_cell.alignment:
                        target_cell.alignment = source_cell.alignment.copy()
        
        # 保存到内存
        logs.append("正在保存提取的内容...")
        output = io.BytesIO()
        new_wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        # 计算匹配率
        match_rate = round((total_rows_extracted / total_rows_original * 100), 2) if total_rows_original > 0 else 0
        
        logs.append(f"提取完成！共提取 {total_rows_extracted} 行数据")
        logs.append(f"原始数据: {total_rows_original} 行")
        logs.append(f"匹配率: {match_rate}%")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'rowsExtracted': total_rows_extracted,
                    'totalRows': total_rows_original,
                    'matchRate': match_rate,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

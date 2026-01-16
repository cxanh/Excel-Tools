# Excel拆分插件 - Python处理脚本
import time
import io
from openpyxl import Workbook, load_workbook
from copy import copy

def process(data):
    """
    拆分Excel文件
    
    参数:
        data: 包含file(bytes)、params(dict)的字典
        params包含:
            - splitMode: 'sheets' 或 'rows'
            - rowsPerFile: 每个文件的行数（当splitMode='rows'时）
    
    返回:
        包含success、logs、details的字典
        details包含files列表，每个文件包含fileName、buffer、rows、sheets
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始拆分文件...")
        
        # 获取文件数据和参数
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        params = data.get('params', {})
        split_mode = params.get('splitMode', 'sheets')
        rows_per_file = params.get('rowsPerFile', 1000)
        
        logs.append(f"拆分模式: {'按工作表' if split_mode == 'sheets' else f'按行数 (每{rows_per_file}行)'}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        result_files = []
        
        if split_mode == 'sheets':
            # 按工作表拆分：每个工作表一个文件
            logs.append(f"共有 {len(wb.sheetnames)} 个工作表")
            
            for sheet_idx, sheet_name in enumerate(wb.sheetnames):
                logs.append(f"处理工作表 {sheet_idx + 1}/{len(wb.sheetnames)}: {sheet_name}")
                
                source_ws = wb[sheet_name]
                
                # 创建新工作簿
                new_wb = Workbook()
                new_wb.remove(new_wb.active)  # 删除默认工作表
                new_ws = new_wb.create_sheet(title=sheet_name)
                
                # 复制数据和格式
                row_count = 0
                for row in source_ws.iter_rows():
                    for cell in row:
                        target_cell = new_ws.cell(
                            row=cell.row,
                            column=cell.column,
                            value=cell.value
                        )
                        
                        # 复制格式
                        if cell.has_style:
                            target_cell.font = copy(cell.font)
                            target_cell.border = copy(cell.border)
                            target_cell.fill = copy(cell.fill)
                            target_cell.number_format = copy(cell.number_format)
                            target_cell.protection = copy(cell.protection)
                            target_cell.alignment = copy(cell.alignment)
                    
                    row_count += 1
                
                # 复制列宽
                for col_letter in source_ws.column_dimensions:
                    if col_letter in source_ws.column_dimensions:
                        new_ws.column_dimensions[col_letter].width = \
                            source_ws.column_dimensions[col_letter].width
                
                # 复制行高
                for row_num in source_ws.row_dimensions:
                    if row_num in source_ws.row_dimensions:
                        new_ws.row_dimensions[row_num].height = \
                            source_ws.row_dimensions[row_num].height
                
                # 保存到内存
                output = io.BytesIO()
                new_wb.save(output)
                output.seek(0)
                result_bytes = output.read()
                
                # 生成文件名
                safe_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')
                file_name = f"{safe_sheet_name}.xlsx"
                
                result_files.append({
                    'fileName': file_name,
                    'buffer': result_bytes,
                    'rows': row_count,
                    'sheets': 1
                })
                
                logs.append(f"  生成文件: {file_name} ({row_count}行)")
        
        else:
            # 按行数拆分
            for sheet_name in wb.sheetnames:
                source_ws = wb[sheet_name]
                max_row = source_ws.max_row
                
                if max_row == 0:
                    logs.append(f"工作表 {sheet_name} 为空，跳过")
                    continue
                
                logs.append(f"处理工作表: {sheet_name} ({max_row}行)")
                
                # 计算需要拆分的文件数
                num_files = (max_row + rows_per_file - 1) // rows_per_file
                
                for file_idx in range(num_files):
                    start_row = file_idx * rows_per_file + 1
                    end_row = min((file_idx + 1) * rows_per_file, max_row)
                    
                    # 创建新工作簿
                    new_wb = Workbook()
                    new_wb.remove(new_wb.active)
                    new_ws = new_wb.create_sheet(title=sheet_name)
                    
                    # 复制指定范围的行
                    new_row_idx = 1
                    for row_idx in range(start_row, end_row + 1):
                        for col_idx in range(1, source_ws.max_column + 1):
                            source_cell = source_ws.cell(row=row_idx, column=col_idx)
                            target_cell = new_ws.cell(
                                row=new_row_idx,
                                column=col_idx,
                                value=source_cell.value
                            )
                            
                            # 复制格式
                            if source_cell.has_style:
                                target_cell.font = copy(source_cell.font)
                                target_cell.border = copy(source_cell.border)
                                target_cell.fill = copy(source_cell.fill)
                                target_cell.number_format = copy(source_cell.number_format)
                                target_cell.protection = copy(source_cell.protection)
                                target_cell.alignment = copy(source_cell.alignment)
                        
                        new_row_idx += 1
                    
                    # 复制列宽
                    for col_letter in source_ws.column_dimensions:
                        if col_letter in source_ws.column_dimensions:
                            new_ws.column_dimensions[col_letter].width = \
                                source_ws.column_dimensions[col_letter].width
                    
                    # 保存到内存
                    output = io.BytesIO()
                    new_wb.save(output)
                    output.seek(0)
                    result_bytes = output.read()
                    
                    # 生成文件名
                    safe_sheet_name = sheet_name.replace('/', '_').replace('\\', '_')
                    file_name = f"{safe_sheet_name}_part{file_idx + 1}.xlsx"
                    
                    result_files.append({
                        'fileName': file_name,
                        'buffer': result_bytes,
                        'rows': end_row - start_row + 1,
                        'sheets': 1
                    })
                    
                    logs.append(f"  生成文件: {file_name} (行{start_row}-{end_row})")
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"拆分完成！共生成 {len(result_files)} 个文件")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'logs': logs,
            'details': {
                'files': result_files,
                'statistics': {
                    'filesGenerated': len(result_files),
                    'splitMode': '按工作表' if split_mode == 'sheets' else f'按行数 (每{rows_per_file}行)',
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

# 按规则修改内容插件 - Python处理脚本
import time
import io
import re
from openpyxl import load_workbook

def process(data):
    """
    处理Excel文件，按规则替换内容
    
    参数:
        data: 包含file(bytes)、params(dict)的字典
        params包含:
            - rules: 规则列表，每个规则包含find、replace、useRegex
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据和参数
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        params = data.get('params', {})
        rules = params.get('rules', [])
        
        if not rules:
            raise ValueError("未提供替换规则")
        
        logs.append(f"共有 {len(rules)} 条替换规则")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        total_replacements = 0
        cells_processed = 0
        rule_details = []
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 获取最大行数和列数
            max_row = ws.max_row
            max_col = ws.max_column
            
            if max_row == 0:
                logs.append(f"  工作表为空，跳过")
                continue
            
            # 按顺序应用每条规则
            for rule_idx, rule in enumerate(rules):
                find_text = rule.get('find', '')
                replace_text = rule.get('replace', '')
                use_regex = rule.get('useRegex', False)
                
                if not find_text:
                    continue
                
                rule_replacements = 0
                
                # 遍历所有单元格
                for row_idx in range(1, max_row + 1):
                    for col_idx in range(1, max_col + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        
                        # 只处理包含文本的单元格
                        if cell.value is None:
                            continue
                        
                        cell_value = str(cell.value)
                        cells_processed += 1
                        
                        # 应用替换规则
                        if use_regex:
                            # 正则表达式替换
                            try:
                                new_value, count = re.subn(find_text, replace_text, cell_value)
                                if count > 0:
                                    cell.value = new_value
                                    rule_replacements += count
                            except re.error as e:
                                logs.append(f"  警告: 规则{rule_idx + 1}的正则表达式无效: {str(e)}")
                        else:
                            # 普通文本替换
                            if find_text in cell_value:
                                count = cell_value.count(find_text)
                                new_value = cell_value.replace(find_text, replace_text)
                                cell.value = new_value
                                rule_replacements += count
                
                total_replacements += rule_replacements
                rule_details.append({
                    'ruleIndex': rule_idx + 1,
                    'replacements': rule_replacements
                })
                
                logs.append(f"  规则{rule_idx + 1}: 替换了 {rule_replacements} 次")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！总共替换 {total_replacements} 次")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'rulesApplied': len(rules),
                    'totalReplacements': total_replacements,
                    'cellsProcessed': cells_processed,
                    'ruleDetails': rule_details,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

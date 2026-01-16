# 删除Excel图片插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook

def process(data):
    """
    删除Excel文件中的所有图片
    
    参数:
        data: 包含file(bytes)的字典
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 记录原始文件大小
        original_size = len(file_bytes)
        logs.append(f"原始文件大小: {original_size} 字节")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        total_images_deleted = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 删除图片
            # openpyxl中图片存储在_images属性中
            if hasattr(ws, '_images') and ws._images:
                images_count = len(ws._images)
                ws._images = []
                total_images_deleted += images_count
                logs.append(f"  删除了 {images_count} 个图片")
            else:
                logs.append(f"  没有找到图片")
            
            # 删除图表（charts）
            if hasattr(ws, '_charts') and ws._charts:
                charts_count = len(ws._charts)
                ws._charts = []
                total_images_deleted += charts_count
                logs.append(f"  删除了 {charts_count} 个图表")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算新文件大小
        new_size = len(result_bytes)
        size_reduction_bytes = original_size - new_size
        size_reduction_percent = (size_reduction_bytes / original_size * 100) if original_size > 0 else 0
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！共删除 {total_images_deleted} 个图片/图表")
        logs.append(f"文件大小从 {original_size} 字节减小到 {new_size} 字节")
        logs.append(f"减小了 {size_reduction_bytes} 字节 ({size_reduction_percent:.1f}%)")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'imagesDeleted': total_images_deleted,
                    'originalSize': original_size,
                    'newSize': new_size,
                    'sizeReduction': f"{size_reduction_percent:.1f}%",
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

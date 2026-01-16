# 替换图片插件 - Python处理脚本
import time
import io
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage

def process(data):
    """
    替换Excel文件中的所有图片为指定的新图片
    
    参数:
        data: 包含file(bytes)、params的字典
        params包含:
            - replaceImage: 新图片的字节数据
            - replaceImageName: 新图片的文件名
    
    返回:
        包含success、buffer、logs、details的字典
    """
    logs = []
    start_time = time.time()
    
    try:
        logs.append("开始处理文件...")
        
        # 获取文件数据
        file_bytes = data.get('file')
        if not file_bytes:
            raise ValueError("未提供文件数据")
        
        # 获取参数
        params = data.get('params', {})
        replace_image_bytes = params.get('replaceImage')
        replace_image_name = params.get('replaceImageName', 'image.png')
        
        if not replace_image_bytes:
            raise ValueError("未提供替换图片")
        
        logs.append(f"替换图片: {replace_image_name}")
        
        # 加载Excel文件
        logs.append("正在加载Excel文件...")
        wb = load_workbook(io.BytesIO(file_bytes))
        
        # 加载替换图片
        logs.append("正在加载替换图片...")
        replace_pil_image = PILImage.open(io.BytesIO(replace_image_bytes))
        
        total_images_replaced = 0
        sheets_processed = 0
        
        # 遍历所有工作表
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            logs.append(f"处理工作表: {sheet_name}")
            
            # 检查是否有图片
            if hasattr(ws, '_images') and ws._images:
                images_count = len(ws._images)
                logs.append(f"  找到 {images_count} 个图片")
                
                # 收集原始图片的位置和大小信息
                image_positions = []
                for img in ws._images:
                    # 保存图片的锚点信息（位置和大小）
                    anchor = img.anchor
                    image_positions.append({
                        'anchor': anchor,
                        'width': img.width if hasattr(img, 'width') else None,
                        'height': img.height if hasattr(img, 'height') else None
                    })
                
                # 清空原有图片
                ws._images = []
                
                # 添加新图片到相同位置
                for pos_info in image_positions:
                    # 创建新图片的副本
                    img_stream = io.BytesIO()
                    replace_pil_image.save(img_stream, format=replace_pil_image.format or 'PNG')
                    img_stream.seek(0)
                    
                    # 创建openpyxl图片对象
                    new_img = Image(img_stream)
                    
                    # 如果有原始尺寸信息，保持原始尺寸
                    if pos_info['width'] and pos_info['height']:
                        new_img.width = pos_info['width']
                        new_img.height = pos_info['height']
                    
                    # 设置锚点（位置）
                    new_img.anchor = pos_info['anchor']
                    
                    # 添加到工作表
                    ws.add_image(new_img)
                
                total_images_replaced += images_count
                sheets_processed += 1
                logs.append(f"  替换了 {images_count} 个图片")
            else:
                logs.append(f"  没有找到图片")
        
        # 保存到内存
        logs.append("正在保存处理后的文件...")
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        result_bytes = output.read()
        
        # 计算处理时间
        processing_time = int((time.time() - start_time) * 1000)
        
        logs.append(f"处理完成！共替换 {total_images_replaced} 个图片")
        logs.append(f"处理了 {sheets_processed} 个工作表")
        logs.append(f"处理时间: {processing_time}ms")
        
        return {
            'success': True,
            'buffer': result_bytes,
            'logs': logs,
            'details': {
                'statistics': {
                    'imagesReplaced': total_images_replaced,
                    'sheetsProcessed': sheets_processed,
                    'processingTime': processing_time
                }
            }
        }
        
    except Exception as e:
        logs.append(f"错误: {str(e)}")
        return {
            'success': False,
            'logs': logs,
            'error': str(e)
        }

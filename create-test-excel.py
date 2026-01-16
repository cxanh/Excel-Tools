#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
创建一个测试用的 Excel 文件
"""

from openpyxl import Workbook

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "测试数据"

# 添加表头
headers = ["姓名", "年龄", "城市", "职业"]
ws.append(headers)

# 添加测试数据
data = [
    ["张三", 28, "北京", "工程师"],
    ["李四", 32, "上海", "设计师"],
    ["王五", 25, "广州", "产品经理"],
    ["", 30, "深圳", ""],  # 包含空白单元格
    ["赵六", 35, "杭州", "架构师"],
    ["", "", "", ""],  # 空白行
    ["钱七", 29, "成都", "测试工程师"],
]

for row in data:
    ws.append(row)

# 添加第二个工作表
ws2 = wb.create_sheet("销售数据")
ws2.append(["产品", "数量", "单价", "总价"])
ws2.append(["笔记本", 10, 5000, 50000])
ws2.append(["鼠标", 50, 50, 2500])
ws2.append(["键盘", 30, 200, 6000])

# 保存文件
filename = "test.xlsx"
wb.save(filename)
print(f"✅ 测试文件已创建: {filename}")
print(f"   - 工作表 1: {ws.title} ({ws.max_row} 行 × {ws.max_column} 列)")
print(f"   - 工作表 2: {ws2.title} ({ws2.max_row} 行 × {ws2.max_column} 列)")

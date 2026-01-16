import sys
import os
import re
from openpyxl import load_workbook
from io import BytesIO

# 设置标准输出为UTF-8
sys.stdout.reconfigure(encoding='utf-8')


def replace_content_in_excel(file_data, options):
    """
    按规则修改Excel内容
    :param file_data: Excel文件的二进制数据
    :param options: 修改选项，包含：
        - sheet_names: 要处理的工作表名称列表
        - replace_rules: 替换规则列表，每个规则包含：
            - find: 要查找的内容
            - replace: 替换后的内容
            - match_type: 匹配类型（exact: 精确匹配, regex: 正则表达式）
            - case_sensitive: 是否区分大小写
            - columns: 要处理的列索引列表（如：[0, 2, 4]）
            - delete: 是否删除匹配内容（设置为True时，replace字段无效）
    :return: 处理结果
    """
    try:
        # 加载Excel文件
        workbook = load_workbook(filename=BytesIO(file_data))
        results = {
            'total_replaced': 0,
            'sheet_results': {}
        }

        # 获取要处理的工作表
        sheet_names = options.get('sheet_names', [])
        if not sheet_names:
            sheet_names = workbook.sheetnames

        # 获取替换规则
        replace_rules = options.get('replace_rules', [])
        if not replace_rules:
            return {
                'success': False,
                'error': '没有提供替换规则'
            }

        # 处理每个工作表
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            sheet_result = {
                'replaced_count': 0
            }

            # 处理每个替换规则
            for rule in replace_rules:
                replaced = replace_in_sheet(sheet, rule)
                sheet_result['replaced_count'] += replaced

            # 保存工作表结果
            results['sheet_results'][sheet_name] = sheet_result
            results['total_replaced'] += sheet_result['replaced_count']

        # 保存修改后的Excel文件
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return {
            'success': True,
            'data': {
                'file_data': output.getvalue().hex(),
                'results': results
            }
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def replace_in_sheet(sheet, rule):
    """
    在单个工作表中执行替换
    :param sheet: Excel工作表对象
    :param rule: 替换规则
    :return: 替换的单元格数量
    """
    find_content = rule.get('find', '')
    replace_content = rule.get('replace', '')
    match_type = rule.get('match_type', 'exact')
    case_sensitive = rule.get('case_sensitive', False)
    columns = rule.get('columns', [])
    delete = rule.get('delete', False)

    replaced_count = 0

    # 编译正则表达式（如果需要）
    regex_pattern = None
    if match_type == 'regex':
        flags = 0 if case_sensitive else re.IGNORECASE
        regex_pattern = re.compile(find_content, flags)

    # 确定要处理的列范围
    min_col = 1
    max_col = sheet.max_column

    if columns:
        # 转换为1-based索引，并确定范围
        min_col = min(columns) + 1
        max_col = max(columns) + 1

    # 遍历所有单元格
    for row in sheet.iter_rows(min_col=min_col, max_col=max_col):
        for cell in row:
            # 检查是否在指定列中（如果有指定）
            if columns and (cell.column - 1) not in columns:
                continue

            # 获取单元格当前值
            current_value = cell.value
            if current_value is None:
                continue

            # 转换为字符串处理
            original_value = str(current_value)
            new_value = original_value
            matched = False

            # 执行匹配和替换
            if match_type == 'exact':
                # 精确匹配
                if case_sensitive:
                    if original_value == find_content:
                        matched = True
                else:
                    if original_value.lower() == find_content.lower():
                        matched = True
            elif match_type == 'regex':
                # 正则表达式匹配
                if regex_pattern.search(original_value):
                    matched = True

            # 如果匹配，执行替换或删除
            if matched:
                if delete:
                    # 删除内容（设为None）
                    new_value = None
                else:
                    # 执行替换
                    if match_type == 'exact':
                        if case_sensitive:
                            new_value = original_value.replace(find_content, replace_content)
                        else:
                            # 不区分大小写的替换
                            new_value = re.sub(re.escape(find_content), replace_content, original_value, flags=re.IGNORECASE)
                    elif match_type == 'regex':
                        # 正则表达式替换
                        new_value = regex_pattern.sub(replace_content, original_value)

                # 更新单元格值
                cell.value = new_value
                replaced_count += 1

    return replaced_count


# 主函数，处理命令行参数
# 注意：在Pyodide环境中，__name__ != '__main__'，所以这段测试代码不会执行
if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': '缺少参数'}), flush=True)
        sys.exit(1)

    try:
        # 解析参数
        params = json.loads(sys.argv[1])
        file_data = bytes.fromhex(params['file_data'])
        options = params['options']

        # 调用处理函数
        result = replace_content_in_excel(file_data, options)
        print(json.dumps(result, ensure_ascii=False), flush=True)
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False), flush=True)
        sys.exit(1)

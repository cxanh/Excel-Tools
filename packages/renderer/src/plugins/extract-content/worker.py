import sys
import os
from openpyxl import load_workbook
import csv
import json
from io import BytesIO, StringIO

# 设置标准输出为UTF-8
sys.stdout.reconfigure(encoding='utf-8')


def extract_content_from_excel(file_data, extract_type, options):
    """
    从Excel文件中提取指定内容
    :param file_data: Excel文件的二进制数据
    :param extract_type: 提取类型：columns, rows, condition
    :param options: 提取选项
    :return: 提取结果
    """
    try:
        # 加载Excel文件
        workbook = load_workbook(filename=BytesIO(file_data))
        result = {}

        # 根据提取类型处理
        if extract_type == 'columns':
            result = extract_by_columns(workbook, options)
        elif extract_type == 'rows':
            result = extract_by_rows(workbook, options)
        elif extract_type == 'condition':
            result = extract_by_condition(workbook, options)
        else:
            return {
                'success': False,
                'error': f'不支持的提取类型: {extract_type}'
            }

        return {
            'success': True,
            'data': result
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def extract_by_columns(workbook, options):
    """
    按列提取内容
    :param workbook: Excel工作簿对象
    :param options: 提取选项，包含：
        - sheet_names: 要处理的工作表名称列表
        - columns: 要提取的列索引列表（如：[0, 2, 4]）
    :return: 提取结果
    """
    sheet_names = options.get('sheet_names', [])
    columns = options.get('columns', [])
    result = {}

    # 如果没有指定工作表，处理所有工作表
    if not sheet_names:
        sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        extracted_data = []

        # 获取表头
        header = []
        for col_idx in columns:
            header_cell = sheet.cell(row=1, column=col_idx + 1)
            header.append(header_cell.value or f'列{col_idx + 1}')
        extracted_data.append(header)

        # 获取数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            extracted_row = []
            for col_idx in columns:
                if col_idx < len(row):
                    extracted_row.append(row[col_idx])
                else:
                    extracted_row.append('')
            extracted_data.append(extracted_row)

        result[sheet_name] = extracted_data

    return result


def extract_by_rows(workbook, options):
    """
    按行提取内容
    :param workbook: Excel工作簿对象
    :param options: 提取选项，包含：
        - sheet_names: 要处理的工作表名称列表
        - start_row: 开始行索引（从0开始）
        - end_row: 结束行索引（从0开始，-1表示到末尾）
    :return: 提取结果
    """
    sheet_names = options.get('sheet_names', [])
    start_row = options.get('start_row', 0)
    end_row = options.get('end_row', -1)
    result = {}

    # 如果没有指定工作表，处理所有工作表
    if not sheet_names:
        sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        extracted_data = []

        # 计算实际的开始行和结束行（openpyxl是从1开始的）
        actual_start_row = start_row + 1
        actual_end_row = sheet.max_row if end_row == -1 else end_row + 1

        # 提取指定行范围的数据
        for row in sheet.iter_rows(min_row=actual_start_row, max_row=actual_end_row, values_only=True):
            extracted_data.append(list(row))

        result[sheet_name] = extracted_data

    return result


def extract_by_condition(workbook, options):
    """
    按条件提取内容
    :param workbook: Excel工作簿对象
    :param options: 提取选项，包含：
        - sheet_names: 要处理的工作表名称列表
        - condition_column: 条件列索引（从0开始）
        - condition_type: 条件类型（eq, ne, gt, lt, ge, le, contains）
        - condition_value: 条件值
    :return: 提取结果
    """
    sheet_names = options.get('sheet_names', [])
    condition_column = options.get('condition_column', 0)
    condition_type = options.get('condition_type', 'eq')
    condition_value = options.get('condition_value', '')
    result = {}

    # 如果没有指定工作表，处理所有工作表
    if not sheet_names:
        sheet_names = workbook.sheetnames

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        extracted_data = []

        # 获取表头
        header = []
        for col in range(1, sheet.max_column + 1):
            header_cell = sheet.cell(row=1, column=col)
            header.append(header_cell.value or f'列{col}')
        extracted_data.append(header)

        # 提取符合条件的数据行
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if len(row) <= condition_column:
                continue

            cell_value = row[condition_column]
            if check_condition(cell_value, condition_value, condition_type):
                extracted_data.append(list(row))

        result[sheet_name] = extracted_data

    return result


def check_condition(cell_value, condition_value, condition_type):
    """
    检查单元格值是否符合条件
    :param cell_value: 单元格值
    :param condition_value: 条件值
    :param condition_type: 条件类型
    :return: 是否符合条件
    """
    try:
        # 尝试将字符串转换为数值进行比较
        if isinstance(cell_value, (int, float)) and isinstance(condition_value, str):
            condition_value = float(condition_value)
        elif isinstance(cell_value, str) and isinstance(condition_value, (int, float)):
            cell_value = float(cell_value)
    except (ValueError, TypeError):
        pass

    if condition_type == 'eq':
        return cell_value == condition_value
    elif condition_type == 'ne':
        return cell_value != condition_value
    elif condition_type == 'gt':
        return cell_value > condition_value
    elif condition_type == 'lt':
        return cell_value < condition_value
    elif condition_type == 'ge':
        return cell_value >= condition_value
    elif condition_type == 'le':
        return cell_value <= condition_value
    elif condition_type == 'contains':
        return str(condition_value) in str(cell_value)
    else:
        return False


# 主函数，处理命令行参数
# 注意：在Pyodide环境中，__name__ != '__main__'，所以这段测试代码不会执行
if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': '缺少参数'}), flush=True)
        sys.exit(1)

    try:
        # 解析参数
        params = json.loads(sys.argv[1])
        file_data = params['file_data'].encode('latin1')  # 从base64解码后的数据
        extract_type = params['extract_type']
        options = params['options']

        # 调用提取函数
        result = extract_content_from_excel(file_data, extract_type, options)
        print(json.dumps(result, ensure_ascii=False), flush=True)
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False), flush=True)
        sys.exit(1)

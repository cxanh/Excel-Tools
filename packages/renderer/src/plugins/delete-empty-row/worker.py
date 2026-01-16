import sys
import os
from openpyxl import load_workbook
from io import BytesIO

# 设置标准输出为UTF-8
sys.stdout.reconfigure(encoding='utf-8')


def delete_empty_content(file_data, options):
    """
    删除Excel文件中的空白内容
    :param file_data: Excel文件的二进制数据
    :param options: 删除选项，包含：
        - delete_empty_sheets: 是否删除空白工作表
        - delete_empty_rows: 是否删除空白行
        - delete_empty_cols: 是否删除空白列
        - sheet_names: 要处理的工作表名称列表
    :return: 处理结果
    """
    try:
        # 加载Excel文件
        workbook = load_workbook(filename=BytesIO(file_data))
        results = {
            'deleted_sheets': [],
            'deleted_rows': {},
            'deleted_cols': {}
        }

        # 获取要处理的工作表
        sheet_names = options.get('sheet_names', [])
        if not sheet_names:
            sheet_names = workbook.sheetnames.copy()  # 复制一份，避免修改时出现问题

        # 1. 删除空白工作表
        if options.get('delete_empty_sheets', False):
            deleted_sheets = delete_empty_worksheets(workbook, sheet_names)
            results['deleted_sheets'] = deleted_sheets
            # 从要处理的工作表列表中移除已删除的工作表
            for sheet in deleted_sheets:
                if sheet in sheet_names:
                    sheet_names.remove(sheet)

        # 2. 处理每个工作表的空白行和空白列
        for sheet_name in sheet_names:
            if sheet_name not in workbook.sheetnames:
                continue

            sheet = workbook[sheet_name]
            deleted_rows = 0
            deleted_cols = 0

            # 删除空白行
            if options.get('delete_empty_rows', False):
                deleted_rows = delete_empty_rows_from_sheet(sheet)
                results['deleted_rows'][sheet_name] = deleted_rows

            # 删除空白列
            if options.get('delete_empty_cols', False):
                deleted_cols = delete_empty_cols_from_sheet(sheet)
                results['deleted_cols'][sheet_name] = deleted_cols

        # 保存修改后的Excel文件
        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return {
            'success': True,
            'data': {
                'file_data': output.getvalue().hex(),
                'results': results
            }
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }


def delete_empty_worksheets(workbook, sheet_names):
    """
    删除空白工作表
    :param workbook: Excel工作簿对象
    :param sheet_names: 要检查的工作表名称列表
    :return: 已删除的工作表名称列表
    """
    deleted_sheets = []

    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            continue

        sheet = workbook[sheet_name]
        is_empty = True

        # 检查工作表是否有数据
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                if cell_value is not None and cell_value != '':
                    is_empty = False
                    break
            if not is_empty:
                break

        if is_empty:
            deleted_sheets.append(sheet_name)
            workbook.remove(sheet)

    return deleted_sheets


def delete_empty_rows_from_sheet(sheet):
    """
    删除工作表中的空白行
    :param sheet: Excel工作表对象
    :return: 删除的行数
    """
    deleted_rows = 0
    rows_to_delete = []

    # 找出所有空白行
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        is_empty = True
        for cell_value in row:
            if cell_value is not None and cell_value != '':
                is_empty = False
                break
        if is_empty:
            rows_to_delete.append(row_idx)

    # 从下往上删除，避免索引混乱
    for row_idx in reversed(rows_to_delete):
        sheet.delete_rows(row_idx)
        deleted_rows += 1

    return deleted_rows


def delete_empty_cols_from_sheet(sheet):
    """
    删除工作表中的空白列
    :param sheet: Excel工作表对象
    :return: 删除的列数
    """
    deleted_cols = 0
    cols_to_delete = []
    max_col = sheet.max_column

    # 找出所有空白列
    for col_idx in range(1, max_col + 1):
        is_empty = True
        for row in sheet.iter_rows(min_col=col_idx, max_col=col_idx, values_only=True):
            cell_value = row[0]
            if cell_value is not None and cell_value != '':
                is_empty = False
                break
        if is_empty:
            cols_to_delete.append(col_idx)

    # 从右往左删除，避免索引混乱
    for col_idx in reversed(cols_to_delete):
        sheet.delete_cols(col_idx)
        deleted_cols += 1

    return deleted_cols


# 主函数，处理命令行参数
# 注意：在Pyodide环境中，__name__ != '__main__'，所以这段测试代码不会执行
if __name__ == '__main__':
    import sys
    import json

    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': '缺少参数'}), flush=True)
        sys.exit(1)

    try:
        # 解析参数
        params = json.loads(sys.argv[1])
        file_data = bytes.fromhex(params['file_data'])
        options = params['options']

        # 调用处理函数
        result = delete_empty_content(file_data, options)
        print(json.dumps(result, ensure_ascii=False), flush=True)
    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}, ensure_ascii=False), flush=True)
        sys.exit(1)

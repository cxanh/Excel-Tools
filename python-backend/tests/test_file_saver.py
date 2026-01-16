#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Saver 单元测试
"""

import sys
import os
import pytest
import time
from openpyxl import Workbook

# 添加父目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from engine.core.saver import FileSaver
from engine.core.loader import FileLoader


class TestFileSaver:
    """File Saver 测试类"""
    
    def setup_method(self):
        """每个测试方法前执行"""
        self.saver = FileSaver()
        self.loader = FileLoader()
        self.test_files_dir = os.path.join(os.path.dirname(__file__), 'test_files')
        
        # 创建测试文件目录
        os.makedirs(self.test_files_dir, exist_ok=True)
    
    def teardown_method(self):
        """每个测试方法后执行"""
        # 清理测试文件
        pass
    
    def _create_test_workbook(self) -> Workbook:
        """创建测试用的工作簿"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws['A1'] = "测试数据"
        ws['B1'] = 123
        return wb
    
    def test_save_new_file(self):
        """测试保存新文件"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_save_new.xlsx")
        
        # 删除文件（如果存在）
        if os.path.exists(file_path):
            os.remove(file_path)
        
        result = self.saver.save({
            "workbook": wb,
            "file_path": file_path
        })
        
        assert result["type"] == "result"
        assert result["status"] == "success"
        assert result["message"] == "文件保存成功"
        assert os.path.exists(file_path)
        
        # 验证文件信息
        data = result["data"]
        assert data["file_name"] == "test_save_new.xlsx"
        assert data["file_size"] > 0
        assert "backup_created" not in data  # 新文件不应创建备份
    
    def test_save_existing_file_with_backup(self):
        """测试保存已存在的文件（创建备份）"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_save_existing.xlsx")
        
        # 先保存一次
        wb.save(file_path)
        time.sleep(0.1)  # 确保时间戳不同
        
        # 再次保存（应该创建备份）
        result = self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": True,
            "create_backup": True
        })
        
        assert result["status"] == "success"
        assert result["data"]["backup_created"] is True
        assert "backup_path" in result["data"]
        
        # 验证备份文件存在
        backup_path = result["data"]["backup_path"]
        assert os.path.exists(backup_path)
    
    def test_save_without_backup(self):
        """测试保存文件（不创建备份）"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_save_no_backup.xlsx")
        
        # 先保存一次
        wb.save(file_path)
        
        # 再次保存（不创建备份）
        result = self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": True,
            "create_backup": False
        })
        
        assert result["status"] == "success"
        assert "backup_created" not in result["data"]
    
    def test_save_file_exists_no_overwrite(self):
        """测试保存文件（文件已存在且不覆盖）"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_save_exists.xlsx")
        
        # 先保存一次
        wb.save(file_path)
        
        # 尝试再次保存（不覆盖）
        result = self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": False
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "FILE_EXISTS"
    
    def test_save_missing_workbook(self):
        """测试保存文件（缺少工作簿）"""
        result = self.saver.save({
            "file_path": "test.xlsx"
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "MISSING_WORKBOOK"
    
    def test_save_missing_file_path(self):
        """测试保存文件（缺少文件路径）"""
        wb = self._create_test_workbook()
        
        result = self.saver.save({
            "workbook": wb
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "MISSING_FILE_PATH"
    
    def test_list_backups(self):
        """测试列出备份"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_list_backups.xlsx")
        
        # 清理旧的备份目录
        file_base = os.path.splitext(os.path.basename(file_path))[0]
        backup_dir = os.path.join(self.test_files_dir, f".{file_base}_backups")
        if os.path.exists(backup_dir):
            import shutil
            shutil.rmtree(backup_dir)
        
        # 保存文件多次以创建多个备份
        wb.save(file_path)
        
        for i in range(3):
            time.sleep(0.1)  # 确保时间戳不同
            self.saver.save({
                "workbook": wb,
                "file_path": file_path,
                "overwrite": True,
                "create_backup": True
            })
        
        # 列出备份
        result = self.saver.list_backups({
            "file_path": file_path
        })
        
        assert result["status"] == "success"
        assert result["data"]["backup_count"] == 3
        assert len(result["data"]["backups"]) == 3
        
        # 验证备份信息
        backup = result["data"]["backups"][0]
        assert "path" in backup
        assert "name" in backup
        assert "time" in backup
        assert "size" in backup
    
    def test_restore_from_backup(self):
        """测试从备份恢复"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_restore.xlsx")
        
        # 保存原始文件
        wb.save(file_path)
        original_size = os.path.getsize(file_path)
        
        time.sleep(0.1)
        
        # 修改并保存（创建备份）
        ws = wb.active
        ws['C1'] = "新数据"
        self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": True,
            "create_backup": True
        })
        
        modified_size = os.path.getsize(file_path)
        
        # 从备份恢复
        result = self.saver.restore({
            "file_path": file_path,
            "backup_index": 0
        })
        
        assert result["status"] == "success"
        assert result["message"] == "文件恢复成功"
        
        # 验证文件大小恢复到原始大小
        restored_size = os.path.getsize(file_path)
        assert restored_size == original_size
    
    def test_restore_no_backups(self):
        """测试恢复文件（没有备份）"""
        result = self.saver.restore({
            "file_path": "nonexistent.xlsx"
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "NO_BACKUPS"
    
    def test_restore_invalid_index(self):
        """测试恢复文件（无效的备份索引）"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_restore_invalid.xlsx")
        
        # 保存文件并创建一个备份
        wb.save(file_path)
        time.sleep(0.1)
        self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": True,
            "create_backup": True
        })
        
        # 尝试使用无效的索引
        result = self.saver.restore({
            "file_path": file_path,
            "backup_index": 10
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "INVALID_BACKUP_INDEX"
    
    def test_delete_backup(self):
        """测试删除备份"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_delete_backup.xlsx")
        
        # 保存文件并创建备份
        wb.save(file_path)
        time.sleep(0.1)
        save_result = self.saver.save({
            "workbook": wb,
            "file_path": file_path,
            "overwrite": True,
            "create_backup": True
        })
        
        backup_path = save_result["data"]["backup_path"]
        
        # 删除备份
        result = self.saver.delete_backup({
            "backup_path": backup_path
        })
        
        assert result["status"] == "success"
        assert not os.path.exists(backup_path)
    
    def test_cleanup_old_backups(self):
        """测试清理旧备份（保留最近 5 个）"""
        wb = self._create_test_workbook()
        file_path = os.path.join(self.test_files_dir, "test_cleanup.xlsx")
        
        # 清理旧的备份目录
        file_base = os.path.splitext(os.path.basename(file_path))[0]
        backup_dir = os.path.join(self.test_files_dir, f".{file_base}_backups")
        if os.path.exists(backup_dir):
            import shutil
            shutil.rmtree(backup_dir)
        
        # 保存文件
        wb.save(file_path)
        
        # 创建 7 个备份
        for i in range(7):
            time.sleep(0.05)  # 减少等待时间
            self.saver.save({
                "workbook": wb,
                "file_path": file_path,
                "overwrite": True,
                "create_backup": True
            })
        
        # 列出备份
        result = self.saver.list_backups({
            "file_path": file_path
        })
        
        # 应该只保留 5 个备份
        assert result["data"]["backup_count"] == 5


class TestFileSaverIntegration:
    """File Saver 集成测试（通过 CLI Router）"""
    
    def setup_method(self):
        """每个测试方法前执行"""
        from cli_router import CLIRouter
        self.router = CLIRouter()
        self.test_files_dir = os.path.join(os.path.dirname(__file__), 'test_files')
        os.makedirs(self.test_files_dir, exist_ok=True)
    
    def _create_test_xlsx(self, filename: str) -> str:
        """创建测试用的 .xlsx 文件"""
        file_path = os.path.join(self.test_files_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws['A1'] = "测试数据"
        wb.save(file_path)
        wb.close()
        
        return file_path
    
    def test_save_file_via_router(self):
        """测试通过路由器保存文件"""
        # 先创建并加载文件
        file_path = self._create_test_xlsx("test_router_save.xlsx")
        
        self.router.route({
            "action": "load_file",
            "params": {"file_path": file_path}
        })
        
        # 保存文件
        result = self.router.route({
            "action": "save_file",
            "params": {
                "file_path": file_path,
                "overwrite": True
            }
        })
        
        assert result["type"] == "result"
        assert result["status"] == "success"
    
    def test_save_file_no_loaded(self):
        """测试保存文件（没有加载文件）"""
        result = self.router.route({
            "action": "save_file",
            "params": {
                "file_path": "test.xlsx"
            }
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "NO_FILE_LOADED"
    
    def test_list_backups_via_router(self):
        """测试通过路由器列出备份"""
        file_path = self._create_test_xlsx("test_router_list.xlsx")
        
        result = self.router.route({
            "action": "list_backups",
            "params": {"file_path": file_path}
        })
        
        assert result["type"] == "result"
        assert result["status"] == "success"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

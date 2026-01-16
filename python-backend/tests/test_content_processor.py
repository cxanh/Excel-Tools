#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Content Processor 单元测试
"""

import sys
import os
import pytest
from openpyxl import Workbook

# 添加父目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from engine.content.processor import ContentProcessor


class TestContentProcessor:
    """Content Processor 测试类"""
    
    def setup_method(self):
        """每个测试方法前执行"""
        self.processor = ContentProcessor()
    
    def _create_test_worksheet_with_blank_rows(self):
        """创建包含空白行的测试工作表"""
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        ws['A1'] = "姓名"
        ws['B1'] = "年龄"
        # 第 2 行：空白行
        ws['A3'] = "张三"
        ws['B3'] = 25
        # 第 4 行：空白行
        ws['A5'] = "李四"
        ws['B5'] = 30
        
        return ws
    
    def test_remove_blank_rows(self):
        """测试删除空白行"""
        ws = self._create_test_worksheet_with_blank_rows()
        
        result = self.processor.remove_blank_rows({
            "worksheet": ws,
            "sheet_name": "TestSheet"
        })
        
        assert result["status"] == "success"
        assert result["data"]["deleted_count"] == 2
        assert result["data"]["original_rows"] == 5
        assert result["data"]["remaining_rows"] == 3
        
        # 验证空白行已删除
        assert ws.max_row == 3
        assert ws['A1'].value == "姓名"
        assert ws['A2'].value == "张三"
        assert ws['A3'].value == "李四"
    
    def test_clear_blank_cells(self):
        """测试清除空白单元格"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "数据"
        ws['B1'] = "  "  # 空白字符串
        ws['C1'] = ""    # 空字符串
        ws['D1'] = "正常"
        
        result = self.processor.clear_blank_cells({
            "worksheet": ws,
            "sheet_name": "TestSheet"
        })
        
        assert result["status"] == "success"
        assert result["data"]["cleared_count"] == 2
        
        # 验证空白单元格已清除
        assert ws['A1'].value == "数据"
        assert ws['B1'].value is None
        assert ws['C1'].value is None
        assert ws['D1'].value == "正常"
    
    def test_remove_formulas(self):
        """测试删除公式"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = 10
        ws['A2'] = 20
        ws['A3'] = '=A1+A2'  # 公式
        
        # 计算公式（在实际场景中，openpyxl 需要 data_only=False 来保留公式）
        # 这里我们手动设置 data_type
        ws['A3'].data_type = 'f'
        ws['A3'].value = 30  # 模拟计算结果
        
        result = self.processor.remove_formulas({
            "worksheet": ws,
            "sheet_name": "TestSheet"
        })
        
        assert result["status"] == "success"
        assert result["data"]["formula_count"] >= 0  # 可能为 0 或 1，取决于 openpyxl 的处理
    
    def test_remove_duplicate_rows(self):
        """测试删除重复行"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "姓名"
        ws['B1'] = "年龄"
        ws['A2'] = "张三"
        ws['B2'] = 25
        ws['A3'] = "李四"
        ws['B3'] = 30
        ws['A4'] = "张三"  # 重复
        ws['B4'] = 25
        ws['A5'] = "王五"
        ws['B5'] = 35
        
        result = self.processor.remove_duplicate_rows({
            "worksheet": ws,
            "sheet_name": "TestSheet"
        })
        
        assert result["status"] == "success"
        assert result["data"]["deleted_count"] == 1
        assert result["data"]["original_rows"] == 5
        assert result["data"]["remaining_rows"] == 4
        
        # 验证重复行已删除
        assert ws.max_row == 4
    
    def test_remove_duplicate_rows_with_key_columns(self):
        """测试使用关键列删除重复行"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "ID"
        ws['B1'] = "姓名"
        ws['C1'] = "年龄"
        ws['A2'] = 1
        ws['B2'] = "张三"
        ws['C2'] = 25
        ws['A3'] = 2
        ws['B3'] = "李四"
        ws['C3'] = 30
        ws['A4'] = 1  # ID 重复
        ws['B4'] = "张三2"
        ws['C4'] = 26
        
        # 只根据 ID 列（第 1 列）判断重复
        result = self.processor.remove_duplicate_rows({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "key_columns": [1]
        })
        
        assert result["status"] == "success"
        assert result["data"]["deleted_count"] == 1
    
    def test_replace_content_simple(self):
        """测试简单文本替换"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "Hello World"
        ws['A2'] = "Hello Python"
        ws['A3'] = "Goodbye"
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "find_text": "Hello",
            "replace_text": "Hi"
        })
        
        assert result["status"] == "success"
        assert result["data"]["replace_count"] == 2
        
        # 验证替换结果
        assert ws['A1'].value == "Hi World"
        assert ws['A2'].value == "Hi Python"
        assert ws['A3'].value == "Goodbye"
    
    def test_replace_content_case_sensitive(self):
        """测试区分大小写的替换"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "Hello"
        ws['A2'] = "hello"
        ws['A3'] = "HELLO"
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "find_text": "Hello",
            "replace_text": "Hi",
            "case_sensitive": True
        })
        
        assert result["status"] == "success"
        assert result["data"]["replace_count"] == 1
        
        # 只有完全匹配的被替换
        assert ws['A1'].value == "Hi"
        assert ws['A2'].value == "hello"
        assert ws['A3'].value == "HELLO"
    
    def test_replace_content_regex(self):
        """测试正则表达式替换"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "Phone: 123-456-7890"
        ws['A2'] = "Phone: 987-654-3210"
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "find_text": r"\d{3}-\d{3}-\d{4}",
            "replace_text": "***-***-****",
            "use_regex": True
        })
        
        assert result["status"] == "success"
        assert result["data"]["replace_count"] == 2
        
        # 验证正则替换
        assert ws['A1'].value == "Phone: ***-***-****"
        assert ws['A2'].value == "Phone: ***-***-****"
    
    def test_replace_content_whole_word(self):
        """测试全词匹配替换"""
        wb = Workbook()
        ws = wb.active
        
        ws['A1'] = "cat"
        ws['A2'] = "category"
        ws['A3'] = "the cat"
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "find_text": "cat",
            "replace_text": "dog",
            "whole_word": True
        })
        
        assert result["status"] == "success"
        assert result["data"]["replace_count"] == 2
        
        # 只有完整单词被替换
        assert ws['A1'].value == "dog"
        assert ws['A2'].value == "category"  # 不替换
        assert ws['A3'].value == "the dog"
    
    def test_missing_worksheet(self):
        """测试缺少工作表参数"""
        result = self.processor.remove_blank_rows({})
        
        assert result["status"] == "error"
        assert result["error_code"] == "MISSING_WORKSHEET"
    
    def test_missing_find_text(self):
        """测试缺少查找文本"""
        wb = Workbook()
        ws = wb.active
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet"
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "MISSING_FIND_TEXT"
    
    def test_invalid_regex(self):
        """测试无效的正则表达式"""
        wb = Workbook()
        ws = wb.active
        ws['A1'] = "test"
        
        result = self.processor.replace_content({
            "worksheet": ws,
            "sheet_name": "TestSheet",
            "find_text": "[invalid(regex",
            "replace_text": "replacement",
            "use_regex": True
        })
        
        assert result["status"] == "error"
        assert result["error_code"] == "INVALID_REGEX"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

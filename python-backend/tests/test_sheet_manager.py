#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sheet Manager Tests - 工作表管理器测试
"""

import os
import sys
import pytest
from openpyxl import Workbook

# 添加项目根目录到 Python 路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from engine.sheet.manager import SheetManager


class TestSheetManager:
    """工作表管理器测试类"""
    
    @pytest.fixture
    def manager(self):
        """创建工作表管理器实例"""
        return SheetManager()
    
    @pytest.fixture
    def workbook(self):
        """创建测试工作簿"""
        wb = Workbook()
        # 默认有一个 Sheet
        return wb
    
    # ========== 插入工作表测试 ==========
    
    def test_insert_sheet_missing_workbook(self, manager):
        """测试插入工作表 - 缺少工作簿"""
        result = manager.insert_sheet({})
        
        assert result['type'] == 'result'
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_WORKBOOK'
    
    def test_insert_sheet_default_name(self, manager, workbook):
        """测试插入工作表 - 使用默认名称"""
        result = manager.insert_sheet({'workbook': workbook})
        
        assert result['type'] == 'result'
        assert result['status'] == 'success'
        assert 'Sheet' in result['data']['sheet_name']
        assert result['data']['total_sheets'] == 2
    
    def test_insert_sheet_custom_name(self, manager, workbook):
        """测试插入工作表 - 自定义名称"""
        result = manager.insert_sheet({
            'workbook': workbook,
            'sheet_name': 'CustomSheet'
        })
        
        assert result['status'] == 'success'
        assert result['data']['sheet_name'] == 'CustomSheet'
        assert 'CustomSheet' in workbook.sheetnames
    
    def test_insert_sheet_at_index(self, manager, workbook):
        """测试插入工作表 - 指定位置"""
        result = manager.insert_sheet({
            'workbook': workbook,
            'sheet_name': 'FirstSheet',
            'index': 0
        })
        
        assert result['status'] == 'success'
        assert workbook.sheetnames[0] == 'FirstSheet'
        assert result['data']['index'] == 0
    
    def test_insert_sheet_duplicate_name(self, manager, workbook):
        """测试插入工作表 - 名称重复"""
        # 先插入一个
        manager.insert_sheet({
            'workbook': workbook,
            'sheet_name': 'TestSheet'
        })
        
        # 再次插入相同名称
        result = manager.insert_sheet({
            'workbook': workbook,
            'sheet_name': 'TestSheet'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'SHEET_NAME_EXISTS'
    
    def test_insert_sheet_name_too_long(self, manager, workbook):
        """测试插入工作表 - 名称过长"""
        long_name = 'A' * 32  # 超过 31 个字符
        result = manager.insert_sheet({
            'workbook': workbook,
            'sheet_name': long_name
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_SHEET_NAME'
    
    def test_insert_sheet_invalid_chars(self, manager, workbook):
        """测试插入工作表 - 包含非法字符"""
        invalid_names = ['Test\\Sheet', 'Test/Sheet', 'Test*Sheet', 
                        'Test?Sheet', 'Test:Sheet', 'Test[Sheet]']
        
        for name in invalid_names:
            result = manager.insert_sheet({
                'workbook': workbook,
                'sheet_name': name
            })
            
            assert result['status'] == 'error'
            assert result['error_code'] == 'INVALID_SHEET_NAME'
    
    # ========== 删除工作表测试 ==========
    
    def test_delete_sheet_missing_workbook(self, manager):
        """测试删除工作表 - 缺少工作簿"""
        result = manager.delete_sheet({})
        
        assert result['type'] == 'result'
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_WORKBOOK'
    
    def test_delete_sheet_missing_name(self, manager, workbook):
        """测试删除工作表 - 缺少名称"""
        result = manager.delete_sheet({'workbook': workbook})
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_SHEET_NAME'
    
    def test_delete_sheet_not_found(self, manager, workbook):
        """测试删除工作表 - 工作表不存在"""
        result = manager.delete_sheet({
            'workbook': workbook,
            'sheet_name': 'NonExistent'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'SHEET_NOT_FOUND'
    
    def test_delete_sheet_last_sheet(self, manager, workbook):
        """测试删除工作表 - 不能删除唯一的工作表"""
        # 工作簿默认只有一个工作表
        result = manager.delete_sheet({
            'workbook': workbook,
            'sheet_name': workbook.sheetnames[0]
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'CANNOT_DELETE_LAST_SHEET'
    
    def test_delete_sheet_success(self, manager, workbook):
        """测试成功删除工作表"""
        # 先添加一个工作表
        workbook.create_sheet('ToDelete')
        initial_count = len(workbook.sheetnames)
        
        # 删除工作表
        result = manager.delete_sheet({
            'workbook': workbook,
            'sheet_name': 'ToDelete'
        })
        
        assert result['status'] == 'success'
        assert result['data']['deleted_sheet'] == 'ToDelete'
        assert result['data']['remaining_sheets'] == initial_count - 1
        assert 'ToDelete' not in workbook.sheetnames
    
    # ========== 重命名工作表测试 ==========
    
    def test_rename_sheet_missing_workbook(self, manager):
        """测试重命名工作表 - 缺少工作簿"""
        result = manager.rename_sheet({})
        
        assert result['type'] == 'result'
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_WORKBOOK'
    
    def test_rename_sheet_missing_old_name(self, manager, workbook):
        """测试重命名工作表 - 缺少原名称"""
        result = manager.rename_sheet({
            'workbook': workbook,
            'new_name': 'NewName'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_OLD_NAME'
    
    def test_rename_sheet_missing_new_name(self, manager, workbook):
        """测试重命名工作表 - 缺少新名称"""
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': 'Sheet'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_NEW_NAME'
    
    def test_rename_sheet_not_found(self, manager, workbook):
        """测试重命名工作表 - 工作表不存在"""
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': 'NonExistent',
            'new_name': 'NewName'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'SHEET_NOT_FOUND'
    
    def test_rename_sheet_duplicate_name(self, manager, workbook):
        """测试重命名工作表 - 新名称已存在"""
        # 添加两个工作表
        workbook.create_sheet('Sheet1')
        workbook.create_sheet('Sheet2')
        
        # 尝试将 Sheet1 重命名为 Sheet2
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': 'Sheet1',
            'new_name': 'Sheet2'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'SHEET_NAME_EXISTS'
    
    def test_rename_sheet_same_name(self, manager, workbook):
        """测试重命名工作表 - 重命名为相同名称"""
        old_name = workbook.sheetnames[0]
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': old_name,
            'new_name': old_name
        })
        
        assert result['status'] == 'success'
    
    def test_rename_sheet_success(self, manager, workbook):
        """测试成功重命名工作表"""
        old_name = workbook.sheetnames[0]
        new_name = 'RenamedSheet'
        
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': old_name,
            'new_name': new_name
        })
        
        assert result['status'] == 'success'
        assert result['data']['old_name'] == old_name
        assert result['data']['new_name'] == new_name
        assert new_name in workbook.sheetnames
        assert old_name not in workbook.sheetnames
    
    def test_rename_sheet_invalid_name(self, manager, workbook):
        """测试重命名工作表 - 无效名称"""
        old_name = workbook.sheetnames[0]
        
        # 测试名称过长
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': old_name,
            'new_name': 'A' * 32
        })
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_SHEET_NAME'
        
        # 测试非法字符
        result = manager.rename_sheet({
            'workbook': workbook,
            'old_name': old_name,
            'new_name': 'Test/Sheet'
        })
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_SHEET_NAME'
    
    # ========== 获取工作表信息测试 ==========
    
    def test_get_sheet_info_missing_workbook(self, manager):
        """测试获取工作表信息 - 缺少工作簿"""
        result = manager.get_sheet_info({})
        
        assert result['type'] == 'result'
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_WORKBOOK'
    
    def test_get_sheet_info_success(self, manager, workbook):
        """测试成功获取工作表信息"""
        # 添加几个工作表
        workbook.create_sheet('Sheet2')
        workbook.create_sheet('Sheet3')
        
        result = manager.get_sheet_info({'workbook': workbook})
        
        assert result['status'] == 'success'
        assert result['data']['total_sheets'] == 3
        assert len(result['data']['sheets']) == 3
        
        # 检查第一个工作表的信息
        first_sheet = result['data']['sheets'][0]
        assert 'name' in first_sheet
        assert 'index' in first_sheet
        assert 'max_row' in first_sheet
        assert 'max_column' in first_sheet
        assert 'visible' in first_sheet


if __name__ == '__main__':
    pytest.main([__file__, '-v'])

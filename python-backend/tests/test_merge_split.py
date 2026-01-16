#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试合并拆分引擎
"""

import os
import sys
import pytest
import openpyxl
import pandas as pd
import tempfile
import shutil

# 添加项目根目录到 Python 路径
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from engine.merge_split.engine import MergeSplitEngine


class TestMergeSplitEngine:
    """测试合并拆分引擎"""
    
    @pytest.fixture
    def engine(self):
        """创建引擎实例"""
        return MergeSplitEngine()
    
    @pytest.fixture
    def temp_dir(self):
        """创建临时目录"""
        temp_dir = tempfile.mkdtemp()
        yield temp_dir
        # 清理
        shutil.rmtree(temp_dir, ignore_errors=True)
    
    @pytest.fixture
    def sample_excel_files(self, temp_dir):
        """创建示例 Excel 文件"""
        files = []
        
        for i in range(3):
            file_path = os.path.join(temp_dir, f"test{i + 1}.xlsx")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Sheet1"
            
            # 添加表头
            ws.append(["Name", "Age", "City"])
            
            # 添加数据
            for j in range(5):
                ws.append([f"Person{i * 5 + j + 1}", 20 + j, f"City{i + 1}"])
            
            wb.save(file_path)
            wb.close()
            files.append(file_path)
        
        return files
    
    @pytest.fixture
    def sample_csv_files(self, temp_dir):
        """创建示例 CSV 文件"""
        files = []
        
        for i in range(3):
            file_path = os.path.join(temp_dir, f"test{i + 1}.csv")
            df = pd.DataFrame({
                "Name": [f"Person{i * 5 + j + 1}" for j in range(5)],
                "Age": [20 + j for j in range(5)],
                "City": [f"City{i + 1}" for _ in range(5)]
            })
            df.to_csv(file_path, index=False)
            files.append(file_path)
        
        return files
    
    # ==================== Excel 合并测试 ====================
    
    def test_merge_excel_missing_file_paths(self, engine, temp_dir):
        """测试缺少文件路径"""
        output_path = os.path.join(temp_dir, "merged.xlsx")
        result = engine.merge_excel_files({
            'output_path': output_path,
            'mode': 'append'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_FILE_PATHS'
    
    def test_merge_excel_missing_output_path(self, engine, sample_excel_files):
        """测试缺少输出路径"""
        result = engine.merge_excel_files({
            'file_paths': sample_excel_files,
            'mode': 'append'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_OUTPUT_PATH'
    
    def test_merge_excel_invalid_mode(self, engine, sample_excel_files, temp_dir):
        """测试无效的合并模式"""
        output_path = os.path.join(temp_dir, "merged.xlsx")
        result = engine.merge_excel_files({
            'file_paths': sample_excel_files,
            'output_path': output_path,
            'mode': 'invalid'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_MODE'
    
    def test_merge_excel_file_not_found(self, engine, temp_dir):
        """测试文件不存在"""
        output_path = os.path.join(temp_dir, "merged.xlsx")
        result = engine.merge_excel_files({
            'file_paths': [os.path.join(temp_dir, "nonexistent.xlsx")],
            'output_path': output_path,
            'mode': 'append'
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'FILE_NOT_FOUND'
    
    def test_merge_excel_append_mode(self, engine, sample_excel_files, temp_dir):
        """测试追加模式合并 Excel"""
        output_path = os.path.join(temp_dir, "merged.xlsx")
        result = engine.merge_excel_files({
            'file_paths': sample_excel_files,
            'output_path': output_path,
            'mode': 'append',
            'has_header': True,
            'sheet_name': 'Merged'
        })
        
        assert result['status'] == 'success'
        assert result['data']['total_files'] == 3
        assert result['data']['total_rows'] == 15  # 3 files * 5 rows
        assert result['data']['mode'] == 'append'
        assert os.path.exists(output_path)
        
        # 验证合并结果
        df = pd.read_excel(output_path, sheet_name='Merged')
        assert len(df) == 15
        assert list(df.columns) == ["Name", "Age", "City"]
    
    def test_merge_excel_separate_mode(self, engine, sample_excel_files, temp_dir):
        """测试独立工作表模式合并 Excel"""
        output_path = os.path.join(temp_dir, "merged.xlsx")
        result = engine.merge_excel_files({
            'file_paths': sample_excel_files,
            'output_path': output_path,
            'mode': 'separate'
        })
        
        assert result['status'] == 'success'
        assert result['data']['total_files'] == 3
        assert result['data']['total_sheets'] == 3
        assert result['data']['mode'] == 'separate'
        assert os.path.exists(output_path)
        
        # 验证合并结果
        wb = openpyxl.load_workbook(output_path)
        assert len(wb.sheetnames) == 3
        assert 'test1_Sheet1' in wb.sheetnames
        assert 'test2_Sheet1' in wb.sheetnames
        assert 'test3_Sheet1' in wb.sheetnames
        wb.close()
    
    # ==================== CSV 合并测试 ====================
    
    def test_merge_csv_missing_file_paths(self, engine, temp_dir):
        """测试缺少文件路径"""
        output_path = os.path.join(temp_dir, "merged.csv")
        result = engine.merge_csv_files({
            'output_path': output_path
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_FILE_PATHS'
    
    def test_merge_csv_missing_output_path(self, engine, sample_csv_files):
        """测试缺少输出路径"""
        result = engine.merge_csv_files({
            'file_paths': sample_csv_files
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_OUTPUT_PATH'
    
    def test_merge_csv_file_not_found(self, engine, temp_dir):
        """测试文件不存在"""
        output_path = os.path.join(temp_dir, "merged.csv")
        result = engine.merge_csv_files({
            'file_paths': [os.path.join(temp_dir, "nonexistent.csv")],
            'output_path': output_path
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'FILE_NOT_FOUND'
    
    def test_merge_csv_success(self, engine, sample_csv_files, temp_dir):
        """测试成功合并 CSV"""
        output_path = os.path.join(temp_dir, "merged.csv")
        result = engine.merge_csv_files({
            'file_paths': sample_csv_files,
            'output_path': output_path,
            'has_header': True
        })
        
        assert result['status'] == 'success'
        assert result['data']['total_files'] == 3
        assert result['data']['total_rows'] == 15
        assert os.path.exists(output_path)
        
        # 验证合并结果
        df = pd.read_csv(output_path)
        assert len(df) == 15
        assert list(df.columns) == ["Name", "Age", "City"]
    
    # ==================== Excel 拆分测试 ====================
    
    def test_split_excel_missing_file_path(self, engine, temp_dir):
        """测试缺少文件路径"""
        result = engine.split_excel_file({
            'output_dir': temp_dir,
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_FILE_PATH'
    
    def test_split_excel_missing_output_dir(self, engine, sample_excel_files):
        """测试缺少输出目录"""
        result = engine.split_excel_file({
            'file_path': sample_excel_files[0],
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_OUTPUT_DIR'
    
    def test_split_excel_invalid_rows_per_file(self, engine, sample_excel_files, temp_dir):
        """测试无效的每文件行数"""
        result = engine.split_excel_file({
            'file_path': sample_excel_files[0],
            'output_dir': temp_dir,
            'rows_per_file': 0
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_ROWS_PER_FILE'
    
    def test_split_excel_file_not_found(self, engine, temp_dir):
        """测试文件不存在"""
        output_dir = os.path.join(temp_dir, "output")
        result = engine.split_excel_file({
            'file_path': os.path.join(temp_dir, "nonexistent.xlsx"),
            'output_dir': output_dir,
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'FILE_NOT_FOUND'
    
    def test_split_excel_success(self, engine, sample_excel_files, temp_dir):
        """测试成功拆分 Excel"""
        output_dir = os.path.join(temp_dir, "output")
        result = engine.split_excel_file({
            'file_path': sample_excel_files[0],
            'output_dir': output_dir,
            'rows_per_file': 2,
            'has_header': True
        })
        
        assert result['status'] == 'success'
        assert result['data']['total_rows'] == 5
        assert result['data']['rows_per_file'] == 2
        assert result['data']['num_files'] == 3  # 5 rows / 2 = 3 files
        assert os.path.exists(output_dir)
        
        # 验证拆分结果
        output_files = result['data']['output_files']
        assert len(output_files) == 3
        
        # 检查第一个文件
        df1 = pd.read_excel(output_files[0]['path'])
        assert len(df1) == 2
        
        # 检查最后一个文件
        df3 = pd.read_excel(output_files[2]['path'])
        assert len(df3) == 1
    
    # ==================== CSV 拆分测试 ====================
    
    def test_split_csv_missing_file_path(self, engine, temp_dir):
        """测试缺少文件路径"""
        result = engine.split_csv_file({
            'output_dir': temp_dir,
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_FILE_PATH'
    
    def test_split_csv_missing_output_dir(self, engine, sample_csv_files):
        """测试缺少输出目录"""
        result = engine.split_csv_file({
            'file_path': sample_csv_files[0],
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'MISSING_OUTPUT_DIR'
    
    def test_split_csv_invalid_rows_per_file(self, engine, sample_csv_files, temp_dir):
        """测试无效的每文件行数"""
        result = engine.split_csv_file({
            'file_path': sample_csv_files[0],
            'output_dir': temp_dir,
            'rows_per_file': -1
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'INVALID_ROWS_PER_FILE'
    
    def test_split_csv_file_not_found(self, engine, temp_dir):
        """测试文件不存在"""
        output_dir = os.path.join(temp_dir, "output")
        result = engine.split_csv_file({
            'file_path': os.path.join(temp_dir, "nonexistent.csv"),
            'output_dir': output_dir,
            'rows_per_file': 5
        })
        
        assert result['status'] == 'error'
        assert result['error_code'] == 'FILE_NOT_FOUND'
    
    def test_split_csv_success(self, engine, sample_csv_files, temp_dir):
        """测试成功拆分 CSV"""
        output_dir = os.path.join(temp_dir, "output")
        result = engine.split_csv_file({
            'file_path': sample_csv_files[0],
            'output_dir': output_dir,
            'rows_per_file': 2,
            'has_header': True
        })
        
        assert result['status'] == 'success'
        assert result['data']['total_rows'] == 5
        assert result['data']['rows_per_file'] == 2
        assert result['data']['num_files'] == 3
        assert os.path.exists(output_dir)
        
        # 验证拆分结果
        output_files = result['data']['output_files']
        assert len(output_files) == 3
        
        # 检查第一个文件
        df1 = pd.read_csv(output_files[0]['path'])
        assert len(df1) == 2
        
        # 检查最后一个文件
        df3 = pd.read_csv(output_files[2]['path'])
        assert len(df3) == 1


if __name__ == '__main__':
    pytest.main([__file__, '-v'])

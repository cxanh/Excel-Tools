#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Loader 单元测试
"""

import sys
import os
import pytest
import openpyxl
from openpyxl import Workbook

# 添加父目录到路径
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from engine.core.loader import FileLoader


class TestFileLoader:
    """File Loader 测试类"""
    
    def setup_method(self):
        """每个测试方法前执行"""
        self.loader = FileLoader()
        self.test_files_dir = os.path.join(os.path.dirname(__file__), 'test_files')
        
        # 创建测试文件目录
        os.makedirs(self.test_files_dir, exist_ok=True)
    
    def teardown_method(self):
        """每个测试方法后执行"""
        # 关闭加载的文件
        if self.loader.current_workbook:
            self.loader.close()
    
    def _create_test_xlsx(self, filename: str) -> str:
        """创建测试用的 .xlsx 文件"""
        file_path = os.path.join(self.test_files_dir, filename)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        
        # 写入一些测试数据
        ws['A1'] = "姓名"
        ws['B1'] = "年龄"
        ws['C1'] = "城市"
        ws['A2'] = "张三"
        ws['B2'] = 25
        ws['C2'] = "北京"
        ws['A3'] = "李四"
        ws['B3'] = 30
        ws['C3'] = "上海"
        
        # 添加第二个工作表
        ws2 = wb.create_sheet("Sheet2")
        ws2['A1'] = "测试数据"
        
        # 保存文件
        wb.save(file_path)
        wb.close()
        
        return file_path
    
    def test_load_xlsx_success(self):
        """测试成功加载 .xlsx 文件"""
        # 创建测试文件
        file_path = self._create_test_xlsx("test_load.xlsx")
        
        # 加载文件
        result = self.loader.load({"file_path": file_path})
        
        # 验证结果
        assert result["type"] == "result"
        assert result["status"] == "success"
        assert result["message"] == "文件加载成功"
        
        # 验证文件信息
        data = result["data"]
        assert data["file_name"] == "test_load.xlsx"
        assert data["file_format"] == "xlsx"
        assert data["sheet_count"] == 2
        assert len(data["sheets"]) == 2
        assert data["sheets"][0]["name"] == "Sheet1"
        assert data["sheets"][1]["name"] == "Sheet2"
        
        # 验证工作簿已加载
        assert self.loader.current_workbook is not None
        assert self.loader.current_file_path == file_path
        assert self.loader.file_format == "xlsx"
    
    def test_load_file_not_found(self):
        """测试加载不存在的文件"""
        result = self.loader.load({"file_path": "nonexistent.xlsx"})
        
        assert result["type"] == "result"
        assert result["status"] == "error"
        assert result["error_code"] == "FILE_NOT_FOUND"
        assert "文件不存在" in result["message"]
    
    def test_load_missing_file_path(self):
        """测试缺少文件路径参数"""
        result = self.loader.load({})
        
        assert result["type"] == "result"
        assert result["status"] == "error"
        assert result["error_code"] == "MISSING_FILE_PATH"
    
    def test_load_unsupported_format(self):
        """测试不支持的文件格式"""
        # 创建一个文本文件
        file_path = os.path.join(self.test_files_dir, "test.txt")
        with open(file_path, 'w') as f:
            f.write("test")
        
        result = self.loader.load({"file_path": file_path})
        
        assert result["type"] == "result"
        assert result["status"] == "error"
        assert result["error_code"] == "FILE_FORMAT_UNSUPPORTED"
        assert ".txt" in result["message"]
    
    def test_load_corrupted_file(self):
        """测试加载损坏的文件"""
        # 创建一个假的 .xlsx 文件（实际是文本文件）
        file_path = os.path.join(self.test_files_dir, "corrupted.xlsx")
        with open(file_path, 'w') as f:
            f.write("This is not a valid Excel file")
        
        result = self.loader.load({"file_path": file_path})
        
        assert result["type"] == "result"
        assert result["status"] == "error"
        assert result["error_code"] == "FILE_CORRUPTED"
    
    def test_load_with_read_only(self):
        """测试以只读模式加载文件"""
        file_path = self._create_test_xlsx("test_readonly.xlsx")
        
        result = self.loader.load({
            "file_path": file_path,
            "read_only": True
        })
        
        assert result["status"] == "success"
        assert self.loader.current_workbook is not None
    
    def test_load_with_data_only(self):
        """测试只读取值模式"""
        file_path = self._create_test_xlsx("test_dataonly.xlsx")
        
        result = self.loader.load({
            "file_path": file_path,
            "data_only": True
        })
        
        assert result["status"] == "success"
        assert self.loader.current_workbook is not None
    
    def test_close_file(self):
        """测试关闭文件"""
        # 先加载文件
        file_path = self._create_test_xlsx("test_close.xlsx")
        self.loader.load({"file_path": file_path})
        
        # 关闭文件
        result = self.loader.close()
        
        assert result["type"] == "result"
        assert result["status"] == "success"
        assert result["message"] == "文件已关闭"
        
        # 验证工作簿已关闭
        assert self.loader.current_workbook is None
        assert self.loader.current_file_path is None
        assert self.loader.file_format is None
    
    def test_close_no_file(self):
        """测试关闭文件（没有打开的文件）"""
        result = self.loader.close()
        
        assert result["type"] == "result"
        assert result["status"] == "success"
        assert result["message"] == "没有打开的文件"
    
    def test_get_file_info(self):
        """测试获取文件信息"""
        file_path = self._create_test_xlsx("test_info.xlsx")
        result = self.loader.load({"file_path": file_path})
        
        data = result["data"]
        
        # 验证基本信息
        assert "file_path" in data
        assert "file_name" in data
        assert "file_size" in data
        assert "file_format" in data
        assert "sheet_count" in data
        assert "sheets" in data
        assert "properties" in data
        
        # 验证工作表信息
        assert len(data["sheets"]) == 2
        sheet1 = data["sheets"][0]
        assert "name" in sheet1
        assert "max_row" in sheet1
        assert "max_column" in sheet1
        assert "visible" in sheet1
    
    def test_supported_formats(self):
        """测试支持的文件格式列表"""
        assert '.xlsx' in FileLoader.SUPPORTED_FORMATS
        assert '.xlsm' in FileLoader.SUPPORTED_FORMATS
        assert '.xls' in FileLoader.SUPPORTED_FORMATS


class TestFileLoaderIntegration:
    """File Loader 集成测试（通过 CLI Router）"""
    
    def setup_method(self):
        """每个测试方法前执行"""
        from cli_router import CLIRouter
        self.router = CLIRouter()
        self.test_files_dir = os.path.join(os.path.dirname(__file__), 'test_files')
        os.makedirs(self.test_files_dir, exist_ok=True)
    
    def _create_test_xlsx(self, filename: str) -> str:
        """创建测试用的 .xlsx 文件"""
        file_path = os.path.join(self.test_files_dir, filename)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws['A1'] = "测试数据"
        wb.save(file_path)
        wb.close()
        
        return file_path
    
    def test_load_file_via_router(self):
        """测试通过路由器加载文件"""
        file_path = self._create_test_xlsx("test_router.xlsx")
        
        command = {
            "action": "load_file",
            "params": {
                "file_path": file_path
            }
        }
        
        result = self.router.route(command)
        
        assert result["type"] == "result"
        assert result["status"] == "success"
        assert "data" in result
        assert result["data"]["file_name"] == "test_router.xlsx"
    
    def test_close_file_via_router(self):
        """测试通过路由器关闭文件"""
        file_path = self._create_test_xlsx("test_router_close.xlsx")
        
        # 先加载文件
        self.router.route({
            "action": "load_file",
            "params": {"file_path": file_path}
        })
        
        # 关闭文件
        result = self.router.route({
            "action": "close_file",
            "params": {}
        })
        
        assert result["type"] == "result"
        assert result["status"] == "success"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])

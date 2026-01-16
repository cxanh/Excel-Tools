#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
File Loader - 文件加载模块
负责加载 Excel 文件（.xlsx 和 .xls 格式）
"""

import os
import sys
from typing import Dict, Any, Optional
from zipfile import BadZipFile
import openpyxl
import xlrd
from openpyxl.utils.exceptions import InvalidFileException


def log(message):
    """将日志输出到 stderr"""
    sys.stderr.write(f"[LOADER] {message}\n")
    sys.stderr.flush()


class FileLoader:
    """文件加载器"""
    
    # 支持的文件格式
    SUPPORTED_FORMATS = {
        '.xlsx': 'Excel 2007+ (*.xlsx)',
        '.xlsm': 'Excel 2007+ Macro-Enabled (*.xlsm)',
        '.xls': 'Excel 97-2003 (*.xls)',
    }
    
    def __init__(self):
        """初始化文件加载器"""
        self.current_workbook = None
        self.current_file_path = None
        self.file_format = None
    
    def load(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        加载 Excel 文件
        
        Args:
            params: dict, 包含以下字段：
                - file_path: str, 文件路径
                - read_only: bool, 是否以只读模式打开（可选，默认 False）
                - data_only: bool, 是否只读取值而非公式（可选，默认 False）
        
        Returns:
            dict, 包含以下字段：
                - type: "result"
                - status: "success" 或 "error"
                - message: 消息
                - data: 文件信息（成功时）
                - error_code: 错误代码（失败时）
        """
        file_path = params.get('file_path')
        read_only = params.get('read_only', False)
        data_only = params.get('data_only', False)
        
        # 验证文件路径
        if not file_path:
            return self._error_response(
                "MISSING_FILE_PATH",
                "缺少文件路径参数"
            )
        
        # 检查文件是否存在
        if not os.path.exists(file_path):
            return self._error_response(
                "FILE_NOT_FOUND",
                f"文件不存在: {file_path}",
                suggested_action="请检查文件路径是否正确"
            )
        
        # 检查文件格式
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext not in self.SUPPORTED_FORMATS:
            return self._error_response(
                "FILE_FORMAT_UNSUPPORTED",
                f"不支持的文件格式: {file_ext}",
                suggested_action=f"支持的格式: {', '.join(self.SUPPORTED_FORMATS.keys())}"
            )
        
        # 根据文件格式选择加载方法
        try:
            if file_ext in ['.xlsx', '.xlsm']:
                return self._load_xlsx(file_path, read_only, data_only)
            elif file_ext == '.xls':
                return self._load_xls(file_path)
            else:
                return self._error_response(
                    "FILE_FORMAT_UNSUPPORTED",
                    f"不支持的文件格式: {file_ext}"
                )
        except Exception as e:
            log(f"Unexpected error loading file: {str(e)}")
            return self._error_response(
                "INTERNAL_ERROR",
                f"加载文件时发生未知错误: {str(e)}"
            )
    
    def _load_xlsx(self, file_path: str, read_only: bool, data_only: bool) -> Dict[str, Any]:
        """
        加载 .xlsx 或 .xlsm 文件
        
        Args:
            file_path: 文件路径
            read_only: 是否只读模式
            data_only: 是否只读取值
        
        Returns:
            dict, 加载结果
        """
        try:
            log(f"Loading .xlsx file: {file_path}")
            
            # 使用 openpyxl 加载文件
            workbook = openpyxl.load_workbook(
                file_path,
                read_only=read_only,
                data_only=data_only
            )
            
            # 保存当前工作簿信息
            self.current_workbook = workbook
            self.current_file_path = file_path
            self.file_format = 'xlsx'
            
            # 获取文件信息
            file_info = self._get_file_info(workbook, file_path)
            
            log(f"Successfully loaded .xlsx file: {file_path}")
            
            return {
                "type": "result",
                "status": "success",
                "message": "文件加载成功",
                "data": file_info
            }
            
        except PermissionError:
            log(f"Permission denied: {file_path}")
            return self._error_response(
                "FILE_IN_USE",
                "文件正在被其他程序使用，请先关闭 Excel 后重试",
                suggested_action="关闭 Microsoft Excel 或其他正在使用该文件的程序"
            )
        
        except (IOError, OSError, ValueError, KeyError, BadZipFile) as e:
            # 捕获文件格式错误（如 "File is not a zip file"）
            error_msg = str(e).lower()
            if 'zip' in error_msg or 'format' in error_msg or 'corrupt' in error_msg:
                log(f"Corrupted file: {file_path}, error: {str(e)}")
                return self._error_response(
                    "FILE_CORRUPTED",
                    "文件已损坏或格式不正确",
                    suggested_action="请检查文件是否完整，或尝试使用 Excel 修复该文件"
                )
            else:
                log(f"Error loading .xlsx file: {str(e)}")
                return self._error_response(
                    "LOAD_ERROR",
                    f"加载 .xlsx 文件失败: {str(e)}"
                )
        
        except InvalidFileException as e:
            log(f"Invalid file: {file_path}, error: {str(e)}")
            return self._error_response(
                "FILE_CORRUPTED",
                "文件已损坏或格式不正确",
                suggested_action="请检查文件是否完整，或尝试使用 Excel 修复该文件"
            )
        
        except Exception as e:
            log(f"Error loading .xlsx file: {str(e)}")
            return self._error_response(
                "LOAD_ERROR",
                f"加载 .xlsx 文件失败: {str(e)}"
            )
    
    def _load_xls(self, file_path: str) -> Dict[str, Any]:
        """
        加载 .xls 文件（Excel 97-2003 格式）
        
        Args:
            file_path: 文件路径
        
        Returns:
            dict, 加载结果
        """
        try:
            log(f"Loading .xls file: {file_path}")
            
            # 使用 xlrd 加载文件
            workbook = xlrd.open_workbook(file_path)
            
            # 保存当前工作簿信息
            self.current_workbook = workbook
            self.current_file_path = file_path
            self.file_format = 'xls'
            
            # 获取文件信息
            file_info = self._get_file_info_xls(workbook, file_path)
            
            log(f"Successfully loaded .xls file: {file_path}")
            
            return {
                "type": "result",
                "status": "success",
                "message": "文件加载成功",
                "data": file_info
            }
            
        except PermissionError:
            log(f"Permission denied: {file_path}")
            return self._error_response(
                "FILE_IN_USE",
                "文件正在被其他程序使用，请先关闭 Excel 后重试",
                suggested_action="关闭 Microsoft Excel 或其他正在使用该文件的程序"
            )
        
        except xlrd.XLRDError as e:
            log(f"Invalid .xls file: {file_path}, error: {str(e)}")
            return self._error_response(
                "FILE_CORRUPTED",
                "文件已损坏或格式不正确",
                suggested_action="请检查文件是否完整，或尝试使用 Excel 修复该文件"
            )
        
        except Exception as e:
            log(f"Error loading .xls file: {str(e)}")
            return self._error_response(
                "LOAD_ERROR",
                f"加载 .xls 文件失败: {str(e)}"
            )
    
    def _get_file_info(self, workbook, file_path: str) -> Dict[str, Any]:
        """
        获取 .xlsx 文件信息
        
        Args:
            workbook: openpyxl Workbook 对象
            file_path: 文件路径
        
        Returns:
            dict, 文件信息
        """
        # 获取文件大小
        file_size = os.path.getsize(file_path)
        
        # 获取工作表信息
        sheets = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheets.append({
                "name": sheet_name,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "visible": sheet.sheet_state == 'visible'
            })
        
        return {
            "file_path": file_path,
            "file_name": os.path.basename(file_path),
            "file_size": file_size,
            "file_format": "xlsx",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "properties": {
                "creator": workbook.properties.creator if workbook.properties else None,
                "last_modified_by": workbook.properties.lastModifiedBy if workbook.properties else None,
                "created": str(workbook.properties.created) if workbook.properties and workbook.properties.created else None,
                "modified": str(workbook.properties.modified) if workbook.properties and workbook.properties.modified else None,
            }
        }
    
    def _get_file_info_xls(self, workbook, file_path: str) -> Dict[str, Any]:
        """
        获取 .xls 文件信息
        
        Args:
            workbook: xlrd Book 对象
            file_path: 文件路径
        
        Returns:
            dict, 文件信息
        """
        # 获取文件大小
        file_size = os.path.getsize(file_path)
        
        # 获取工作表信息
        sheets = []
        for sheet_index in range(workbook.nsheets):
            sheet = workbook.sheet_by_index(sheet_index)
            sheets.append({
                "name": sheet.name,
                "max_row": sheet.nrows,
                "max_column": sheet.ncols,
                "visible": sheet.visibility == 0
            })
        
        return {
            "file_path": file_path,
            "file_name": os.path.basename(file_path),
            "file_size": file_size,
            "file_format": "xls",
            "sheet_count": len(sheets),
            "sheets": sheets,
            "properties": {
                "creator": None,
                "last_modified_by": None,
                "created": None,
                "modified": None,
            }
        }
    
    def _error_response(self, error_code: str, message: str, suggested_action: Optional[str] = None) -> Dict[str, Any]:
        """
        生成错误响应
        
        Args:
            error_code: 错误代码
            message: 错误消息
            suggested_action: 建议的解决方案（可选）
        
        Returns:
            dict, 错误响应
        """
        response = {
            "type": "result",
            "status": "error",
            "error_code": error_code,
            "message": message
        }
        
        if suggested_action:
            response["suggested_action"] = suggested_action
        
        return response
    
    def get_current_workbook(self):
        """获取当前加载的工作簿"""
        return self.current_workbook
    
    def get_current_file_path(self) -> Optional[str]:
        """获取当前文件路径"""
        return self.current_file_path
    
    def get_file_format(self) -> Optional[str]:
        """获取当前文件格式"""
        return self.file_format
    
    def close(self, params: Dict[str, Any] = None) -> Dict[str, Any]:
        """
        关闭当前工作簿
        
        Args:
            params: dict, 参数（可选，保持接口一致性）
        
        Returns:
            dict, 操作结果
        """
        if self.current_workbook:
            try:
                if self.file_format == 'xlsx':
                    self.current_workbook.close()
                
                self.current_workbook = None
                self.current_file_path = None
                self.file_format = None
                
                return {
                    "type": "result",
                    "status": "success",
                    "message": "文件已关闭"
                }
            except Exception as e:
                log(f"Error closing workbook: {str(e)}")
                return self._error_response(
                    "CLOSE_ERROR",
                    f"关闭文件失败: {str(e)}"
                )
        else:
            return {
                "type": "result",
                "status": "success",
                "message": "没有打开的文件"
            }

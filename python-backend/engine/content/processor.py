#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Content Processor - 内容处理模块
负责 Excel 内容的各种处理操作
"""

import sys
import re
from typing import Dict, Any, List, Set, Optional
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter


def log(message):
    """将日志输出到 stderr"""
    sys.stderr.write(f"[PROCESSOR] {message}\n")
    sys.stderr.flush()


def send_progress(progress, message="", data=None):
    """发送进度消息（安全版本，不会抛异常）"""
    try:
        import json
        progress_msg = {
            "type": "progress",
            "progress": progress,
            "message": message
        }
        if data:
            progress_msg["data"] = data
        
        print(json.dumps(progress_msg, ensure_ascii=False), flush=True)
    except Exception as e:
        # 进度发送失败不应该影响主流程
        log(f"Failed to send progress: {str(e)}")


def send_progress_safe(current, total, stage="processing", message_prefix="正在处理"):
    """
    安全的进度发送封装
    
    Args:
        current: 当前进度
        total: 总数
        stage: 阶段名称
        message_prefix: 消息前缀
    """
    try:
        if total <= 0:
            return
        
        # 小表格也至少发送一次
        if total < 100 or current % max(1, total // 20) == 0 or current == total:
            progress = min(100, int(current / total * 100))
            message = f"{message_prefix}第 {current}/{total} 行"
            send_progress(progress, message)
    except Exception as e:
        log(f"send_progress_safe error: {str(e)}")


class ContentProcessor:
    """内容处理器"""
    
    def __init__(self):
        """初始化内容处理器"""
        pass
    
    def remove_blank_rows(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        删除空白行（稳态版本）
        
        Args:
            params: dict, 包含以下字段：
                - worksheet: Worksheet 对象（必需）
                - sheet_name: str, 工作表名称（用于日志）
        
        Returns:
            dict, 处理结果
        """
        worksheet = params.get('worksheet')
        sheet_name = params.get('sheet_name', 'Unknown')
        
        if not worksheet:
            return self._error_response(
                "MISSING_WORKSHEET",
                "缺少工作表对象"
            )
        
        try:
            log(f"Removing blank rows from sheet: {sheet_name}")
            
            # 获取所有行
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            
            # 标记要删除的行
            rows_to_delete = []
            
            # 扫描阶段
            for row_idx in range(1, max_row + 1):
                try:
                    # 检查该行是否为空
                    is_blank = True
                    for col_idx in range(1, max_col + 1):
                        cell_value = worksheet.cell(row_idx, col_idx).value
                        if cell_value is not None and str(cell_value).strip() != '':
                            is_blank = False
                            break
                    
                    if is_blank:
                        rows_to_delete.append(row_idx)
                except Exception as e:
                    log(f"Error reading row {row_idx}: {str(e)}")
                    # 读取失败的行跳过，不影响其他行
                    continue
                
                # 安全发送进度
                send_progress_safe(row_idx, max_row, "scan", "正在扫描")
            
            # 删除阶段：从后往前删除（避免索引变化）
            deleted_count = 0
            total_to_delete = len(rows_to_delete)
            
            for idx, row_idx in enumerate(reversed(rows_to_delete)):
                try:
                    worksheet.delete_rows(row_idx, 1)
                    deleted_count += 1
                except Exception as e:
                    log(f"Error deleting row {row_idx}: {str(e)}")
                    # 删除失败的行跳过，不中断整个流程
                    continue
                
                # 安全发送删除进度
                if total_to_delete > 0:
                    send_progress_safe(idx + 1, total_to_delete, "delete", "正在删除")
            
            log(f"Removed {deleted_count} blank rows from sheet: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功删除 {deleted_count} 个空白行",
                "data": {
                    "sheet_name": sheet_name,
                    "deleted_count": deleted_count,
                    "original_rows": max_row,
                    "remaining_rows": max_row - deleted_count
                }
            }
            
        except Exception as e:
            log(f"Unexpected error in remove_blank_rows: {str(e)}")
            return self._error_response(
                "PROCESS_ERROR",
                f"删除空白行失败: {str(e)}"
            )
    
    def clear_blank_cells(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        清除空白单元格内容（稳态版本）
        
        Args:
            params: dict, 包含以下字段：
                - worksheet: Worksheet 对象（必需）
                - sheet_name: str, 工作表名称（用于日志）
        
        Returns:
            dict, 处理结果
        """
        worksheet = params.get('worksheet')
        sheet_name = params.get('sheet_name', 'Unknown')
        
        if not worksheet:
            return self._error_response(
                "MISSING_WORKSHEET",
                "缺少工作表对象"
            )
        
        try:
            log(f"Clearing blank cells from sheet: {sheet_name}")
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            cleared_count = 0
            
            for row_idx in range(1, max_row + 1):
                try:
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row_idx, col_idx)
                        if cell.value is not None and str(cell.value).strip() == '':
                            cell.value = None
                            cleared_count += 1
                except Exception as e:
                    log(f"Error processing row {row_idx}: {str(e)}")
                    continue
                
                # 安全发送进度
                send_progress_safe(row_idx, max_row, "clear", "正在清除")
            
            log(f"Cleared {cleared_count} blank cells from sheet: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功清除 {cleared_count} 个空白单元格",
                "data": {
                    "sheet_name": sheet_name,
                    "cleared_count": cleared_count
                }
            }
            
        except Exception as e:
            log(f"Unexpected error in clear_blank_cells: {str(e)}")
            return self._error_response(
                "PROCESS_ERROR",
                f"清除空白单元格失败: {str(e)}"
            )
    
    def remove_formulas(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        删除公式（稳态版本）
        
        Args:
            params: dict, 包含以下字段：
                - worksheet: Worksheet 对象（必需）
                - sheet_name: str, 工作表名称（用于日志）
        
        Returns:
            dict, 处理结果
        """
        worksheet = params.get('worksheet')
        sheet_name = params.get('sheet_name', 'Unknown')
        
        if not worksheet:
            return self._error_response(
                "MISSING_WORKSHEET",
                "缺少工作表对象"
            )
        
        try:
            log(f"Removing formulas from sheet: {sheet_name}")
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            formula_count = 0
            
            for row_idx in range(1, max_row + 1):
                try:
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row_idx, col_idx)
                        
                        # 如果单元格包含公式
                        if cell.data_type == 'f':
                            # 保存当前值
                            current_value = cell.value
                            # 将公式替换为值
                            cell.value = current_value
                            formula_count += 1
                except Exception as e:
                    log(f"Error processing row {row_idx}: {str(e)}")
                    continue
                
                # 安全发送进度
                send_progress_safe(row_idx, max_row, "formula", "正在处理")
            
            log(f"Removed {formula_count} formulas from sheet: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功删除 {formula_count} 个公式",
                "data": {
                    "sheet_name": sheet_name,
                    "formula_count": formula_count
                }
            }
            
        except Exception as e:
            log(f"Unexpected error in remove_formulas: {str(e)}")
            return self._error_response(
                "PROCESS_ERROR",
                f"删除公式失败: {str(e)}"
            )
    
    def remove_duplicate_rows(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        删除重复行（稳态模板）
        
        Args:
            params: dict
                - worksheet: Worksheet 对象（必需）
                - sheet_name: str，工作表名称（用于日志，默认 'Unknown'）
                - key_columns: list，用于判断重复的关键列索引（可选，默认所有列）
                - keep_first: bool，是否保留第一次出现的行（可选，默认 True）
        
        Returns:
            dict, 处理结果（统一 JSON 格式）
        """
        worksheet = params.get('worksheet')
        sheet_name = params.get('sheet_name', 'Unknown')
        key_columns = params.get('key_columns')
        keep_first = params.get('keep_first', True)

        # 必须有 worksheet
        if not worksheet:
            return self._error_response(
                "MISSING_WORKSHEET",
                "缺少工作表对象"
            )

        try:
            log(f"Removing duplicate rows from sheet: {sheet_name}")

            max_row = worksheet.max_row
            max_col = worksheet.max_column

            # 默认使用所有列
            if not key_columns:
                key_columns = list(range(1, max_col + 1))

            seen_rows: set[tuple] = set()
            rows_to_delete: list[int] = []

            # 扫描阶段
            for row_idx in range(1, max_row + 1):
                try:
                    row_key = tuple(
                        str(worksheet.cell(row=row_idx, column=col_idx).value or '')
                        for col_idx in key_columns
                    )
                except Exception as e:
                    log(f"Error reading row {row_idx}: {str(e)}")
                    row_key = tuple('' for _ in key_columns)  # 异常行当作空行

                if row_key in seen_rows:
                    rows_to_delete.append(row_idx)
                else:
                    seen_rows.add(row_key)

                # 安全发送扫描进度
                if row_idx % 100 == 0:
                    try:
                        progress = int(row_idx / max_row * 50)  # 扫描阶段占 50%
                        send_progress(progress, f"正在扫描第 {row_idx}/{max_row} 行")
                    except Exception as e:
                        log(f"send_progress error during scan: {str(e)}")

            # 删除阶段
            deleted_count = 0
            total_to_delete = len(rows_to_delete) or 1  # 避免除零

            for idx, row_idx in enumerate(reversed(rows_to_delete)):
                try:
                    worksheet.delete_rows(row_idx, 1)
                    deleted_count += 1
                except Exception as e:
                    log(f"Error deleting row {row_idx}: {str(e)}")
                    continue  # 异常行跳过，不中断

                # 安全发送删除进度
                if idx % 10 == 0:
                    try:
                        progress = 50 + int(idx / total_to_delete * 50)  # 删除阶段占 50%
                        send_progress(progress, f"正在删除第 {idx + 1}/{total_to_delete} 个重复行")
                    except Exception as e:
                        log(f"send_progress error during delete: {str(e)}")

            log(f"Removed {deleted_count} duplicate rows from sheet: {sheet_name}")

            # 返回统一 JSON
            return {
                "type": "result",
                "status": "success",
                "message": f"成功删除 {deleted_count} 个重复行",
                "data": {
                    "sheet_name": sheet_name,
                    "deleted_count": deleted_count,
                    "original_rows": max_row,
                    "remaining_rows": max_row - deleted_count,
                    "key_columns": key_columns
                }
            }

        except Exception as e:
            log(f"Unexpected error in remove_duplicate_rows: {str(e)}")
            return self._error_response(
                "PROCESS_ERROR",
                f"删除重复行失败: {str(e)}"
            )
    
    def replace_content(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        按规则替换内容（稳态版本）
        
        Args:
            params: dict, 包含以下字段：
                - worksheet: Worksheet 对象（必需）
                - sheet_name: str, 工作表名称（用于日志）
                - find_text: str, 要查找的文本（必需）
                - replace_text: str, 替换为的文本（必需）
                - use_regex: bool, 是否使用正则表达式（可选，默认 False）
                - case_sensitive: bool, 是否区分大小写（可选，默认 False）
                - whole_word: bool, 是否全词匹配（可选，默认 False）
        
        Returns:
            dict, 处理结果
        """
        worksheet = params.get('worksheet')
        sheet_name = params.get('sheet_name', 'Unknown')
        find_text = params.get('find_text')
        replace_text = params.get('replace_text', '')
        use_regex = params.get('use_regex', False)
        case_sensitive = params.get('case_sensitive', False)
        whole_word = params.get('whole_word', False)
        
        if not worksheet:
            return self._error_response(
                "MISSING_WORKSHEET",
                "缺少工作表对象"
            )
        
        if find_text is None:
            return self._error_response(
                "MISSING_FIND_TEXT",
                "缺少查找文本参数"
            )
        
        try:
            log(f"Replacing content in sheet: {sheet_name}")
            
            max_row = worksheet.max_row
            max_col = worksheet.max_column
            replace_count = 0
            
            # 编译正则表达式（如果使用）
            pattern = None
            if use_regex:
                flags = 0 if case_sensitive else re.IGNORECASE
                try:
                    pattern = re.compile(find_text, flags)
                except re.error as e:
                    return self._error_response(
                        "INVALID_REGEX",
                        f"无效的正则表达式: {str(e)}"
                    )
            
            for row_idx in range(1, max_row + 1):
                try:
                    for col_idx in range(1, max_col + 1):
                        cell = worksheet.cell(row_idx, col_idx)
                        
                        if cell.value is None:
                            continue
                        
                        cell_value = str(cell.value)
                        
                        # 执行替换
                        if use_regex:
                            # 正则表达式替换
                            new_value = pattern.sub(replace_text, cell_value)
                            if new_value != cell_value:
                                cell.value = new_value
                                replace_count += 1
                        else:
                            # 普通文本替换
                            if whole_word:
                                # 全词匹配
                                word_pattern = r'\b' + re.escape(find_text) + r'\b'
                                flags = 0 if case_sensitive else re.IGNORECASE
                                new_value = re.sub(word_pattern, replace_text, cell_value, flags=flags)
                            else:
                                # 普通替换
                                if case_sensitive:
                                    new_value = cell_value.replace(find_text, replace_text)
                                else:
                                    # 不区分大小写的替换
                                    case_pattern = re.compile(re.escape(find_text), re.IGNORECASE)
                                    new_value = case_pattern.sub(replace_text, cell_value)
                            
                            if new_value != cell_value:
                                cell.value = new_value
                                replace_count += 1
                except Exception as e:
                    log(f"Error processing cell ({row_idx}, {col_idx}): {str(e)}")
                    continue
                
                # 安全发送进度
                send_progress_safe(row_idx, max_row, "replace", "正在替换")
            
            log(f"Replaced {replace_count} cells in sheet: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功替换 {replace_count} 个单元格",
                "data": {
                    "sheet_name": sheet_name,
                    "replace_count": replace_count,
                    "find_text": find_text,
                    "replace_text": replace_text,
                    "use_regex": use_regex,
                    "case_sensitive": case_sensitive,
                    "whole_word": whole_word
                }
            }
            
        except Exception as e:
            log(f"Unexpected error in replace_content: {str(e)}")
            return self._error_response(
                "PROCESS_ERROR",
                f"替换内容失败: {str(e)}"
            )
    
    def _error_response(self, error_code: str, message: str, suggested_action: Optional[str] = None) -> Dict[str, Any]:
        """
        生成错误响应
        
        Args:
            error_code: 错误代码
            message: 错误消息
            suggested_action: 建议的解决方案（可选）
        
        Returns:
            dict, 错误响应
        """
        response = {
            "type": "result",
            "status": "error",
            "error_code": error_code,
            "message": message
        }
        
        if suggested_action:
            response["suggested_action"] = suggested_action
        
        return response

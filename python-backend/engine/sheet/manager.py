#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sheet Manager - 工作表管理器
管理 Excel 工作表的插入、删除、重命名等操作
"""

import sys
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet


def log(message):
    """将日志输出到 stderr"""
    sys.stderr.write(f"[SHEET_MANAGER] {message}\n")
    sys.stderr.flush()


class SheetManager:
    """工作表管理器"""
    
    def insert_sheet(self, params):
        """
        插入新工作表
        
        Args:
            params: dict, 包含以下字段：
                - workbook: Workbook, 工作簿对象
                - sheet_name: str, 新工作表名称（可选，默认：Sheet{n}）
                - index: int, 插入位置（可选，默认：末尾）
        
        Returns:
            dict, 包含操作结果
        """
        workbook = params.get('workbook')
        sheet_name = params.get('sheet_name')
        index = params.get('index')
        
        # 验证参数
        if not workbook:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_WORKBOOK",
                "message": "缺少工作簿对象"
            }
        
        try:
            # 如果没有指定名称，生成默认名称
            if not sheet_name:
                # 查找可用的默认名称
                base_name = "Sheet"
                counter = 1
                while f"{base_name}{counter}" in workbook.sheetnames:
                    counter += 1
                sheet_name = f"{base_name}{counter}"
            
            # 检查名称是否已存在
            if sheet_name in workbook.sheetnames:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "SHEET_NAME_EXISTS",
                    "message": f"工作表名称已存在: {sheet_name}",
                    "suggested_action": "请使用不同的工作表名称"
                }
            
            # 检查名称是否有效（Excel 限制）
            if len(sheet_name) > 31:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "INVALID_SHEET_NAME",
                    "message": "工作表名称过长（最多 31 个字符）",
                    "suggested_action": "请使用较短的名称"
                }
            
            # 检查名称中是否包含非法字符
            invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
            for char in invalid_chars:
                if char in sheet_name:
                    return {
                        "type": "result",
                        "status": "error",
                        "error_code": "INVALID_SHEET_NAME",
                        "message": f"工作表名称包含非法字符: {char}",
                        "suggested_action": f"工作表名称不能包含以下字符: {', '.join(invalid_chars)}"
                    }
            
            # 创建新工作表
            if index is not None:
                # 在指定位置插入
                new_sheet = workbook.create_sheet(title=sheet_name, index=index)
                log(f"在位置 {index} 插入工作表: {sheet_name}")
            else:
                # 在末尾添加
                new_sheet = workbook.create_sheet(title=sheet_name)
                log(f"在末尾添加工作表: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功插入工作表: {sheet_name}",
                "data": {
                    "sheet_name": sheet_name,
                    "index": workbook.sheetnames.index(sheet_name),
                    "total_sheets": len(workbook.sheetnames)
                }
            }
        
        except Exception as e:
            log(f"插入工作表时发生错误: {str(e)}")
            return {
                "type": "result",
                "status": "error",
                "error_code": "INSERT_ERROR",
                "message": f"插入工作表时发生错误: {str(e)}"
            }
    
    def delete_sheet(self, params):
        """
        删除工作表
        
        Args:
            params: dict, 包含以下字段：
                - workbook: Workbook, 工作簿对象
                - sheet_name: str, 要删除的工作表名称
        
        Returns:
            dict, 包含操作结果
        """
        workbook = params.get('workbook')
        sheet_name = params.get('sheet_name')
        
        # 验证参数
        if not workbook:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_WORKBOOK",
                "message": "缺少工作簿对象"
            }
        
        if not sheet_name:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_SHEET_NAME",
                "message": "缺少工作表名称参数"
            }
        
        try:
            # 检查工作表是否存在
            if sheet_name not in workbook.sheetnames:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "SHEET_NOT_FOUND",
                    "message": f"工作表不存在: {sheet_name}"
                }
            
            # 检查是否是唯一的工作表
            if len(workbook.sheetnames) == 1:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "CANNOT_DELETE_LAST_SHEET",
                    "message": "不能删除唯一的工作表",
                    "suggested_action": "工作簿至少需要保留一个工作表"
                }
            
            # 删除工作表
            sheet = workbook[sheet_name]
            workbook.remove(sheet)
            log(f"删除工作表: {sheet_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功删除工作表: {sheet_name}",
                "data": {
                    "deleted_sheet": sheet_name,
                    "remaining_sheets": len(workbook.sheetnames),
                    "sheet_names": workbook.sheetnames
                }
            }
        
        except Exception as e:
            log(f"删除工作表时发生错误: {str(e)}")
            return {
                "type": "result",
                "status": "error",
                "error_code": "DELETE_ERROR",
                "message": f"删除工作表时发生错误: {str(e)}"
            }
    
    def rename_sheet(self, params):
        """
        重命名工作表
        
        Args:
            params: dict, 包含以下字段：
                - workbook: Workbook, 工作簿对象
                - old_name: str, 原工作表名称
                - new_name: str, 新工作表名称
        
        Returns:
            dict, 包含操作结果
        """
        workbook = params.get('workbook')
        old_name = params.get('old_name')
        new_name = params.get('new_name')
        
        # 验证参数
        if not workbook:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_WORKBOOK",
                "message": "缺少工作簿对象"
            }
        
        if not old_name:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_OLD_NAME",
                "message": "缺少原工作表名称参数"
            }
        
        if not new_name:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_NEW_NAME",
                "message": "缺少新工作表名称参数"
            }
        
        try:
            # 检查原工作表是否存在
            if old_name not in workbook.sheetnames:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "SHEET_NOT_FOUND",
                    "message": f"工作表不存在: {old_name}"
                }
            
            # 检查新名称是否已存在（排除自己）
            if new_name != old_name and new_name in workbook.sheetnames:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "SHEET_NAME_EXISTS",
                    "message": f"工作表名称已存在: {new_name}",
                    "suggested_action": "请使用不同的工作表名称"
                }
            
            # 检查名称是否有效
            if len(new_name) > 31:
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "INVALID_SHEET_NAME",
                    "message": "工作表名称过长（最多 31 个字符）",
                    "suggested_action": "请使用较短的名称"
                }
            
            # 检查名称中是否包含非法字符
            invalid_chars = ['\\', '/', '*', '?', ':', '[', ']']
            for char in invalid_chars:
                if char in new_name:
                    return {
                        "type": "result",
                        "status": "error",
                        "error_code": "INVALID_SHEET_NAME",
                        "message": f"工作表名称包含非法字符: {char}",
                        "suggested_action": f"工作表名称不能包含以下字符: {', '.join(invalid_chars)}"
                    }
            
            # 重命名工作表
            sheet = workbook[old_name]
            sheet.title = new_name
            log(f"重命名工作表: {old_name} -> {new_name}")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功重命名工作表: {old_name} -> {new_name}",
                "data": {
                    "old_name": old_name,
                    "new_name": new_name,
                    "index": workbook.sheetnames.index(new_name)
                }
            }
        
        except Exception as e:
            log(f"重命名工作表时发生错误: {str(e)}")
            return {
                "type": "result",
                "status": "error",
                "error_code": "RENAME_ERROR",
                "message": f"重命名工作表时发生错误: {str(e)}"
            }
    
    def get_sheet_info(self, params):
        """
        获取工作表信息
        
        Args:
            params: dict, 包含以下字段：
                - workbook: Workbook, 工作簿对象
        
        Returns:
            dict, 包含工作表信息
        """
        workbook = params.get('workbook')
        
        if not workbook:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_WORKBOOK",
                "message": "缺少工作簿对象"
            }
        
        try:
            sheets_info = []
            
            for idx, sheet_name in enumerate(workbook.sheetnames):
                sheet = workbook[sheet_name]
                sheets_info.append({
                    "name": sheet_name,
                    "index": idx,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "visible": sheet.sheet_state == 'visible'
                })
            
            return {
                "type": "result",
                "status": "success",
                "message": f"共有 {len(sheets_info)} 个工作表",
                "data": {
                    "total_sheets": len(sheets_info),
                    "sheets": sheets_info
                }
            }
        
        except Exception as e:
            log(f"获取工作表信息时发生错误: {str(e)}")
            return {
                "type": "result",
                "status": "error",
                "error_code": "INFO_ERROR",
                "message": f"获取工作表信息时发生错误: {str(e)}"
            }

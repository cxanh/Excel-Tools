#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Merge and Split Engine - 文件合并与拆分引擎
负责 Excel 和 CSV 文件的合并与拆分操作
"""

import os
import sys
import openpyxl
import pandas as pd
from typing import List, Dict, Any, Optional


def send_progress(progress: int, message: str = ""):
    """发送进度消息到 stdout"""
    progress_msg = {
        "type": "progress",
        "progress": progress,
        "message": message
    }
    import json
    print(json.dumps(progress_msg), flush=True)


class MergeSplitEngine:
    """文件合并与拆分引擎"""
    
    def merge_excel_files(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        合并多个 Excel 文件
        
        Args:
            params: {
                'file_paths': List[str],  # 要合并的文件路径列表
                'output_path': str,  # 输出文件路径
                'mode': str,  # 合并模式: 'append' 或 'separate'
                'has_header': bool,  # 是否包含表头（默认 True）
                'align_mode': str,  # 列对齐模式: 'by_name', 'by_order'（默认 'by_name'）
                'sheet_name': str  # 目标工作表名称（append 模式使用，默认 'Merged'）
            }
        
        Returns:
            dict: 操作结果
        """
        # 验证参数
        file_paths = params.get('file_paths')
        output_path = params.get('output_path')
        mode = params.get('mode', 'append')
        has_header = params.get('has_header', True)
        align_mode = params.get('align_mode', 'by_name')
        sheet_name = params.get('sheet_name', 'Merged')
        
        if not file_paths:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_FILE_PATHS",
                "message": "缺少文件路径列表"
            }
        
        if not output_path:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_OUTPUT_PATH",
                "message": "缺少输出文件路径"
            }
        
        if mode not in ['append', 'separate']:
            return {
                "type": "result",
                "status": "error",
                "error_code": "INVALID_MODE",
                "message": f"无效的合并模式: {mode}，必须是 'append' 或 'separate'"
            }
        
        # 检查文件是否存在
        for file_path in file_paths:
            if not os.path.exists(file_path):
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "FILE_NOT_FOUND",
                    "message": f"文件不存在: {file_path}"
                }
        
        try:
            if mode == 'append':
                return self._merge_excel_append(
                    file_paths, output_path, has_header, align_mode, sheet_name
                )
            else:  # separate
                return self._merge_excel_separate(file_paths, output_path)
        
        except Exception as e:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MERGE_ERROR",
                "message": f"合并文件时发生错误: {str(e)}"
            }
    
    def _merge_excel_append(
        self,
        file_paths: List[str],
        output_path: str,
        has_header: bool,
        align_mode: str,
        sheet_name: str
    ) -> Dict[str, Any]:
        """追加模式合并 Excel 文件"""
        all_data = []
        header = None
        total_rows = 0
        
        for idx, file_path in enumerate(file_paths):
            send_progress(
                int((idx / len(file_paths)) * 80),
                f"正在读取文件 {idx + 1}/{len(file_paths)}: {os.path.basename(file_path)}"
            )
            
            # 使用 pandas 读取 Excel
            df = pd.read_excel(file_path, header=0 if has_header else None)
            
            if has_header:
                if header is None:
                    # 第一个文件，保存表头
                    header = df.columns.tolist()
                else:
                    # 后续文件，根据对齐模式处理
                    if align_mode == 'by_name':
                        # 按列名对齐
                        df = df.reindex(columns=header, fill_value='')
                    # by_order 模式不需要特殊处理
            
            all_data.append(df)
            total_rows += len(df)
        
        send_progress(85, "正在合并数据...")
        
        # 合并所有数据
        merged_df = pd.concat(all_data, ignore_index=True)
        
        send_progress(90, "正在保存文件...")
        
        # 保存到 Excel
        merged_df.to_excel(output_path, sheet_name=sheet_name, index=False)
        
        send_progress(100, "合并完成")
        
        return {
            "type": "result",
            "status": "success",
            "message": f"成功合并 {len(file_paths)} 个文件",
            "data": {
                "total_files": len(file_paths),
                "total_rows": total_rows,
                "output_path": output_path,
                "mode": "append",
                "sheet_name": sheet_name
            }
        }
    
    def _merge_excel_separate(
        self,
        file_paths: List[str],
        output_path: str
    ) -> Dict[str, Any]:
        """独立工作表模式合并 Excel 文件"""
        # 创建新工作簿
        output_wb = openpyxl.Workbook()
        output_wb.remove(output_wb.active)  # 删除默认工作表
        
        total_sheets = 0
        
        for idx, file_path in enumerate(file_paths):
            send_progress(
                int((idx / len(file_paths)) * 90),
                f"正在处理文件 {idx + 1}/{len(file_paths)}: {os.path.basename(file_path)}"
            )
            
            # 读取源文件
            source_wb = openpyxl.load_workbook(file_path, data_only=True)
            file_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # 复制所有工作表
            for sheet in source_wb.worksheets:
                # 生成新工作表名称
                new_sheet_name = f"{file_name}_{sheet.title}"
                
                # 确保名称唯一且不超过 31 字符
                if len(new_sheet_name) > 31:
                    new_sheet_name = new_sheet_name[:31]
                
                # 检查名称是否已存在
                counter = 1
                original_name = new_sheet_name
                while new_sheet_name in output_wb.sheetnames:
                    suffix = f"_{counter}"
                    max_len = 31 - len(suffix)
                    new_sheet_name = original_name[:max_len] + suffix
                    counter += 1
                
                # 创建新工作表
                new_sheet = output_wb.create_sheet(title=new_sheet_name)
                
                # 复制数据
                for row in sheet.iter_rows():
                    for cell in row:
                        new_cell = new_sheet[cell.coordinate]
                        new_cell.value = cell.value
                        
                        # 复制样式（可选）
                        if cell.has_style:
                            new_cell.font = cell.font.copy()
                            new_cell.border = cell.border.copy()
                            new_cell.fill = cell.fill.copy()
                            new_cell.number_format = cell.number_format
                            new_cell.protection = cell.protection.copy()
                            new_cell.alignment = cell.alignment.copy()
                
                total_sheets += 1
            
            source_wb.close()
        
        send_progress(95, "正在保存文件...")
        
        # 保存输出文件
        output_wb.save(output_path)
        output_wb.close()
        
        send_progress(100, "合并完成")
        
        return {
            "type": "result",
            "status": "success",
            "message": f"成功合并 {len(file_paths)} 个文件的 {total_sheets} 个工作表",
            "data": {
                "total_files": len(file_paths),
                "total_sheets": total_sheets,
                "output_path": output_path,
                "mode": "separate"
            }
        }
    
    def merge_csv_files(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        合并多个 CSV 文件
        
        Args:
            params: {
                'file_paths': List[str],  # 要合并的文件路径列表
                'output_path': str,  # 输出文件路径
                'has_header': bool,  # 是否包含表头（默认 True）
                'encoding': str,  # 编码格式（默认 'utf-8'）
                'delimiter': str  # 分隔符（默认 ','）
            }
        
        Returns:
            dict: 操作结果
        """
        # 验证参数
        file_paths = params.get('file_paths')
        output_path = params.get('output_path')
        has_header = params.get('has_header', True)
        encoding = params.get('encoding', 'utf-8')
        delimiter = params.get('delimiter', ',')
        
        if not file_paths:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_FILE_PATHS",
                "message": "缺少文件路径列表"
            }
        
        if not output_path:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_OUTPUT_PATH",
                "message": "缺少输出文件路径"
            }
        
        # 检查文件是否存在
        for file_path in file_paths:
            if not os.path.exists(file_path):
                return {
                    "type": "result",
                    "status": "error",
                    "error_code": "FILE_NOT_FOUND",
                    "message": f"文件不存在: {file_path}"
                }
        
        try:
            all_data = []
            total_rows = 0
            
            for idx, file_path in enumerate(file_paths):
                send_progress(
                    int((idx / len(file_paths)) * 80),
                    f"正在读取文件 {idx + 1}/{len(file_paths)}: {os.path.basename(file_path)}"
                )
                
                # 读取 CSV
                df = pd.read_csv(
                    file_path,
                    header=0 if has_header else None,
                    encoding=encoding,
                    delimiter=delimiter
                )
                
                all_data.append(df)
                total_rows += len(df)
            
            send_progress(85, "正在合并数据...")
            
            # 合并所有数据
            merged_df = pd.concat(all_data, ignore_index=True)
            
            send_progress(90, "正在保存文件...")
            
            # 保存到 CSV
            merged_df.to_csv(
                output_path,
                index=False,
                encoding=encoding,
                sep=delimiter
            )
            
            send_progress(100, "合并完成")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功合并 {len(file_paths)} 个 CSV 文件",
                "data": {
                    "total_files": len(file_paths),
                    "total_rows": total_rows,
                    "output_path": output_path
                }
            }
        
        except Exception as e:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MERGE_ERROR",
                "message": f"合并 CSV 文件时发生错误: {str(e)}"
            }
    
    def split_excel_file(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        拆分 Excel 文件
        
        Args:
            params: {
                'file_path': str,  # 要拆分的文件路径
                'output_dir': str,  # 输出目录
                'rows_per_file': int,  # 每个文件的行数
                'has_header': bool,  # 是否包含表头（默认 True）
                'sheet_name': str  # 要拆分的工作表名称（可选，默认第一个）
            }
        
        Returns:
            dict: 操作结果
        """
        # 验证参数
        file_path = params.get('file_path')
        output_dir = params.get('output_dir')
        rows_per_file = params.get('rows_per_file')
        has_header = params.get('has_header', True)
        sheet_name = params.get('sheet_name')
        
        if not file_path:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_FILE_PATH",
                "message": "缺少文件路径"
            }
        
        if not output_dir:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_OUTPUT_DIR",
                "message": "缺少输出目录"
            }
        
        if not rows_per_file or rows_per_file <= 0:
            return {
                "type": "result",
                "status": "error",
                "error_code": "INVALID_ROWS_PER_FILE",
                "message": "每个文件的行数必须大于 0"
            }
        
        if not os.path.exists(file_path):
            return {
                "type": "result",
                "status": "error",
                "error_code": "FILE_NOT_FOUND",
                "message": f"文件不存在: {file_path}"
            }
        
        try:
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            send_progress(10, "正在读取文件...")
            
            # 读取 Excel
            df = pd.read_excel(
                file_path,
                sheet_name=sheet_name if sheet_name else 0,
                header=0 if has_header else None
            )
            
            total_rows = len(df)
            num_files = (total_rows + rows_per_file - 1) // rows_per_file
            
            send_progress(20, f"将拆分为 {num_files} 个文件...")
            
            # 获取文件名（不含扩展名）
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            
            output_files = []
            
            for i in range(num_files):
                start_row = i * rows_per_file
                end_row = min((i + 1) * rows_per_file, total_rows)
                
                send_progress(
                    20 + int((i / num_files) * 70),
                    f"正在生成文件 {i + 1}/{num_files}..."
                )
                
                # 提取数据
                chunk_df = df.iloc[start_row:end_row]
                
                # 生成输出文件名
                output_file = os.path.join(output_dir, f"{base_name}_part{i + 1}.xlsx")
                
                # 保存文件
                chunk_df.to_excel(output_file, index=False)
                
                output_files.append({
                    "file_name": os.path.basename(output_file),
                    "path": output_file,
                    "rows": len(chunk_df)
                })
            
            send_progress(100, "拆分完成")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功拆分为 {num_files} 个文件",
                "data": {
                    "total_rows": total_rows,
                    "rows_per_file": rows_per_file,
                    "num_files": num_files,
                    "output_dir": output_dir,
                    "output_files": output_files
                }
            }
        
        except Exception as e:
            return {
                "type": "result",
                "status": "error",
                "error_code": "SPLIT_ERROR",
                "message": f"拆分文件时发生错误: {str(e)}"
            }
    
    def split_csv_file(self, params: Dict[str, Any]) -> Dict[str, Any]:
        """
        拆分 CSV 文件
        
        Args:
            params: {
                'file_path': str,  # 要拆分的文件路径
                'output_dir': str,  # 输出目录
                'rows_per_file': int,  # 每个文件的行数
                'has_header': bool,  # 是否包含表头（默认 True）
                'encoding': str,  # 编码格式（默认 'utf-8'）
                'delimiter': str  # 分隔符（默认 ','）
            }
        
        Returns:
            dict: 操作结果
        """
        # 验证参数
        file_path = params.get('file_path')
        output_dir = params.get('output_dir')
        rows_per_file = params.get('rows_per_file')
        has_header = params.get('has_header', True)
        encoding = params.get('encoding', 'utf-8')
        delimiter = params.get('delimiter', ',')
        
        if not file_path:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_FILE_PATH",
                "message": "缺少文件路径"
            }
        
        if not output_dir:
            return {
                "type": "result",
                "status": "error",
                "error_code": "MISSING_OUTPUT_DIR",
                "message": "缺少输出目录"
            }
        
        if not rows_per_file or rows_per_file <= 0:
            return {
                "type": "result",
                "status": "error",
                "error_code": "INVALID_ROWS_PER_FILE",
                "message": "每个文件的行数必须大于 0"
            }
        
        if not os.path.exists(file_path):
            return {
                "type": "result",
                "status": "error",
                "error_code": "FILE_NOT_FOUND",
                "message": f"文件不存在: {file_path}"
            }
        
        try:
            # 创建输出目录
            os.makedirs(output_dir, exist_ok=True)
            
            send_progress(10, "正在读取文件...")
            
            # 读取 CSV
            df = pd.read_csv(
                file_path,
                header=0 if has_header else None,
                encoding=encoding,
                delimiter=delimiter
            )
            
            total_rows = len(df)
            num_files = (total_rows + rows_per_file - 1) // rows_per_file
            
            send_progress(20, f"将拆分为 {num_files} 个文件...")
            
            # 获取文件名（不含扩展名）
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            
            output_files = []
            
            for i in range(num_files):
                start_row = i * rows_per_file
                end_row = min((i + 1) * rows_per_file, total_rows)
                
                send_progress(
                    20 + int((i / num_files) * 70),
                    f"正在生成文件 {i + 1}/{num_files}..."
                )
                
                # 提取数据
                chunk_df = df.iloc[start_row:end_row]
                
                # 生成输出文件名
                output_file = os.path.join(output_dir, f"{base_name}_part{i + 1}.csv")
                
                # 保存文件
                chunk_df.to_csv(
                    output_file,
                    index=False,
                    encoding=encoding,
                    sep=delimiter
                )
                
                output_files.append({
                    "file_name": os.path.basename(output_file),
                    "path": output_file,
                    "rows": len(chunk_df)
                })
            
            send_progress(100, "拆分完成")
            
            return {
                "type": "result",
                "status": "success",
                "message": f"成功拆分为 {num_files} 个文件",
                "data": {
                    "total_rows": total_rows,
                    "rows_per_file": rows_per_file,
                    "num_files": num_files,
                    "output_dir": output_dir,
                    "output_files": output_files
                }
            }
        
        except Exception as e:
            return {
                "type": "result",
                "status": "error",
                "error_code": "SPLIT_ERROR",
                "message": f"拆分 CSV 文件时发生错误: {str(e)}"
            }

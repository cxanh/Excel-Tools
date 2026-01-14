import openpyxl
import os

# 测试文件路径
file_path = "../test/课程信息表1.xlsx"

print(f"检查文件: {file_path}")
print(f"文件存在: {os.path.exists(file_path)}")

# 读取Excel文件
try:
    wb = openpyxl.load_workbook(file_path)
    print(f"\n工作表列表: {wb.sheetnames}")
    
    # 遍历所有工作表
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n=== 工作表: {sheet_name} ===")
        print(f"最大行: {ws.max_row}, 最大列: {ws.max_column}")
        
        # 读取前5行数据
        print("\n前5行数据:")
        for row in ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True):
            print(row)
    
    wb.close()
except Exception as e:
    print(f"读取文件时出错: {str(e)}")

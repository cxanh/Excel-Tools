#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
测试 remove_duplicate_rows 函数的脚本
"""

import sys
import os
sys.path.append(os.path.join(os.path.dirname(__file__), 'python-backend'))

from engine.content.processor import ContentProcessor
from openpyxl import Workbook


def test_remove_duplicate_rows():
    """测试删除重复行功能"""
    print("开始测试 remove_duplicate_rows 功能...")
    
    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    
    # 添加一些测试数据，包括重复行
    test_data = [
        ["姓名", "年龄", "城市"],      # 标题行
        ["张三", 25, "北京"],        # 数据行1
        ["李四", 30, "上海"],        # 数据行2
        ["张三", 25, "北京"],        # 重复数据行1 (与第一行数据相同)
        ["王五", 28, "广州"],        # 数据行3
        ["赵六", 35, "深圳"],        # 数据行4
        ["李四", 30, "上海"],        # 重复数据行2 (与第二行数据相同)
    ]
    
    # 将数据添加到工作表
    for row_data in test_data:
        ws.append(row_data)
    
    print(f"原始数据行数: {ws.max_row}")
    print("原始数据:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    # 创建 ContentProcessor 实例
    processor = ContentProcessor()
    
    # 调用 remove_duplicate_rows 方法
    params = {
        'worksheet': ws,
        'sheet_name': 'TestSheet'
    }
    
    result = processor.remove_duplicate_rows(params)
    
    print("\n处理结果:")
    print(result)
    
    print(f"\n处理后数据行数: {ws.max_row}")
    print("处理后数据:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    if result['status'] == 'success':
        print("✅ 删除重复行功能测试通过!")
        return True
    else:
        print("❌ 删除重复行功能测试失败!")
        return False


def test_remove_duplicate_rows_specific_columns():
    """测试基于特定列删除重复行的功能"""
    print("\n" + "="*50)
    print("开始测试基于特定列删除重复行功能...")
    
    # 创建一个新的工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet2"
    
    # 添加一些测试数据，在某些列上有重复
    test_data = [
        ["姓名", "年龄", "城市"],
        ["张三", 25, "北京"],      # 第1行
        ["李四", 30, "上海"],      # 第2行
        ["张三", 26, "北京"],      # 第3行 - 姓名和城市与第1行相同，但年龄不同
        ["王五", 28, "广州"],      # 第4行
        ["张三", 25, "北京"],      # 第5行 - 完全与第1行相同
    ]
    
    # 将数据添加到工作表
    for row_data in test_data:
        ws.append(row_data)
    
    print(f"原始数据行数: {ws.max_row}")
    print("原始数据:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    # 创建 ContentProcessor 实例
    processor = ContentProcessor()
    
    # 只基于姓名和城市列删除重复行
    params = {
        'worksheet': ws,
        'sheet_name': 'TestSheet2',
        'key_columns': [1, 3]  # 只考虑姓名和城市列（索引从1开始）
    }
    
    result = processor.remove_duplicate_rows(params)
    
    print("\n基于姓名和城市列处理结果:")
    print(result)
    
    print(f"\n处理后数据行数: {ws.max_row}")
    print("处理后数据:")
    for row in ws.iter_rows(values_only=True):
        print(row)
    
    if result['status'] == 'success':
        print("✅ 基于特定列删除重复行功能测试通过!")
        return True
    else:
        print("❌ 基于特定列删除重复行功能测试失败!")
        return False


if __name__ == "__main__":
    success1 = test_remove_duplicate_rows()
    success2 = test_remove_duplicate_rows_specific_columns()
    
    if success1 and success2:
        print("\n🎉 所有测试通过!")
    else:
        print("\n💥 有些测试未通过!")
        sys.exit(1)